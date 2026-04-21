from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, session, g, make_response, current_app, send_file,
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import re
import os
import mysql.connector
import io
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import Config
from mail_utils import (
    notificar_nueva_solicitud_permiso,
    notificar_resolucion_permiso,
    notificar_resolucion_vacaciones,
    notificar_nueva_solicitud_vacaciones,
    notificar_encargado_nueva_solicitud,
    notificar_gh_resolucion_por_jefe,
)
import tempfile

app = Flask(__name__)
app.config.from_object(Config)
app.permanent_session_lifetime = timedelta(minutes=app.config.get("SESSION_TIMEOUT_MINUTES", 30))


@app.before_request
def enforce_session_timeout():
    """Cierra sesión cuando se supera el tiempo de inactividad configurado."""
    timeout_minutes = int(app.config.get("SESSION_TIMEOUT_MINUTES", 30))
    if timeout_minutes <= 0:
        return None

    # Evitar bucles en login/logout y excluir recursos estáticos.
    if request.endpoint in {"login", "logout", "register", "static"}:
        return None

    if "user_id" not in session:
        return None

    now_ts = datetime.utcnow().timestamp()
    last_activity = session.get("last_activity_ts")

    if last_activity and (now_ts - float(last_activity) > timeout_minutes * 60):
        session.clear()
        if _is_api_request():
            return jsonify({"error": "Sesion expirada por inactividad"}), 401
        flash("Sesion expirada por inactividad. Inicia sesion de nuevo.", "warning")
        return redirect(url_for("login"))

    session["last_activity_ts"] = now_ts
    session.permanent = True
    return None


# ── DB helpers ────────────────────────────────────────────────

def get_db():
    return mysql.connector.connect(
        host=app.config["MYSQL_HOST"],
        port=app.config["MYSQL_PORT"],
        user=app.config["MYSQL_USER"],
        password=app.config["MYSQL_PASSWORD"],
        database=app.config["MYSQL_DATABASE"],
    )


def query(sql, params=None, one=False):
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, params or ())
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows[0] if one and rows else rows if not one else None


def execute(sql, params=None):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(sql, params or ())
    conn.commit()
    lid = cursor.lastrowid
    cursor.close()
    conn.close()
    return lid


# ── Auth helpers (carga desde BD: rol_permiso, rol_modulo) ───

# Fallback si no existen tablas o están vacías
_ROLE_PERMISSIONS_FALLBACK = {
    "ADMIN": "ALL", "COORD. GH": "ALL", "GH INFORMADA": "READ",
    "GESTOR DE CONTRATACION": "WRITE", "BIENESTAR SOCIAL": "READ",
    "GESTOR DE NOMINA": "WRITE", "GESTOR SST": "READ",
}


def _load_role_permissions_from_db():
    """Devuelve dict rol -> nivel (READ/WRITE/ALL) desde rol_permiso, o None si no existe tabla."""
    try:
        rows = query("SELECT rol_nombre, nivel FROM rol_permiso")
        if rows:
            return {r["rol_nombre"]: r["nivel"] for r in rows}
    except Exception:
        pass
    return None


def _load_role_modules_from_db():
    """Devuelve dict rol -> { modulo_key: bool } desde rol_modulo, o None si no existe tabla."""
    try:
        rows = query("SELECT rol_nombre, modulo_key, visible FROM rol_modulo")
        if not rows:
            return None
        by_rol = {}
        for r in rows:
            by_rol.setdefault(r["rol_nombre"], {})[r["modulo_key"]] = bool(r["visible"])
        return by_rol
    except Exception:
        pass
    return None


def get_role_permission(rol):
    """Nivel de permiso del rol (READ/WRITE/ALL). Desde BD o fallback."""
    if not hasattr(g, "_role_permissions"):
        g._role_permissions = _load_role_permissions_from_db() or _ROLE_PERMISSIONS_FALLBACK
    return g._role_permissions.get(rol, "READ")


def get_acciones_for_rol(rol):
    """Texto de acciones para usuario según permiso del rol. Se guarda en BD en usuario.acciones."""
    perm = get_role_permission(rol)
    return {"ALL": "TODOS LOS CAMBIOS", "WRITE": "AGREGAR Y MODIFICAR", "READ": "VISTA"}.get(perm, "VISTA")


def get_current_user():
    if "user_id" not in session:
        return None
    if not hasattr(g, "_user"):
        g._user = query(
            "SELECT * FROM usuario WHERE id_user = %s AND estado = 1",
            (session["user_id"],), one=True,
        )
    return g._user


def _is_api_request():
    """True si la petición espera JSON (rutas /api/ o Accept: application/json)."""
    return request.path.startswith("/api/") or "application/json" in request.accept_mimetypes


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if user is None:
            if _is_api_request():
                return jsonify({"error": "No autenticado"}), 401
            flash("Debes iniciar sesión", "error")
            return redirect(url_for("login"))
        # Obligatorio cambiar clave de una vez (flag en BD o entró con contraseña estándar)
        must_change = user.get("debe_cambiar_clave") or session.get("force_change_password")
        if must_change and request.endpoint not in ("cambiar_clave_obligatorio", "logout"):
            flash("Debe cambiar su contraseña para continuar.", "warning")
            return redirect(url_for("cambiar_clave_obligatorio"))
        return f(*args, **kwargs)
    return decorated


def role_required(*allowed):
    """allowed can be 'ALL', 'WRITE', 'READ' (minimum level)."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = get_current_user()
            if user is None:
                if _is_api_request():
                    return jsonify({"error": "No autenticado"}), 401
                return redirect(url_for("login"))
            perm = get_role_permission(user["rol"])
            levels = {"READ": 0, "WRITE": 1, "ALL": 2}
            min_level = max(levels.get(a, 0) for a in allowed)
            if levels.get(perm, 0) < min_level:
                if _is_api_request():
                    return jsonify({"error": "Sin permisos"}), 403
                flash("No tienes permisos para esta acción", "error")
                return redirect(url_for("home"))
            return f(*args, **kwargs)
        return decorated
    return decorator


def admin_only(f):
    """Solo el rol ADMIN (ej. tecnologia@colbeef.com) puede acceder. Coordinación y otros no."""
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if user is None:
            return redirect(url_for("login"))
        if (user.get("rol") or "").strip() != "ADMIN":
            flash("Solo el administrador del sistema puede acceder a esta sección.", "error")
            return redirect(url_for("home"))
        return f(*args, **kwargs)
    return decorated


# Qué módulos del sidebar/home puede ver cada rol.
# Claves usadas en base.html y home.html con {{ vm.* }}
_ROLE_MODULES = {
    "ADMIN": {
        "organizacion": True,   # Áreas, Deptos, Perfiles
        "personal":     True,   # Personal Activo/Inactivo
        "retiro":       True,   # Retiro de Personal
        "familia":      True,   # Hijos Activos/Inactivos
        "eventos":      True,   # Cumpleaños, Aniversario
        "eps":          True,   # EPS
        "fondos":       True,   # Fondo de Pensiones
        "reportes":     True,   # Total Hijos, Dashboard
        "dashboard":    True,   # Gráficas (siempre visible para ADMIN)
        "admin":        True,   # Home Setting, Catálogos, Usuarios (todo)
        "admin_usuarios": True,
        "permisos":     True,   # Solicitud permiso/licencia
    },
    "COORD. GH": {
        "organizacion": True,
        "personal":     True,
        "retiro":       True,
        "familia":      True,
        "eventos":      True,
        "eps":          True,
        "fondos":       True,
        "reportes":     True,
        "admin":        False,   # Sin Home Setting ni Catálogos
        "admin_usuarios": True,  # Solo ver usuarios e inactivar
        "permisos":     True,    # Coordinación aprueba/rechaza permisos
    },
    "GH INFORMADA": {
        "organizacion": True,
        "personal":     True,
        "retiro":       True,
        "familia":      True,
        "eventos":      True,
        "eps":          True,
        "fondos":       True,
        "reportes":     True,
        "admin":        False,
        "admin_usuarios": False,
        "permisos":     True,    # Solo consulta; no aprueba ni rechaza
    },
    "GESTOR DE CONTRATACION": {
        "organizacion": False,
        "personal":     True,
        "retiro":       True,
        "familia":      True,
        "eventos":      True,
        "eps":          True,
        "fondos":       True,
        "reportes":     True,
        "admin":        False,   # Sin Home Setting ni Catálogos
        "admin_usuarios": True,  # Ver, crear, editar roles, inactivar (no asignar ADMIN)
        "permisos":     True,
    },
    "BIENESTAR SOCIAL": {
        "organizacion": False,
        "personal":     True,
        "retiro":       False,
        "familia":      True,
        "eventos":      True,
        "eps":          False,
        "fondos":       False,
        "reportes":     True,   # solo Total Hijos
        "dashboard":    False,  # sin Dashboard
        "admin":        False,
        "permisos":     True,
    },
    "GESTOR DE NOMINA": {
        "organizacion": False,
        "personal":     True,
        "retiro":       True,
        "familia":      False,
        "eventos":      False,
        "eps":          True,
        "fondos":       True,
        "reportes":     True,   # solo Dashboard
        "total_hijos":  False,
        "admin":        False,
        "permisos":     True,
    },
    "GESTOR SST": {
        "organizacion": False,
        "personal":     True,   # solo Activo (restringido en ruta)
        "retiro":       False,
        "familia":      False,
        "eventos":      False,
        "eps":          False,
        "fondos":       False,
        "reportes":     True,   # solo Dashboard
        "total_hijos":  False,
        "admin":        False,
        "permisos":     False,  # GESTOR SST no ve Permisos (solo Incidencias SISO si aplica)
    },
    "EMPLEADO": {
        "organizacion": False,
        "personal":     False,
        "retiro":       False,
        "familia":      False,
        "eventos":      False,
        "eps":          False,
        "fondos":       False,
        "reportes":     False,
        "admin":        False,
        "permisos":     True,   # solo Solicitud de permiso (portal empleado)
    },
    "SISO": {
        "incidencias": True,   # Registro y análisis INATEL (incidentes, accidentes, enfermedades laborales)
        "incidencias_dashboard": True,
    },
}

# Contraseña inicial para empleados creados por Gestor (y para dar de alta en BD)
EMPLEADO_PASSWORD_DEFAULT = "Colbeef2026*"
# Contraseña estándar al restablecer desde admin; el usuario debe cambiarla al ingresar
PASSWORD_ESTANDAR = "Colbeef2026*"

# Módulo por defecto (sin rol): nada visible
_DEFAULT_MODULES = {k: False for k in [
    "organizacion", "personal", "personal_inactivo", "retiro",
    "familia", "eventos", "eps", "fondos",
    "reportes", "dashboard", "total_hijos", "admin", "admin_usuarios", "permisos",
    "incidencias", "incidencias_dashboard", "suite_principal",
    "locker",
]}


def _normalize_rol(rol):
    """Normaliza el rol para comparar: sin tildes, mayúsculas, espacios colapsados."""
    if not rol or not isinstance(rol, str):
        return ""
    s = " ".join(str(rol).upper().split())
    for old, new in [("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ñ", "N")]:
        s = s.replace(old, new)
    return s


def _rol_match(rol_from_db):
    """Devuelve la clave de _ROLE_MODULES que corresponde a rol_from_db (exacta o normalizada)."""
    if rol_from_db in _ROLE_MODULES:
        return rol_from_db
    norm = _normalize_rol(rol_from_db)
    for key in _ROLE_MODULES:
        if _normalize_rol(key) == norm:
            return key
    return rol_from_db


def _normalize_email(s):
    if s is None:
        return ""
    return " ".join(str(s).split()).strip().lower()


def _is_locker_user(user):
    """Locker: visible para gestor de contratación y correo de gerencia."""
    if not user:
        return False
    email = _normalize_email(user.get("email"))
    gestor_mail = _normalize_email(app.config.get("MAIL_GESTOR_CONTRATACION") or "gestor.contratacion@colbeef.com")
    extra_allowed = {
        "gerencia@colbeef",
        "gerencia@colbeef.com",
    }
    if gestor_mail and email == gestor_mail:
        return True
    if email in extra_allowed:
        return True
    return _rol_match(user.get("rol")) == "GESTOR DE CONTRATACION"


def _get_effective_modules(rol):
    """Calcula el mapa completo de módulos visibles para un rol.
    Usa _ROLE_MODULES como base; la BD solo puede sumar visibilidad (no quitar), así
    gestor.contratacion@colbeef.com (GESTOR DE CONTRATACION) siempre ve Permisos."""
    if not hasattr(g, "_role_modules_db"):
        g._role_modules_db = _load_role_modules_from_db()
    rol_key = _rol_match(rol)
    base = dict(_DEFAULT_MODULES)
    base.update(_ROLE_MODULES.get(rol_key, {}))
    if g._role_modules_db and rol in g._role_modules_db:
        for k, v in g._role_modules_db[rol].items():
            if v:
                base[k] = True
    vm = dict(_DEFAULT_MODULES)
    vm.update(base)

    # dashboard y total_hijos heredan de reportes si no están definidos
    if "dashboard" not in base:
        vm["dashboard"] = vm.get("reportes", False)
    if "total_hijos" not in base:
        vm["total_hijos"] = vm.get("reportes", False)
    if "personal_inactivo" not in base:
        vm["personal_inactivo"] = vm.get("personal", False)

    # Restricciones específicas por rol (igual que antes)
    if rol_key == "GESTOR DE NOMINA":
        vm["total_hijos"] = False
    if rol_key == "BIENESTAR SOCIAL":
        vm["dashboard"] = False
    if rol_key == "GESTOR SST":
        vm["personal_inactivo"] = False
        vm["permisos"] = False  # GESTOR SST no ve Permisos (prevalece sobre rol_modulo en BD)

    return vm


def module_required(module_key):
    """Restringe el acceso a rutas según visibilidad de módulo por rol."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = get_current_user()
            if user is None:
                return redirect(url_for("login"))
            # Incidencias: visible también si el usuario es Siso@colbeef.com (aunque tenga otro rol)
            email_lower = (user.get("email") or "").strip().lower()
            tiene_incidencias = (
                _get_effective_modules(user["rol"]).get(module_key, False)
                or (module_key in ("incidencias", "incidencias_dashboard") and email_lower == "siso@colbeef.com")
            )
            if not tiene_incidencias:
                flash("No tienes acceso a este módulo", "error")
                if (user.get("rol") or "").strip().upper() == "EMPLEADO":
                    return redirect(url_for("empleado_portal"))
                return redirect(url_for("home"))
            return f(*args, **kwargs)
        return decorated
    return decorator


@app.context_processor
def inject_user():
    user = get_current_user()
    can_write = False
    can_admin = False
    vm = dict(_DEFAULT_MODULES)
    show_permisos_menu = False
    if user:
        rol = user.get("rol") or ""
        perm = get_role_permission(rol)
        can_write = perm in ("WRITE", "ALL")
        can_admin = perm == "ALL"
        vm = _get_effective_modules(rol)
        # Acceso directo a la suite principal (todos los usuarios autenticados)
        vm["suite_principal"] = True
        vm["locker"] = _is_locker_user(user)
        # Usuario Siso@colbeef.com: siempre ve el módulo Incidencias (aunque su rol sea GESTOR SST u otro)
        if (user.get("email") or "").strip().lower() == "siso@colbeef.com":
            vm["incidencias"] = True
            vm["incidencias_dashboard"] = True
        # GESTOR SST no ve Permisos (ni por código ni por BD)
        if _rol_match(rol) == "GESTOR SST":
            vm["permisos"] = False
        # Mostrar Permisos si vm lo tiene O si el rol es Gestor de Contratación (por nombre)
        show_permisos_menu = vm.get("permisos") is True or (
            "GESTOR" in rol.upper() and "CONTRAT" in rol.upper()
        )
    return dict(
        current_user=user,
        can_write=can_write,
        can_admin=can_admin,
        vm=vm,
        show_permisos_menu=show_permisos_menu,
    )


# ── AUTH ROUTES ───────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if get_current_user():
        return redirect(url_for("home"))

    if request.method == "POST":
        login_value = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        if "@" in login_value:
            user = query(
                "SELECT * FROM usuario WHERE LOWER(email) = %s AND estado = 1",
                (login_value.lower(),), one=True,
            )
        else:
            # Empleados sin correo: pueden entrar con solo su cédula (contraseña Colbeef2026*)
            cedula = "".join(c for c in login_value if c.isdigit())
            if cedula:
                user = query(
                    "SELECT * FROM usuario WHERE (id_cedula = %s OR id_user = %s) AND estado = 1",
                    (cedula, "EMP-" + cedula), one=True,
                )
            else:
                user = None

        if user and user.get("password_hash") and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id_user"]
            session.permanent = True
            # Obligar a cambiar la clave de una vez: si tiene flag en BD o si entró con la estándar
            if user.get("debe_cambiar_clave") or (password.strip() == PASSWORD_ESTANDAR):
                if password.strip() == PASSWORD_ESTANDAR:
                    session["force_change_password"] = True
                return redirect(url_for("cambiar_clave_obligatorio"))
            if (user.get("rol") or "").strip().upper() == "EMPLEADO":
                return redirect(url_for("empleado_portal"))
            return redirect(url_for("home"))

        flash("Correo o contraseña incorrectos", "error")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    """Ruta deshabilitada: la creación de usuarios es exclusiva del ADMIN desde /users."""
    flash("El acceso al sistema debe ser solicitado al administrador.", "error")
    return redirect(url_for("login"))


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada", "info")
    return redirect(url_for("login"))


@app.route("/locker")
@login_required
def locker():
    """Redirige a la URL configurada; solo el usuario con correo MAIL_GESTOR_CONTRATACION."""
    user = get_current_user()
    if not _is_locker_user(user):
        flash("No tienes acceso a este enlace.", "error")
        return redirect(url_for("home"))
    url = (app.config.get("GESTOR_CONTRATACION_PORTAL_URL") or "").strip()
    if not url:
        flash("Locker no configurado.", "warning")
        return redirect(url_for("home"))
    return redirect(url)


@app.route("/cambiar-clave", methods=["GET", "POST"])
@login_required
def cambiar_clave_obligatorio():
    """Cambio de contraseña obligatorio: al entrar lo obliga de una vez (estándar o flag en BD)."""
    user = get_current_user()
    must_change = user.get("debe_cambiar_clave") or session.get("force_change_password")
    if not must_change:
        return redirect(url_for("empleado_portal" if (user.get("rol") or "").strip().upper() == "EMPLEADO" else "home"))

    if request.method == "POST":
        actual = (request.form.get("password_actual") or "").strip()
        nueva = (request.form.get("nueva_password") or "").strip()
        repetir = (request.form.get("repetir_password") or "").strip()
        if not actual:
            flash("Indique su contraseña actual.", "error")
            return redirect(url_for("cambiar_clave_obligatorio"))
        u = query("SELECT password_hash FROM usuario WHERE id_user = %s", (user["id_user"],), one=True)
        if not u or not u.get("password_hash") or not check_password_hash(u["password_hash"], actual):
            flash("La contraseña actual no es correcta.", "error")
            return redirect(url_for("cambiar_clave_obligatorio"))
        ok, msg = _validar_password(nueva)
        if not ok:
            flash(msg, "error")
            return redirect(url_for("cambiar_clave_obligatorio"))
        if nueva != repetir:
            flash("La nueva contraseña y la repetición no coinciden.", "error")
            return redirect(url_for("cambiar_clave_obligatorio"))
        try:
            execute(
                "UPDATE usuario SET password_hash = %s, debe_cambiar_clave = 0 WHERE id_user = %s",
                (generate_password_hash(nueva), user["id_user"]),
            )
        except Exception as e:
            if "debe_cambiar_clave" in str(e):
                execute(
                    "UPDATE usuario SET password_hash = %s WHERE id_user = %s",
                    (generate_password_hash(nueva), user["id_user"]),
                )
            else:
                raise
        if hasattr(g, "_user"):
            g._user["debe_cambiar_clave"] = 0
        session.pop("force_change_password", None)
        flash("Contraseña actualizada. Ya puede usar su clave personal.", "success")
        if (user.get("rol") or "").strip().upper() == "EMPLEADO":
            return redirect(url_for("empleado_portal"))
        return redirect(url_for("home"))

    return render_template(
        "cambiar_clave_obligatorio.html",
        active_page="Cambiar contraseña",
        is_empleado=(user.get("rol") or "").strip().upper() == "EMPLEADO",
    )


# ── FOTO DE PERFIL ───────────────────────────────────────────

AVATAR_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
AVATAR_MAX_SIZE = 2 * 1024 * 1024  # 2 MB

EMPLEADO_FOTO_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
EMPLEADO_FOTO_MAX_SIZE = 4 * 1024 * 1024  # 4 MB


def _guardar_foto_empleado(id_cedula, file_storage):
    """Guarda la foto subida para un empleado dentro de static/empleados/<cedula>.<ext>.

    - Devuelve la ruta relativa (ej. 'empleados/1005123456.jpg') si se guardó
      correctamente, o None si no se subió nada.
    - Lanza ValueError con un mensaje descriptivo si el archivo es inválido
      (extensión o tamaño).
    """
    if not file_storage or not getattr(file_storage, "filename", ""):
        return None
    nombre_seguro = secure_filename(file_storage.filename)
    ext = os.path.splitext(nombre_seguro)[1].lower()
    if ext not in EMPLEADO_FOTO_EXTENSIONS:
        raise ValueError("Formato de foto no permitido (use JPG, JPEG, PNG o WEBP).")
    file_storage.seek(0, 2)
    size = file_storage.tell()
    file_storage.seek(0)
    if size > EMPLEADO_FOTO_MAX_SIZE:
        raise ValueError("La foto no debe superar 4 MB.")
    static_folder = current_app.static_folder
    carpeta = os.path.join(static_folder, "empleados")
    os.makedirs(carpeta, exist_ok=True)
    # Borrar fotos previas del mismo empleado con otras extensiones
    for prev_ext in EMPLEADO_FOTO_EXTENSIONS:
        prev = os.path.join(carpeta, f"{id_cedula}{prev_ext}")
        if prev_ext != ext and os.path.exists(prev):
            try:
                os.remove(prev)
            except Exception:
                pass
    filename = f"{id_cedula}{ext}"
    filepath = os.path.join(carpeta, filename)
    file_storage.save(filepath)
    return f"empleados/{filename}"


def _actualizar_foto_empleado_db(id_cedula, ruta_rel):
    """Intenta actualizar empleado.foto; si la columna aún no existe, avisa."""
    if not ruta_rel:
        return True
    try:
        execute("UPDATE empleado SET foto = %s WHERE id_cedula = %s", (ruta_rel, id_cedula))
        return True
    except Exception as e:
        if "foto" in str(e).lower():
            flash(
                "La foto se guardó en el servidor, pero falta aplicar la migración "
                "database/migration_foto_empleado.sql para que se asocie al empleado.",
                "warning",
            )
            return False
        raise


@app.route("/perfil/foto", methods=["POST"])
@login_required
def perfil_subir_foto():
    """Sube la foto de perfil del usuario. Guarda en static/avatars/id_user.ext y actualiza usuario.foto_perfil."""
    user = get_current_user()
    if not user:
        flash("Debe iniciar sesión.", "error")
        return redirect(url_for("login"))
    id_user = user["id_user"]
    file = request.files.get("foto")
    if not file or not file.filename:
        flash("Seleccione una imagen.", "error")
        return redirect(request.referrer or url_for("home"))
    ext = os.path.splitext(secure_filename(file.filename))[1].lower()
    if ext not in AVATAR_EXTENSIONS:
        ext = ".jpg"
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > AVATAR_MAX_SIZE:
        flash("La imagen no debe superar 2 MB.", "error")
        return redirect(request.referrer or url_for("home"))
    static_folder = current_app.static_folder
    avatars_dir = os.path.join(static_folder, "avatars")
    os.makedirs(avatars_dir, exist_ok=True)
    filename = id_user + ext
    filepath = os.path.join(avatars_dir, filename)
    try:
        file.save(filepath)
    except Exception as e:
        flash("No se pudo guardar la imagen.", "error")
        return redirect(request.referrer or url_for("home"))
    ruta_db = "avatars/" + filename
    try:
        execute("UPDATE usuario SET foto_perfil = %s WHERE id_user = %s", (ruta_db, id_user))
    except Exception:
        flash("Guardado en servidor; falta ejecutar migración: database/migration_foto_perfil.sql", "warning")
        return redirect(request.referrer or url_for("home"))
    if hasattr(g, "_user"):
        g._user["foto_perfil"] = ruta_db
    flash("Foto de perfil actualizada.", "success")
    return redirect(request.referrer or url_for("home"))


# ── PORTAL EMPLEADO (registro, panel, mis solicitudes) ─────────

@app.route("/empleado/registro", methods=["GET", "POST"])
def empleado_registro():
    """Auto-registro deshabilitado: las cuentas solo las crea Gestión Humana
    (ADMIN o COORD. GH) desde el módulo Personal. Cualquier intento de entrar
    a esta URL se redirige al login con un mensaje informativo.
    """
    flash(
        "El auto-registro está deshabilitado. Contacte a Gestión Humana para que le generen su cuenta.",
        "info",
    )
    return redirect(url_for("login"))


@app.route("/empleado")
@login_required
def empleado_portal():
    """Panel del empleado: solo acceso a Solicitud de permiso y Mis solicitudes."""
    user = get_current_user()
    if (user.get("rol") or "").strip().upper() != "EMPLEADO":
        return redirect(url_for("home"))
    return render_template("empleado_portal.html", active_page="Portal Empleado")


@app.route("/empleado/mis-solicitudes")
@login_required
def empleado_mis_solicitudes():
    """Listado de solicitudes del empleado (solo las suyas)."""
    user = get_current_user()
    if (user.get("rol") or "").strip().upper() != "EMPLEADO":
        return redirect(url_for("home"))
    id_cedula = (user.get("id_cedula") or "").strip()
    if not id_cedula:
        flash("No tiene cédula vinculada. Contacte al administrador.", "error")
        return redirect(url_for("empleado_portal"))
    solicitudes = query(
        "SELECT id, tipo, fecha_desde, fecha_hasta, estado, fecha_solicitud, observaciones, resuelto_por, fecha_resolucion "
        "FROM solicitud_permiso WHERE id_cedula = %s ORDER BY fecha_solicitud DESC",
        (id_cedula,),
    )
    for s in solicitudes:
        if s.get("fecha_solicitud"):
            d = s["fecha_solicitud"]
            s["fecha_solicitud_str"] = d.strftime("%d/%m/%Y %H:%M") if hasattr(d, "strftime") else str(d)
        else:
            s["fecha_solicitud_str"] = ""
        if s.get("fecha_resolucion"):
            d = s["fecha_resolucion"]
            s["fecha_resolucion_str"] = d.strftime("%d/%m/%Y %H:%M") if hasattr(d, "strftime") else str(d)
        else:
            s["fecha_resolucion_str"] = ""
    return render_template(
        "empleado_mis_solicitudes.html",
        active_page="Mis solicitudes",
        solicitudes=solicitudes,
    )


@app.route("/empleado/cambiar-password", methods=["GET", "POST"])
@login_required
def empleado_cambiar_password():
    """El empleado puede cambiar su contraseña (p. ej. desde la inicial Colbeef2026*)."""
    user = get_current_user()
    if (user.get("rol") or "").strip().upper() != "EMPLEADO":
        return redirect(url_for("home"))
    if request.method == "POST":
        actual = request.form.get("password_actual", "")
        nueva = (request.form.get("nueva_password") or "").strip()
        repetir = (request.form.get("repetir_password") or "").strip()
        if not actual:
            flash("Indique su contraseña actual.", "error")
            return redirect(url_for("empleado_cambiar_password"))
        u = query("SELECT password_hash FROM usuario WHERE id_user = %s", (user["id_user"],), one=True)
        if not u or not u.get("password_hash") or not check_password_hash(u["password_hash"], actual):
            flash("Contraseña actual incorrecta.", "error")
            return redirect(url_for("empleado_cambiar_password"))
        if len(nueva) < 6:
            flash("La nueva contraseña debe tener al menos 6 caracteres.", "error")
            return redirect(url_for("empleado_cambiar_password"))
        if nueva != repetir:
            flash("La nueva contraseña y la repetición no coinciden.", "error")
            return redirect(url_for("empleado_cambiar_password"))
        execute(
            "UPDATE usuario SET password_hash = %s WHERE id_user = %s",
            (generate_password_hash(nueva), user["id_user"]),
        )
        flash("Contraseña actualizada. Use la nueva contraseña en su próximo inicio de sesión.", "success")
        return redirect(url_for("empleado_portal"))
    return render_template("empleado_cambiar_password.html", active_page="Cambiar contraseña")


# ── HOME Y DESCARGA PRESENTACIÓN ─────────────────────────────

@app.route("/descargar/presentacion")
@login_required
def descargar_presentacion():
    """Descarga la presentación PowerPoint del proyecto (presentacion_gestio_humana.pptx)."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "presentacion_gestio_humana.pptx")
    if not os.path.isfile(path):
        flash("El archivo de presentación no está disponible.", "info")
        return redirect(url_for("home"))
    return send_file(path, as_attachment=True, download_name="presentacion_gestio_humana.pptx", mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation")


@app.route("/")
@login_required
def home():
    if (get_current_user().get("rol") or "").strip().upper() == "EMPLEADO":
        return redirect(url_for("empleado_portal"))
    return render_template("home.html", active_page="Home")


# ── CUMPLEAÑOS ────────────────────────────────────────────────

def parse_fecha(fecha_str):
    """Parsea fecha desde string. Prioriza DD/MM/YYYY (Colombia) para que cumpleaños coincidan con la fecha real."""
    if not fecha_str:
        return None
    if isinstance(fecha_str, date):
        return fecha_str
    s = str(fecha_str).strip()
    # Orden: DD/MM primero (estándar Colombia/Latam), luego ISO, luego MM/DD (US)
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def _calendar_label_maps():
    """Mapas id -> etiqueta para tipo_documento, nivel_educativo, profesion (para mostrar en calendario)."""
    def _s(v):
        return str(v).strip() if v is not None else ""

    tipo_map = {}
    for row in query("SELECT id_tipo_documento, tipo_documento FROM tipo_documento"):
        tid, tlabel = _s(row.get("id_tipo_documento")), _s(row.get("tipo_documento"))
        if tid: tipo_map[tid] = row["tipo_documento"]
        if tlabel: tipo_map[tlabel] = row["tipo_documento"]
    nivel_map = {}
    for row in query("SELECT id_nivel, nivel FROM nivel_educativo"):
        nid, nlabel = _s(row.get("id_nivel")), _s(row.get("nivel"))
        if nid: nivel_map[nid] = row["nivel"]
        if nlabel: nivel_map[nlabel] = row["nivel"]
    prof_map = {}
    for row in query("SELECT id_profesion, profesion FROM profesion"):
        pid, plabel = _s(row.get("id_profesion")), _s(row.get("profesion"))
        if pid: prof_map[pid] = row["profesion"]
        if plabel: prof_map[plabel] = row["profesion"]
    return tipo_map, nivel_map, prof_map


def _normalize_celular(val):
    """Evita mostrar celular como 3145831927.0; devuelve string sin decimal."""
    if val is None:
        return ""
    # Decimal (MySQL) o float
    if hasattr(val, "__float__"):
        try:
            f = float(val)
            return str(int(f)) if f == int(f) else str(val)
        except (ValueError, TypeError):
            pass
    s = str(val).strip()
    if not s:
        return ""
    if s.endswith(".0"):
        try:
            return str(int(float(s)))
        except ValueError:
            pass
    return s


def _looks_like_id(val):
    """True si el valor parece un id (código) y no un nombre legible."""
    if not val or len(val) > 80:
        return False
    s = str(val).strip()
    return bool(re.match(r"^[a-z0-9]{6,20}$", s, re.I))


def enrich_calendar_row(row_dict, tipo_map, nivel_map, prof_map):
    """Pone en row_dict celular normalizado y etiquetas para tipo_documento, nivel_educativo, profesion."""
    if "celular" in row_dict:
        row_dict["celular"] = _normalize_celular(row_dict.get("celular"))
    if "telefono_contacto" in row_dict:
        row_dict["telefono_contacto"] = _normalize_celular(row_dict.get("telefono_contacto"))
    v = str(row_dict.get("tipo_documento") or "").strip()
    if v:
        row_dict["tipo_documento"] = tipo_map.get(v, v)
    v = str(row_dict.get("nivel_educativo") or "").strip()
    if v:
        row_dict["nivel_educativo"] = nivel_map.get(v, v)
    v = str(row_dict.get("profesion") or "").strip()
    if v:
        row_dict["profesion"] = prof_map.get(v, v)


def _resolve_calendar_ids_in_results(results):
    """Resuelve profesion/tipo_documento/nivel_educativo que sigan como id en results (fallback por si el mapa falló)."""
    ids_prof = set()
    ids_tipo = set()
    ids_nivel = set()
    for item in results:
        p = str(item.get("profesion") or "").strip()
        if p and _looks_like_id(p):
            ids_prof.add(p)
        t = str(item.get("tipo_documento") or "").strip()
        if t and _looks_like_id(t) and len(t) <= 20:
            ids_tipo.add(t)
        n = str(item.get("nivel_educativo") or "").strip()
        if n and _looks_like_id(n):
            ids_nivel.add(n)
    if ids_prof:
        rows = query(
            "SELECT id_profesion, profesion FROM profesion WHERE id_profesion IN (" + ",".join(["%s"] * len(ids_prof)) + ")",
            tuple(ids_prof),
        )
        prof_by_id = {str(r["id_profesion"]).strip(): r["profesion"] for r in rows}
        for item in results:
            p = str(item.get("profesion") or "").strip()
            if p in prof_by_id:
                item["profesion"] = prof_by_id[p]
    if ids_tipo:
        rows = query(
            "SELECT id_tipo_documento, tipo_documento FROM tipo_documento WHERE id_tipo_documento IN (" + ",".join(["%s"] * len(ids_tipo)) + ")",
            tuple(ids_tipo),
        )
        tipo_by_id = {str(r["id_tipo_documento"]).strip(): r["tipo_documento"] for r in rows}
        for item in results:
            t = str(item.get("tipo_documento") or "").strip()
            if t in tipo_by_id:
                item["tipo_documento"] = tipo_by_id[t]
    if ids_nivel:
        rows = query(
            "SELECT id_nivel, nivel FROM nivel_educativo WHERE id_nivel IN (" + ",".join(["%s"] * len(ids_nivel)) + ")",
            tuple(ids_nivel),
        )
        nivel_by_id = {str(r["id_nivel"]).strip(): r["nivel"] for r in rows}
        for item in results:
            n = str(item.get("nivel_educativo") or "").strip()
            if n in nivel_by_id:
                item["nivel_educativo"] = nivel_by_id[n]


def enrich_empleados(empleados):
    """Add calculated fields, related hijos count and related retirados count."""
    today = date.today()
    hijos_counts = query(
        "SELECT id_cedula, COUNT(*) AS cnt FROM hijo GROUP BY id_cedula"
    )
    hijos_map = {r["id_cedula"]: int(r["cnt"]) for r in hijos_counts}
    ret_counts = query(
        "SELECT id_cedula, COUNT(*) AS cnt FROM retirado GROUP BY id_cedula"
    )
    ret_map = {r["id_cedula"]: int(r["cnt"]) for r in ret_counts}

    perfil_rows = query("SELECT id_perfil, perfil_ocupacional FROM perfil_ocupacional")
    perfil_map = {str(r["id_perfil"]): r["perfil_ocupacional"] for r in perfil_rows}

    for emp in empleados:
        cid = emp["id_cedula"]
        emp["perfil_ocupacional_nombre"] = perfil_map.get(str(emp.get("id_perfil_ocupacional", "")), "")
        raw_bd = emp.get("fecha_nacimiento", "")
        bd = raw_bd if isinstance(raw_bd, date) else parse_fecha(raw_bd)
        raw_fi = emp.get("fecha_ingreso", "")
        fi = raw_fi if isinstance(raw_fi, date) else parse_fecha(raw_fi)
        emp["mes_cumple"] = bd.month if bd else ""
        emp["cumpleanos"] = ""
        if bd:
            try:
                d = date(today.year, bd.month, bd.day)
                emp["cumpleanos"] = f"{d.month}/{d.day}/{d.year}"
            except ValueError:
                pass
        emp["aniversario_laboral"] = ""
        emp["antiguedad"] = ""
        if fi:
            try:
                d = date(today.year, fi.month, fi.day)
                emp["aniversario_laboral"] = f"{d.month}/{d.day}/{d.year}"
            except ValueError:
                pass
            emp["antiguedad"] = today.year - fi.year
        emp["contador"] = 1
        h_cnt = hijos_map.get(cid, 0)
        emp["related_hijos"] = f"Related Hijos ({h_cnt})" if h_cnt else ""
        r_cnt = ret_map.get(cid, 0)
        emp["related_retirados"] = f"Related Retirados ({r_cnt})" if r_cnt else ""
    return empleados


INLINE_COLUMNS = [
    ("apellidos_nombre", "Apellidos y Nombres"),
    ("id_cedula", "Numero de Cedula"),
    ("lugar_expedicion", "Lugar Expedicion Documento de Ide..."),
    ("fecha_expedicion", "Fecha Expedicion"),
    ("departamento", "Departamento"),
    ("area", "Area"),
    ("perfil_ocupacional_nombre", "Perfil Ocupacional"),
    ("fecha_ingreso", "Fecha de Ingreso a la compañia"),
    ("sexo", "Sexo"),
    ("rh", "Rh"),
    ("direccion_residencia", "Direccion de Residencia"),
    ("barrio_residencia", "Barrio de Residencia"),
    ("ciudad_residencia", "Ciudad de Residencia"),
    ("telefono", "Telefono"),
    ("celular", "Celular"),
    ("direccion_email", "Correo Electronico"),
    ("eps", "EPS"),
    ("fondo_pensiones", "Fondo de Pensiones"),
    ("fecha_nacimiento", "Fecha de Nacimiento"),
    ("hijos", "¿Tiene Hijos?"),
    ("estado", "Estado"),
    ("tipo_documento", "Tipo Documento"),
    ("nivel_educativo", "Nivel Educativo"),
    ("profesion", "Profesion"),
    ("contacto_emergencia", "contactoEmergencia"),
    ("telefono_contacto", "TelefonoContacto"),
    ("parentezco", "parentezco"),
    ("related_hijos", "Related Hijos"),
    ("contador", "Contador"),
    ("mes_cumple", "Mes Cumple"),
    ("cumpleanos", "Cumpleaños"),
    ("aniversario_laboral", "Aniversario Laboral"),
    ("antiguedad", "Antigüedad"),
    ("related_retirados", "Related Retirados"),
]


def export_excel_response_generic(rows, columns, filename_prefix):
    """Generate a styled Excel (.xlsx) download with given columns list of (key, label)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2D9E3F", end_color="2D9E3F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    alt_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    headers = [label for _, label in columns]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r, row in enumerate(rows, 2):
        for c, (key, _) in enumerate(columns, 1):
            val = row.get(key, "")
            if val is None:
                val = ""
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = alt_fill

    for c, (key, label) in enumerate(columns, 1):
        max_len = len(label)
        for r in range(2, len(rows) + 2):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 45)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today_str = date.today().strftime("%Y-%m-%d")
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="{filename_prefix}.ViewData.{today_str}.xlsx"'
    )
    return resp


def export_excel_response(empleados, filename_prefix):
    """Shortcut for empleado exports using INLINE_COLUMNS."""
    return export_excel_response_generic(empleados, INLINE_COLUMNS, filename_prefix)


@app.route("/cumpleanos")
@login_required
@module_required("eventos")
def cumpleanos():
    return render_template("cumpleanos.html", active_page="Cumpleaños")


@app.route("/api/cumpleanos")
@login_required
def api_cumpleanos():
    year = request.args.get("year", date.today().year, type=int)
    month = request.args.get("month", date.today().month, type=int)

    rows = query(
        "SELECT id_cedula, apellidos_nombre, fecha_nacimiento, departamento, "
        "area, sexo, celular, direccion_email, estado, tipo_documento, "
        "nivel_educativo, profesion, contacto_emergencia, telefono_contacto, parentezco "
        "FROM empleado WHERE estado = 'ACTIVO'"
    )

    tipo_map, nivel_map, prof_map = _calendar_label_maps()
    results = []
    for r in rows:
        bd = parse_fecha(r["fecha_nacimiento"])
        if bd and bd.month == month:
            try:
                cumple_date = date(year, bd.month, bd.day)
            except ValueError:
                continue
            item = {
                "id_cedula": r["id_cedula"],
                "apellidos_nombre": r["apellidos_nombre"],
                "fecha_nacimiento": r["fecha_nacimiento"],
                "departamento": r["departamento"] or "",
                "area": r["area"] or "",
                "sexo": r["sexo"] or "",
                "celular": r["celular"] or "",
                "correo": r["direccion_email"] or "",
                "estado": r["estado"] or "",
                "tipo_documento": r["tipo_documento"] or "",
                "nivel_educativo": r["nivel_educativo"] or "",
                "profesion": r["profesion"] or "",
                "contacto_emergencia": r["contacto_emergencia"] or "",
                "telefono_contacto": r["telefono_contacto"] or "",
                "parentezco": r["parentezco"] or "",
                "dia": bd.day,
                "mes_cumple": bd.month,
                "cumpleanos": cumple_date.strftime("%d/%m/%Y"),
            }
            enrich_calendar_row(item, tipo_map, nivel_map, prof_map)
            results.append(item)

    _resolve_calendar_ids_in_results(results)
    return jsonify(results)


@app.route("/cumpleanos/tarjeta")
@login_required
@module_required("eventos")
def cumpleanos_tarjeta():
    """Genera la tarjeta de cumpleaños del mes (formato bienestar: Nombre - Área: Día)."""
    year = request.args.get("year", date.today().year, type=int)
    month = request.args.get("month", date.today().month, type=int)
    rows = query(
        "SELECT apellidos_nombre, fecha_nacimiento, departamento, area "
        "FROM empleado WHERE estado = 'ACTIVO'"
    )
    MESES = ("", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre")
    lista = []
    for r in rows:
        bd = parse_fecha(r["fecha_nacimiento"])
        if not bd or bd.month != month:
            continue
        try:
            _ = date(year, bd.month, bd.day)
        except ValueError:
            continue
        rol = (r.get("area") or "").strip() or (r.get("departamento") or "").strip() or "—"
        nombre = (r.get("apellidos_nombre") or "").strip()
        lista.append({"nombre": nombre, "rol": rol, "dia": bd.day})
    lista.sort(key=lambda x: (x["dia"], x["nombre"]))
    return render_template(
        "cumpleanos_tarjeta.html",
        active_page="Tarjeta Cumpleaños",
        year=year,
        month=month,
        mes_nombre=MESES[month] if 1 <= month <= 12 else "",
        lista=lista,
        meses=list(enumerate(MESES[1:], 1)),
    )


# ── ANIVERSARIO LABORAL ────────────────────────────────────────

@app.route("/aniversario")
@login_required
@module_required("eventos")
def aniversario():
    return render_template("aniversario.html", active_page="Aniversario Laboral")


# ── SOLICITUD DE PERMISO / LICENCIA ───────────────────────────

def _puede_aprobar_permisos():
    """Roles con visibilidad global: ADMIN y COORD. GH ven todas las solicitudes."""
    user = get_current_user()
    return user and user.get("rol") in ("ADMIN", "COORD. GH")


def _es_admin_o_coord(user=None):
    user = user or get_current_user()
    return user and user.get("rol") in ("ADMIN", "COORD. GH")


def _puede_resolver_solicitud(solicitud):
    """Un usuario puede resolver (aprobar/rechazar) una solicitud si:
    - Es ADMIN o COORD. GH (visibilidad global), o
    - Es el encargado asignado al empleado de esa solicitud (empleado.id_user_encargado == user.id_user).
    """
    user = get_current_user()
    if not user or not solicitud:
        return False
    if _es_admin_o_coord(user):
        return True
    try:
        emp = query(
            "SELECT id_user_encargado FROM empleado WHERE id_cedula = %s",
            (solicitud.get("id_cedula"),), one=True,
        )
    except Exception:
        return False
    if not emp:
        return False
    return (emp.get("id_user_encargado") or "") == (user.get("id_user") or "")


def _puede_ver_listado_solicitudes():
    """Puede ver el listado de solicitudes si es ADMIN/COORD. GH, o si es encargado
    de al menos 1 empleado (es decir, tiene solicitudes de sus asignados)."""
    user = get_current_user()
    if not user:
        return False
    if _es_admin_o_coord(user):
        return True
    try:
        row = query(
            "SELECT COUNT(*) AS c FROM empleado WHERE id_user_encargado = %s",
            (user.get("id_user"),), one=True,
        )
        return bool(row and row.get("c"))
    except Exception:
        return False


def _resolver_email_empleado(id_cedula):
    """Devuelve el mejor correo disponible para el empleado, en este orden:
    1. empleado.direccion_email (si está registrado en la ficha).
    2. usuario.email del usuario vinculado a esa cédula (si no es un placeholder
       tipo "<cedula>@empleado.colbeef.local").
    Devuelve None si no se encuentra un correo válido.
    Útil para que la notificación de resolución (aprobación/rechazo) llegue al
    correo corporativo del empleado aunque no esté cargado en la ficha.
    """
    if not id_cedula:
        return None
    try:
        emp = query(
            "SELECT direccion_email FROM empleado WHERE id_cedula = %s",
            (id_cedula,), one=True,
        )
    except Exception:
        emp = None
    em = ((emp or {}).get("direccion_email") or "").strip()
    if em and "@" in em and not em.lower().endswith("@empleado.colbeef.local"):
        return em
    try:
        u = query(
            "SELECT email FROM usuario WHERE id_cedula = %s AND COALESCE(estado, 1) = 1 "
            "ORDER BY (rol = 'EMPLEADO') ASC, id_user LIMIT 1",
            (id_cedula,), one=True,
        )
    except Exception:
        u = None
    em2 = ((u or {}).get("email") or "").strip()
    if em2 and "@" in em2 and not em2.lower().endswith("@empleado.colbeef.local"):
        return em2
    return em or em2 or None


def _obtener_encargado_de(id_cedula):
    """Devuelve (id_user, nombre, email) del encargado asignado al empleado.
    None si el empleado no tiene encargado asignado o si la columna aún no existe.
    """
    if not id_cedula:
        return None
    try:
        emp = query("SELECT id_user_encargado FROM empleado WHERE id_cedula = %s", (id_cedula,), one=True)
    except Exception:
        return None
    if not emp or not emp.get("id_user_encargado"):
        return None
    u = query(
        "SELECT id_user, nombre, email FROM usuario WHERE id_user = %s AND COALESCE(estado, 1) = 1",
        (emp["id_user_encargado"],), one=True,
    )
    return u


def _sql_filtro_encargado(alias="e"):
    """Devuelve (where_fragment | None, params list) para filtrar solicitudes por encargado.
    Si el usuario es ADMIN/COORD. GH, no aplica filtro (devuelve None).
    En cualquier otro caso restringe a los empleados donde id_user_encargado == user.id_user.
    """
    user = get_current_user()
    if _es_admin_o_coord(user):
        return (None, [])
    return (f"{alias}.id_user_encargado = %s", [user.get("id_user") if user else None])


def _permisos_query(filtro_estado=None, buscar=None, area=None, tipo=None, orden=None, aplicar_filtro_encargado=True):
    """Construye SQL y params para listado de solicitudes (coordinadora).
    Si aplicar_filtro_encargado=True y el usuario no es ADMIN/COORD. GH,
    se restringe a las solicitudes de empleados cuyo encargado es él mismo.
    """
    sql = (
        "SELECT p.*, e.apellidos_nombre, e.direccion_email, "
        "COALESCE(p.area, e.area) AS area, e.id_user_encargado "
        "FROM solicitud_permiso p JOIN empleado e ON p.id_cedula = e.id_cedula "
    )
    params = []
    where = []
    if filtro_estado:
        where.append("p.estado = %s")
        params.append(filtro_estado)
    if buscar:
        where.append("(e.apellidos_nombre LIKE %s OR p.id_cedula LIKE %s)")
        params.extend(["%{}%".format(buscar.replace("%", "\\%")), "%{}%".format(buscar.replace("%", "\\%"))])
    if area:
        where.append("COALESCE(p.area, e.area) = %s")
        params.append(area)
    if tipo:
        where.append("p.tipo = %s")
        params.append(tipo)
    if aplicar_filtro_encargado:
        enc_where, enc_params = _sql_filtro_encargado("e")
        if enc_where:
            where.append(enc_where)
            params.extend(enc_params)
    if where:
        sql += " WHERE " + " AND ".join(where)
    order = (orden or "").strip().lower()
    if order == "fecha_desde":
        sql += " ORDER BY p.fecha_desde ASC, p.fecha_solicitud DESC"
    elif order == "nombre":
        sql += " ORDER BY e.apellidos_nombre, p.fecha_solicitud DESC"
    else:
        sql += " ORDER BY (p.estado = 'PENDIENTE') DESC, p.fecha_desde ASC, p.fecha_solicitud DESC"
    return sql, tuple(params)


@app.route("/permisos")
@login_required
@module_required("permisos")
def permisos_index():
    """Listado de solicitudes de permiso. ADMIN y COORD. GH ven todas; otros roles
    (jefes/encargados) ven solo las de los empleados asignados a ellos.
    """
    if not _puede_ver_listado_solicitudes():
        flash("No tiene solicitudes asignadas para revisar.", "info")
        return redirect(url_for("home"))
    filtro_estado = request.args.get("estado", "").strip().upper()
    if filtro_estado not in ("PENDIENTE", "APROBADO", "RECHAZADO"):
        filtro_estado = None
    buscar = request.args.get("buscar", "").strip()
    area = request.args.get("area", "").strip() or None
    tipo = request.args.get("tipo", "").strip() or None
    orden = request.args.get("orden", "").strip() or None
    sql, params = _permisos_query(filtro_estado=filtro_estado, buscar=buscar, area=area, tipo=tipo, orden=orden)
    solicitudes = query(sql, params)
    cur_user = get_current_user() or {}
    es_admin_coord = _es_admin_o_coord(cur_user)
    for s in solicitudes:
        if s.get("fecha_solicitud"):
            d = s["fecha_solicitud"]
            s["fecha_solicitud_str"] = d.strftime("%d/%m/%Y %H:%M") if hasattr(d, "strftime") else str(d)
        else:
            s["fecha_solicitud_str"] = ""
        s["_puede_resolver"] = es_admin_coord or (
            (s.get("id_user_encargado") or "") == (cur_user.get("id_user") or "")
        )
        s["_es_proximo"] = False
        if s.get("estado") == "PENDIENTE" and s.get("fecha_desde"):
            fd = s["fecha_desde"]
            try:
                hoy = date.today()
                if hasattr(fd, "date"):
                    fd_date = fd.date()
                elif hasattr(fd, "strftime"):
                    fd_date = fd
                elif isinstance(fd, str):
                    fd_date = datetime.strptime(fd[:10], "%Y-%m-%d").date()
                else:
                    fd_date = None
                if fd_date is not None:
                    delta = (fd_date - hoy).days
                    s["_es_proximo"] = 0 <= delta <= 7
            except Exception:
                pass
    areas_distinct = query(
        "SELECT DISTINCT COALESCE(p.area, e.area) AS area FROM solicitud_permiso p "
        "JOIN empleado e ON p.id_cedula = e.id_cedula WHERE COALESCE(p.area, e.area) IS NOT NULL AND COALESCE(p.area, e.area) != '' ORDER BY 1"
    )
    tipos_distinct = query("SELECT DISTINCT tipo FROM solicitud_permiso WHERE tipo IS NOT NULL AND tipo != '' ORDER BY tipo")
    from urllib.parse import urlencode
    export_params = {}
    if filtro_estado:
        export_params["estado"] = filtro_estado
    if buscar:
        export_params["buscar"] = buscar
    if area:
        export_params["area"] = area
    if tipo:
        export_params["tipo"] = tipo
    if orden:
        export_params["orden"] = orden
    permisos_export_url_full = url_for("permisos_export") + ("?" + urlencode(export_params) if export_params else "")
    return render_template(
        "permisos_list.html",
        active_page="Solicitud de permiso",
        solicitudes=solicitudes,
        puede_aprobar=_puede_ver_listado_solicitudes(),
        filtro_estado=filtro_estado,
        filtro_buscar=buscar,
        filtro_area=area,
        filtro_tipo=tipo,
        filtro_orden=orden,
        areas_list=[r["area"] for r in areas_distinct],
        tipos_list=[r["tipo"] for r in tipos_distinct],
        permisos_export_url=url_for("permisos_export"),
        permisos_export_url_full=permisos_export_url_full,
    )


# Roles que pueden solicitar permiso "para sí mismos" (si tienen id_cedula en usuario)
# Cualquier rol con id_cedula vinculada puede solicitar permiso/vacaciones "para sí" (mismos campos bloqueados + validación en servidor)
_ROLES_SOLICITAR_PARA_SI = (
    "EMPLEADO", "BIENESTAR SOCIAL", "ADMIN", "COORD. GH",
    "GESTOR DE CONTRATACION", "GESTOR DE NOMINA", "GESTOR SST",
)


@app.route("/permisos/solicitar", methods=["GET", "POST"])
@login_required
@module_required("permisos")
def permiso_solicitar():
    """Formulario GH-FR-007: permiso o licencia (área, remunerado/no, hora inicio/fin)."""
    user = get_current_user()
    rol = (user.get("rol") or "").strip().upper()
    is_empleado = rol == "EMPLEADO"
    # Cualquier usuario con id_cedula (todos los roles) usa flujo "para sí" con campos bloqueados y validación en servidor
    id_cedula_empleado = (user.get("id_cedula") or "").strip() or None

    if request.method == "POST":
        id_cedula = (request.form.get("id_cedula") or "").strip()
        if id_cedula_empleado:
            id_cedula = id_cedula_empleado
            if request.form.get("id_cedula", "").strip() != id_cedula_empleado:
                flash("No puede enviar solicitudes a nombre de otro empleado.", "error")
                return redirect(url_for("permiso_solicitar"))
        tipo = (request.form.get("tipo") or "Permiso").strip()
        _tipos_permitidos = ("Permiso", "Licencia", "Médico", "Personal", "Capacitación", "Calamidad doméstica", "Otro")
        if tipo not in _tipos_permitidos:
            tipo = "Permiso"
        fecha_desde = request.form.get("fecha_desde")
        fecha_hasta = request.form.get("fecha_hasta")
        motivo = (request.form.get("motivo") or "").strip()
        area = (request.form.get("area") or "").strip() or None
        pr = request.form.get("permiso_remunerado")
        permiso_remunerado = int(pr) if pr in ("0", "1") else None
        permiso_no_remunerado = 0 if permiso_remunerado == 1 else (1 if permiso_remunerado == 0 else None)
        hora_inicio = (request.form.get("hora_inicio") or "").strip() or None
        hora_fin = (request.form.get("hora_fin") or "").strip() or None
        if not id_cedula or not fecha_desde or not fecha_hasta:
            flash("Complete empleado, fecha desde y fecha hasta.", "error")
            return redirect(url_for("permiso_solicitar"))
        if permiso_remunerado is None:
            flash("Indique si el permiso es remunerado o no.", "error")
            return redirect(url_for("permiso_solicitar"))
        try:
            d_desde = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
            d_hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
            if d_desde > d_hasta:
                flash("La fecha desde no puede ser mayor que la fecha hasta.", "error")
                return redirect(url_for("permiso_solicitar"))
        except ValueError:
            flash("Fechas inválidas.", "error")
            return redirect(url_for("permiso_solicitar"))
        if permiso_remunerado == 0:
            evidencia_file = request.files.get("evidencia")
            if not evidencia_file or not evidencia_file.filename:
                flash("Para permiso no remunerado debe adjuntar evidencia (PDF o imagen).", "error")
                return redirect(url_for("permiso_solicitar"))
        emp = query("SELECT id_cedula, apellidos_nombre, direccion_email, area FROM empleado WHERE id_cedula = %s AND estado = 'ACTIVO'", (id_cedula,), one=True)
        if not emp:
            flash("No se encontró un empleado activo con esa cédula.", "error")
            return redirect(url_for("permiso_solicitar"))
        # Solicitud "para sí": no confiar en el formulario para cédula ni área (vienen de la ficha)
        if id_cedula_empleado:
            area = emp.get("area")
        elif not area and emp.get("area"):
            area = emp["area"]

        evidencia_ruta = None
        if permiso_remunerado == 0:
            evidencia_file = request.files.get("evidencia")
            if evidencia_file and evidencia_file.filename:
                ext = os.path.splitext(secure_filename(evidencia_file.filename))[1].lower()
                if ext not in (".pdf", ".jpg", ".jpeg", ".png"):
                    flash("La evidencia debe ser PDF o imagen (JPG, PNG).", "error")
                    return redirect(url_for("permiso_solicitar"))
                evidencia_file.seek(0, 2)
                size = evidencia_file.tell()
                evidencia_file.seek(0)
                if size > 5 * 1024 * 1024:
                    flash("La evidencia no debe superar 5 MB.", "error")
                    return redirect(url_for("permiso_solicitar"))
                upload_dir = os.path.join(current_app.instance_path, "uploads", "permisos")
                os.makedirs(upload_dir, exist_ok=True)
                nombre_safe = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{id_cedula}{ext}"
                evidencia_ruta = os.path.join("permisos", nombre_safe)
                evidencia_full_path = os.path.join(upload_dir, nombre_safe)
                evidencia_file.save(evidencia_full_path)

        try:
            execute(
                "INSERT INTO solicitud_permiso (id_cedula, tipo, fecha_desde, fecha_hasta, motivo, solicitante_email, area, permiso_remunerado, permiso_no_remunerado, hora_inicio, hora_fin, evidencia) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (id_cedula, tipo, fecha_desde, fecha_hasta, motivo, get_current_user() and get_current_user().get("email"), area, permiso_remunerado, permiso_no_remunerado, hora_inicio, hora_fin, evidencia_ruta),
            )
        except Exception as e:
            if "Unknown column" in str(e):
                execute(
                    "INSERT INTO solicitud_permiso (id_cedula, tipo, fecha_desde, fecha_hasta, motivo, solicitante_email, area, permiso_remunerado, permiso_no_remunerado, hora_inicio, hora_fin) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (id_cedula, tipo, fecha_desde, fecha_hasta, motivo, get_current_user() and get_current_user().get("email"), area, permiso_remunerado, permiso_no_remunerado, hora_inicio, hora_fin),
                )
            else:
                raise
        row = query("SELECT * FROM solicitud_permiso WHERE id_cedula = %s ORDER BY id DESC LIMIT 1", (id_cedula,), one=True)
        if not evidencia_ruta:
            evidencia_full_path = None
        else:
            fp = os.path.join(current_app.instance_path, "uploads", evidencia_ruta)
            evidencia_full_path = fp if os.path.isfile(fp) else None
        correos_ok = notificar_nueva_solicitud_permiso(app, row, emp["apellidos_nombre"], emp.get("direccion_email"), evidencia_path=evidencia_full_path)
        encargado = _obtener_encargado_de(id_cedula)
        if encargado and encargado.get("email"):
            try:
                notificar_encargado_nueva_solicitud(
                    app, row, emp["apellidos_nombre"],
                    encargado["email"], encargado.get("nombre"),
                    tipo="permiso", evidencia_path=evidencia_full_path,
                )
            except Exception:
                pass
        if correos_ok:
            flash("Solicitud registrada. Se envió correo a Coordinación GH" + (" y al encargado del empleado." if _obtener_encargado_de(id_cedula) else "."), "success")
        else:
            flash("Solicitud registrada. Revisar configuración de correo (MAIL_ENABLED, MAIL_PASSWORD) si no llegaron los avisos.", "info")
        if is_empleado:
            return redirect(url_for("empleado_mis_solicitudes"))
        return redirect(url_for("permisos_index"))
    now_fecha = datetime.now().strftime("%d-%m-%Y")
    # Solicitud para sí mismo: usuario con id_cedula (EMPLEADO o BIENESTAR SOCIAL u otro rol en la lista)
    if id_cedula_empleado:
        emp_actual = query(
            "SELECT id_cedula, apellidos_nombre, area FROM empleado WHERE id_cedula = %s AND estado = 'ACTIVO'",
            (id_cedula_empleado,), one=True,
        )
        if emp_actual:
            return render_template(
                "permiso_form.html",
                active_page="Solicitud de permiso",
                empleados=None,
                is_empleado=is_empleado,
                empleado_actual=emp_actual,
                now_fecha=now_fecha,
                nombreCompleto_valor=emp_actual.get("apellidos_nombre"),
                cedula_valor=emp_actual.get("id_cedula"),
            )
    empleados = query("SELECT id_cedula, apellidos_nombre, area FROM empleado WHERE estado = 'ACTIVO' ORDER BY apellidos_nombre")
    return render_template(
        "permiso_form.html",
        active_page="Solicitud de permiso",
        empleados=empleados,
        is_empleado=False,
        empleado_actual=None,
        now_fecha=now_fecha,
        nombreCompleto_valor=None,
        cedula_valor=None,
    )


# ── SOLICITUD DE VACACIONES ───────────────────────────────────

@app.route("/vacaciones/solicitar", methods=["GET", "POST"])
@login_required
@module_required("permisos")
def vacaciones_solicitar():
    """Formulario Solicitud de vacaciones (Gestión Humana - Colbeef)."""
    user = get_current_user()
    rol = (user.get("rol") or "").strip().upper()
    is_empleado = rol == "EMPLEADO"
    # Cualquier usuario con id_cedula (todos los roles) usa flujo "para sí" con validación en servidor
    id_cedula_empleado = (user.get("id_cedula") or "").strip() or None

    if request.method == "POST":
        id_cedula = (request.form.get("id_cedula") or "").strip()
        if id_cedula_empleado:
            id_cedula = id_cedula_empleado
            if request.form.get("id_cedula", "").strip() != id_cedula_empleado:
                flash("No puede enviar solicitudes a nombre de otro empleado.", "error")
                return redirect(url_for("vacaciones_solicitar"))
        fecha_solicitud = request.form.get("fecha_solicitud")
        periodo_causado = (request.form.get("periodo_causado") or "").strip() or None
        dias_tiempo = request.form.get("dias_en_tiempo")
        dias_comp = request.form.get("dias_compensados_dinero")
        dias_en_tiempo = int(dias_tiempo) if dias_tiempo not in (None, "") else None
        dias_compensados_dinero = int(dias_comp) if dias_comp not in (None, "") else None
        fecha_inicio = request.form.get("fecha_inicio")
        fecha_fin = request.form.get("fecha_fin")
        fecha_regreso = request.form.get("fecha_regreso")
        pago_anticipado = request.form.get("pago_anticipado")
        pago_anticipado = 1 if pago_anticipado == "1" else (0 if pago_anticipado == "0" else None)
        if not id_cedula or not fecha_solicitud or not fecha_inicio or not fecha_fin or not fecha_regreso:
            flash("Complete cédula, fecha de solicitud y fechas de inicio, fin y regreso.", "error")
            return redirect(url_for("vacaciones_solicitar"))
        emp = query("SELECT id_cedula, apellidos_nombre, direccion_email FROM empleado WHERE id_cedula = %s AND estado = 'ACTIVO'", (id_cedula,), one=True)
        if not emp:
            flash("No se encontró un empleado activo con esa cédula.", "error")
            return redirect(url_for("vacaciones_solicitar"))
        try:
            execute(
                "INSERT INTO solicitud_vacaciones (id_cedula, fecha_solicitud, periodo_causado, dias_en_tiempo, dias_compensados_dinero, fecha_inicio, fecha_fin, fecha_regreso, pago_anticipado, solicitante_email) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (id_cedula, fecha_solicitud, periodo_causado, dias_en_tiempo, dias_compensados_dinero, fecha_inicio, fecha_fin, fecha_regreso, pago_anticipado, get_current_user() and get_current_user().get("email")),
            )
        except Exception as e:
            if "doesn't exist" in str(e).lower():
                flash("Ejecute la migración: database/migration_solicitud_vacaciones.sql", "error")
                return redirect(url_for("vacaciones_solicitar"))
            raise
        try:
            new_row = query(
                "SELECT * FROM solicitud_vacaciones WHERE id_cedula = %s ORDER BY id DESC LIMIT 1",
                (id_cedula,), one=True,
            )
            emp_row = query(
                "SELECT apellidos_nombre FROM empleado WHERE id_cedula = %s", (id_cedula,), one=True,
            )
            emp_nombre = emp_row.get("apellidos_nombre") if emp_row else ""
            if new_row and emp_nombre:
                notificar_nueva_solicitud_vacaciones(app, new_row, emp_nombre)
                encargado = _obtener_encargado_de(id_cedula)
                if encargado and encargado.get("email"):
                    notificar_encargado_nueva_solicitud(
                        app, new_row, emp_nombre,
                        encargado["email"], encargado.get("nombre"),
                        tipo="vacaciones",
                    )
        except Exception:
            pass
        flash("Solicitud de vacaciones registrada correctamente.", "success")
        if is_empleado:
            return redirect(url_for("vacaciones_mis_solicitudes"))
        return redirect(url_for("vacaciones_index"))

    if id_cedula_empleado:
        emp_actual = query(
            "SELECT id_cedula, apellidos_nombre FROM empleado WHERE id_cedula = %s AND estado = 'ACTIVO'",
            (id_cedula_empleado,), one=True,
        )
        if emp_actual:
            return render_template(
                "vacaciones_form.html",
                active_page="Solicitud de vacaciones",
                empleados=None,
                is_empleado=is_empleado,
                empleado_actual=emp_actual,
                today=date.today().isoformat(),
            )
    empleados = query("SELECT id_cedula, apellidos_nombre FROM empleado WHERE estado = 'ACTIVO' ORDER BY apellidos_nombre")
    return render_template(
        "vacaciones_form.html",
        active_page="Solicitud de vacaciones",
        empleados=empleados,
        is_empleado=False,
        empleado_actual=None,
        today=date.today().isoformat(),
    )


@app.route("/vacaciones")
@login_required
@module_required("permisos")
def vacaciones_index():
    """Listado de solicitudes de vacaciones.
    ADMIN y COORD. GH ven TODAS. Otros roles (encargados) ven solo las de sus empleados asignados.
    """
    if not _puede_ver_listado_solicitudes():
        return redirect(url_for("vacaciones_solicitar"))
    sql = (
        "SELECT v.id, v.id_cedula, e.apellidos_nombre, e.direccion_email, e.id_user_encargado, "
        "v.fecha_solicitud, v.periodo_causado, v.dias_en_tiempo, v.dias_compensados_dinero, "
        "v.fecha_inicio, v.fecha_fin, v.fecha_regreso, v.pago_anticipado, "
        "v.estado, v.observaciones, v.resuelto_por, v.fecha_resolucion, "
        "u.nombre AS resuelto_por_nombre "
        "FROM solicitud_vacaciones v "
        "JOIN empleado e ON e.id_cedula = v.id_cedula "
        "LEFT JOIN usuario u ON u.id_user = v.resuelto_por "
    )
    params = []
    enc_where, enc_params = _sql_filtro_encargado("e")
    if enc_where:
        sql += " WHERE " + enc_where
        params.extend(enc_params)
    sql += " ORDER BY CASE v.estado WHEN 'PENDIENTE' THEN 0 ELSE 1 END, v.fecha_solicitud DESC, v.id DESC"
    rows = query(sql, tuple(params))
    cur_user = get_current_user() or {}
    es_admin_coord = _es_admin_o_coord(cur_user)
    for r in rows:
        r["_puede_resolver"] = es_admin_coord or (
            (r.get("id_user_encargado") or "") == (cur_user.get("id_user") or "")
        )
    return render_template(
        "vacaciones_list.html",
        active_page="Listado vacaciones",
        rows=rows,
        puede_aprobar=_puede_ver_listado_solicitudes(),
    )


@app.route("/vacaciones/<int:id>/aprobar", methods=["POST"])
@login_required
@module_required("permisos")
def vacaciones_aprobar(id):
    observaciones = (request.form.get("observaciones") or "").strip()
    solicitud = query("SELECT * FROM solicitud_vacaciones WHERE id = %s", (id,), one=True)
    if not solicitud:
        flash("Solicitud de vacaciones no encontrada.", "error")
        return redirect(url_for("vacaciones_index"))
    if not _puede_resolver_solicitud(solicitud):
        flash("No tiene permiso para aprobar esta solicitud (no es el encargado asignado).", "error")
        return redirect(url_for("vacaciones_index"))
    if solicitud["estado"] != "PENDIENTE":
        flash("La solicitud ya fue resuelta. Se actualiza la lista.", "info")
        return redirect(url_for("vacaciones_index"))
    execute(
        "UPDATE solicitud_vacaciones SET estado = 'APROBADO', observaciones = %s, "
        "resuelto_por = %s, fecha_resolucion = NOW() WHERE id = %s",
        (observaciones or None, get_current_user()["id_user"], id),
    )
    registrar_audit("Solicitud de vacaciones aprobada", "vacaciones", f"id={id} cédula={solicitud.get('id_cedula')}")
    emp = query(
        "SELECT apellidos_nombre, direccion_email FROM empleado WHERE id_cedula = %s",
        (solicitud["id_cedula"],), one=True,
    )
    email_empleado = _resolver_email_empleado(solicitud["id_cedula"])
    correo_ok = notificar_resolucion_vacaciones(
        app, solicitud,
        emp["apellidos_nombre"] if emp else "",
        email_empleado,
        aprobado=True, observaciones=observaciones,
    )
    _cur = get_current_user() or {}
    if not _es_admin_o_coord(_cur):
        try:
            notificar_gh_resolucion_por_jefe(
                app, solicitud, emp["apellidos_nombre"] if emp else "",
                tipo="vacaciones", aprobado=True,
                jefe_nombre=_cur.get("nombre"), observaciones=observaciones,
            )
        except Exception:
            pass
    if correo_ok:
        flash("Solicitud de vacaciones aprobada. Se notificó al empleado por correo.", "success")
    else:
        flash("Solicitud de vacaciones aprobada. No se pudo enviar el correo (revisar consola y MAIL_PASSWORD en .env).", "warning")
    return redirect(url_for("vacaciones_index"))


@app.route("/vacaciones/<int:id>/rechazar", methods=["POST"])
@login_required
@module_required("permisos")
def vacaciones_rechazar(id):
    observaciones = (request.form.get("observaciones") or "").strip()
    solicitud = query("SELECT * FROM solicitud_vacaciones WHERE id = %s", (id,), one=True)
    if not solicitud:
        flash("Solicitud de vacaciones no encontrada.", "error")
        return redirect(url_for("vacaciones_index"))
    if not _puede_resolver_solicitud(solicitud):
        flash("No tiene permiso para rechazar esta solicitud (no es el encargado asignado).", "error")
        return redirect(url_for("vacaciones_index"))
    if solicitud["estado"] != "PENDIENTE":
        flash("La solicitud ya fue resuelta. Se actualiza la lista.", "info")
        return redirect(url_for("vacaciones_index"))
    execute(
        "UPDATE solicitud_vacaciones SET estado = 'RECHAZADO', observaciones = %s, "
        "resuelto_por = %s, fecha_resolucion = NOW() WHERE id = %s",
        (observaciones or None, get_current_user()["id_user"], id),
    )
    registrar_audit("Solicitud de vacaciones rechazada", "vacaciones", f"id={id} cédula={solicitud.get('id_cedula')}")
    emp = query(
        "SELECT apellidos_nombre, direccion_email FROM empleado WHERE id_cedula = %s",
        (solicitud["id_cedula"],), one=True,
    )
    email_empleado = _resolver_email_empleado(solicitud["id_cedula"])
    correo_ok = notificar_resolucion_vacaciones(
        app, solicitud,
        emp["apellidos_nombre"] if emp else "",
        email_empleado,
        aprobado=False, observaciones=observaciones,
    )
    _cur = get_current_user() or {}
    if not _es_admin_o_coord(_cur):
        try:
            notificar_gh_resolucion_por_jefe(
                app, solicitud, emp["apellidos_nombre"] if emp else "",
                tipo="vacaciones", aprobado=False,
                jefe_nombre=_cur.get("nombre"), observaciones=observaciones,
            )
        except Exception:
            pass
    if correo_ok:
        flash("Solicitud de vacaciones rechazada. Se notificó al empleado por correo.", "success")
    else:
        flash("Solicitud de vacaciones rechazada. No se pudo enviar el correo (revisar consola y MAIL_PASSWORD en .env).", "warning")
    return redirect(url_for("vacaciones_index"))


@app.route("/vacaciones/mis-solicitudes")
@login_required
@module_required("permisos")
def vacaciones_mis_solicitudes():
    """Mis solicitudes de vacaciones (empleado o quien solicita para sí)."""
    user = get_current_user()
    id_cedula = (user.get("id_cedula") or "").strip()
    if not id_cedula:
        flash("No tiene cédula vinculada.", "error")
        return redirect(url_for("vacaciones_solicitar"))
    rows = query(
        "SELECT id, fecha_solicitud, periodo_causado, dias_en_tiempo, dias_compensados_dinero, fecha_inicio, fecha_fin, fecha_regreso, pago_anticipado, estado, fecha_resolucion "
        "FROM solicitud_vacaciones WHERE id_cedula = %s ORDER BY fecha_solicitud DESC",
        (id_cedula,),
    )
    return render_template(
        "vacaciones_mis_solicitudes.html",
        active_page="Mis solicitudes de vacaciones",
        rows=rows,
    )




@app.route("/permisos/<int:id>/aprobar", methods=["POST"])
@login_required
@module_required("permisos")
def permiso_aprobar(id):
    observaciones = (request.form.get("observaciones") or "").strip()
    solicitud = query("SELECT * FROM solicitud_permiso WHERE id = %s", (id,), one=True)
    if solicitud and not _puede_resolver_solicitud(solicitud):
        flash("No tiene permiso para aprobar esta solicitud (no es el encargado asignado).", "error")
        if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(ok=False, error="Sin permiso para aprobar esta solicitud."), 403
        return redirect(url_for("permisos_index"))
    if not solicitud:
        flash("Solicitud no encontrada. En este equipo la base de datos puede ser distinta.", "error")
        if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            try:
                p = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'PENDIENTE'", one=True)["c"]
                a = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'APROBADO'", one=True)["c"]
                r = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'RECHAZADO'", one=True)["c"]
            except Exception:
                p, a, r = 0, 0, 0
            return jsonify(ok=False, error="Solicitud no encontrada. Revise que la base de datos sea la correcta en este equipo.", pendientes=p, aprobadas=a, rechazadas=r), 404
        return redirect(url_for("permisos_index"))
    if solicitud["estado"] != "PENDIENTE":
        flash("La solicitud ya fue resuelta. Se actualiza la lista.", "info")
        if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            p = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'PENDIENTE'", one=True)["c"]
            a = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'APROBADO'", one=True)["c"]
            r = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'RECHAZADO'", one=True)["c"]
            return jsonify(ok=True, ya_resuelta=True, pendientes=p, aprobadas=a, rechazadas=r)
        return redirect(url_for("permisos_index"))
    execute(
        "UPDATE solicitud_permiso SET estado = 'APROBADO', observaciones = %s, resuelto_por = %s, fecha_resolucion = NOW() WHERE id = %s",
        (observaciones, get_current_user()["id_user"], id),
    )
    registrar_audit("Solicitud aprobada", "permisos", f"id={id} cédula={solicitud.get('id_cedula')}")
    emp = query("SELECT apellidos_nombre, direccion_email FROM empleado WHERE id_cedula = %s", (solicitud["id_cedula"],), one=True)
    attachments = []
    evidencia_ruta = (solicitud.get("evidencia") or "").strip()
    # Firma digital: misma imagen que en el correo. Prioridad: "firma digital cindy.png" en la raíz, luego SIGNATURE_IMAGE_PATH.
    root = getattr(current_app, "root_path", None) or os.path.dirname(os.path.abspath(__file__))
    firma_en_raiz = os.path.join(root, "firma digital cindy.png")
    firma_cfg = (current_app.config.get("SIGNATURE_IMAGE_PATH") or "").strip()
    if os.path.isfile(firma_en_raiz):
        firma_path_abs = firma_en_raiz
    elif firma_cfg and os.path.isfile(firma_cfg):
        firma_path_abs = firma_cfg if os.path.isabs(firma_cfg) else os.path.join(current_app.static_folder, firma_cfg)
    else:
        firma_path_abs = None
    # PDF informe: formulario GH-FR-007 con datos diligenciados y firma de Coordinación (siempre al aprobar)
    temp_informe = None
    try:
        from pdf_informe_permiso import generar_informe_permiso_pdf
        fd, temp_informe = tempfile.mkstemp(suffix=".pdf")
        os.close(fd)
        if generar_informe_permiso_pdf(solicitud, emp["apellidos_nombre"] if emp else "", temp_informe, firma_image_path=firma_path_abs):
            attachments.append(("Informe_permiso_GH-FR-007.pdf", temp_informe))
        elif temp_informe and os.path.isfile(temp_informe):
            try:
                os.unlink(temp_informe)
            except Exception:
                pass
            temp_informe = None
    except Exception:
        if temp_informe and os.path.isfile(temp_informe):
            try:
                os.unlink(temp_informe)
            except Exception:
                pass
        temp_informe = None
    # Opcional: si la evidencia subida era PDF, además se envía el PDF firmado (estampado)
    if evidencia_ruta and firma_path_abs and os.path.isfile(firma_path_abs):
        evidencia_full = os.path.join(current_app.instance_path, "uploads", evidencia_ruta)
        if os.path.isfile(evidencia_full) and evidencia_full.lower().endswith(".pdf"):
            temp_pdf = None
            try:
                from pdf_firma import firmar_pdf
                fd, temp_pdf = tempfile.mkstemp(suffix=".pdf")
                os.close(fd)
                if firmar_pdf(evidencia_full, firma_path_abs, temp_pdf, posicion="gh_celda_firma"):
                    attachments.append(("Formato_permiso_firmado.pdf", temp_pdf))
                elif temp_pdf and os.path.isfile(temp_pdf):
                    try:
                        os.unlink(temp_pdf)
                    except Exception:
                        pass
            except Exception:
                if temp_pdf and os.path.isfile(temp_pdf):
                    try:
                        os.unlink(temp_pdf)
                    except Exception:
                        pass
    email_empleado = _resolver_email_empleado(solicitud["id_cedula"])
    correo_ok = notificar_resolucion_permiso(app, solicitud, emp["apellidos_nombre"] if emp else "", email_empleado, aprobado=True, observaciones=observaciones, attachments=attachments if attachments else None)
    _cur = get_current_user() or {}
    if not _es_admin_o_coord(_cur):
        try:
            notificar_gh_resolucion_por_jefe(
                app, solicitud, emp["apellidos_nombre"] if emp else "",
                tipo="permiso", aprobado=True,
                jefe_nombre=_cur.get("nombre"), observaciones=observaciones,
            )
        except Exception:
            pass
    for _nom, path in attachments:
        if os.path.isfile(path):
            try:
                os.unlink(path)
            except Exception:
                pass
    try:
        execute(
            "UPDATE solicitud_permiso SET correo_resolucion_enviado = %s, correo_resolucion_at = IF(%s, NOW(), NULL) WHERE id = %s",
            (1 if correo_ok else 0, correo_ok, id),
        )
    except Exception:
        pass  # columnas no existen si no se ejecutó migration_correo_resolucion_validar.sql
    if correo_ok:
        if attachments:
            flash("Solicitud aprobada. Se notificó al empleado por correo con el informe en PDF (formato GH-FR-007 y firma digital) adjunto.", "success")
        else:
            flash("Solicitud aprobada. Se ha notificado al empleado por correo.", "success")
    else:
        flash("Solicitud aprobada. No se pudo enviar el correo al empleado (revisar consola y MAIL_PASSWORD en .env).", "warning")
    if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        p = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'PENDIENTE'", one=True)["c"]
        a = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'APROBADO'", one=True)["c"]
        r = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'RECHAZADO'", one=True)["c"]
        return jsonify(ok=True, pendientes=p, aprobadas=a, rechazadas=r)
    return redirect(url_for("permisos_index"))


@app.route("/permisos/<int:id>/rechazar", methods=["POST"])
@login_required
@module_required("permisos")
def permiso_rechazar(id):
    observaciones = (request.form.get("observaciones") or "").strip()
    solicitud = query("SELECT * FROM solicitud_permiso WHERE id = %s", (id,), one=True)
    if solicitud and not _puede_resolver_solicitud(solicitud):
        flash("No tiene permiso para rechazar esta solicitud (no es el encargado asignado).", "error")
        if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(ok=False, error="Sin permiso para rechazar esta solicitud."), 403
        return redirect(url_for("permisos_index"))
    if not solicitud:
        flash("Solicitud no encontrada. En este equipo la base de datos puede ser distinta.", "error")
        if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            try:
                p = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'PENDIENTE'", one=True)["c"]
                a = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'APROBADO'", one=True)["c"]
                r = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'RECHAZADO'", one=True)["c"]
            except Exception:
                p, a, r = 0, 0, 0
            return jsonify(ok=False, error="Solicitud no encontrada. Revise la base de datos.", pendientes=p, aprobadas=a, rechazadas=r), 404
        return redirect(url_for("permisos_index"))
    if solicitud["estado"] != "PENDIENTE":
        flash("La solicitud ya fue resuelta. Se actualiza la lista.", "info")
        if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            p = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'PENDIENTE'", one=True)["c"]
            a = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'APROBADO'", one=True)["c"]
            r = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'RECHAZADO'", one=True)["c"]
            return jsonify(ok=True, ya_resuelta=True, pendientes=p, aprobadas=a, rechazadas=r)
        return redirect(url_for("permisos_index"))
    execute(
        "UPDATE solicitud_permiso SET estado = 'RECHAZADO', observaciones = %s, resuelto_por = %s, fecha_resolucion = NOW() WHERE id = %s",
        (observaciones, get_current_user()["id_user"], id),
    )
    registrar_audit("Solicitud rechazada", "permisos", f"id={id} cédula={solicitud.get('id_cedula')}")
    emp = query("SELECT apellidos_nombre, direccion_email FROM empleado WHERE id_cedula = %s", (solicitud["id_cedula"],), one=True)
    email_empleado = _resolver_email_empleado(solicitud["id_cedula"])
    correo_ok = notificar_resolucion_permiso(app, solicitud, emp["apellidos_nombre"] if emp else "", email_empleado, aprobado=False, observaciones=observaciones)
    _cur = get_current_user() or {}
    if not _es_admin_o_coord(_cur):
        try:
            notificar_gh_resolucion_por_jefe(
                app, solicitud, emp["apellidos_nombre"] if emp else "",
                tipo="permiso", aprobado=False,
                jefe_nombre=_cur.get("nombre"), observaciones=observaciones,
            )
        except Exception:
            pass
    try:
        execute(
            "UPDATE solicitud_permiso SET correo_resolucion_enviado = %s, correo_resolucion_at = IF(%s, NOW(), NULL) WHERE id = %s",
            (1 if correo_ok else 0, correo_ok, id),
        )
    except Exception:
        pass  # columnas no existen si no se ejecutó migration_correo_resolucion_validar.sql
    if correo_ok:
        flash("Solicitud rechazada. Se ha notificado al empleado por correo.", "success")
    else:
        flash("Solicitud rechazada. No se pudo enviar el correo al empleado (revisar consola y MAIL_PASSWORD en .env).", "warning")
    if _is_api_request() or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        p = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'PENDIENTE'", one=True)["c"]
        a = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'APROBADO'", one=True)["c"]
        r = query("SELECT COUNT(*) as c FROM solicitud_permiso WHERE estado = 'RECHAZADO'", one=True)["c"]
        return jsonify(ok=True, pendientes=p, aprobadas=a, rechazadas=r)
    return redirect(url_for("permisos_index"))


@app.route("/permisos/<int:id>/evidencia")
@login_required
@module_required("permisos")
def permiso_evidencia(id):
    """La coordinación (o quien tenga módulo permisos) puede ver/descargar la evidencia adjunta de la solicitud."""
    solicitud = query("SELECT id, evidencia FROM solicitud_permiso WHERE id = %s", (id,), one=True)
    if not solicitud or not (solicitud.get("evidencia") or "").strip():
        flash("No hay evidencia adjunta para esta solicitud.", "info")
        return redirect(url_for("permisos_index"))
    evidencia_ruta = (solicitud["evidencia"] or "").strip()
    if ".." in evidencia_ruta or evidencia_ruta.startswith("/"):
        flash("Ruta de evidencia no válida.", "error")
        return redirect(url_for("permisos_index"))
    uploads_dir = os.path.join(current_app.instance_path, "uploads")
    full_path = os.path.normpath(os.path.join(uploads_dir, evidencia_ruta))
    if not full_path.startswith(os.path.normpath(uploads_dir)) or not os.path.isfile(full_path):
        flash("Archivo de evidencia no encontrado.", "error")
        return redirect(url_for("permisos_index"))
    nombre_descarga = os.path.basename(evidencia_ruta)
    return send_file(full_path, as_attachment=False, download_name=nombre_descarga, mimetype=None)


# ── INCIDENCIAS (INATEL) – Solo rol SISO ─────────────────────────────────────

@app.route("/incidencias")
@login_required
@module_required("incidencias")
def incidencias_index():
    """Listado de incidencias (accidentes, incidentes, enfermedades laborales). Solo SISO."""
    try:
        rows = query(
            "SELECT * FROM incidencia_at ORDER BY COALESCE(fecha_accidente, creado_en) DESC, id DESC"
        )
    except Exception:
        rows = []
    return render_template("incidencias_list.html", rows=rows, active_page="Incidencias")


_MESES_VALIDOS = ("ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE")


@app.route("/incidencias/dashboard")
@login_required
@module_required("incidencias")
def incidencias_dashboard():
    """Dashboard de incidencias para SISO. Filtros: ?mes=ENERO y ?anio=2026 para estadísticas por mes y/o año."""
    filtro_mes = (request.args.get("mes") or "").strip().upper()
    if filtro_mes and filtro_mes not in _MESES_VALIDOS:
        filtro_mes = ""
    try:
        filtro_anio = request.args.get("anio", "").strip()
        filtro_anio = int(filtro_anio) if filtro_anio and filtro_anio.isdigit() and 2000 <= int(filtro_anio) <= 2100 else None
    except (ValueError, TypeError):
        filtro_anio = None

    try:
        # Años que tienen al menos un registro (para el selector)
        años_rows = query(
            "SELECT DISTINCT YEAR(fecha_accidente) AS anio FROM incidencia_at WHERE fecha_accidente IS NOT NULL ORDER BY anio DESC"
        )
        años_disponibles = [r["anio"] for r in (años_rows or []) if r.get("anio")]

        # Condiciones y params según filtros
        cond_mes = " AND mes = %s" if filtro_mes else ""
        cond_anio = " AND YEAR(fecha_accidente) = %s" if filtro_anio else ""
        params_filtro = []
        if filtro_mes:
            params_filtro.append(filtro_mes)
        if filtro_anio:
            params_filtro.append(filtro_anio)
        params_filtro = tuple(params_filtro) if params_filtro else None

        where_total = ("WHERE 1=1" + cond_mes + cond_anio) if params_filtro else ""
        total = query(
            "SELECT COUNT(*) AS c FROM incidencia_at " + where_total,
            params_filtro,
            one=True,
        )
        total = (total or {}).get("c") or 0

        placeholders = ",".join(["%s"] * len(_MESES_VALIDOS))
        base_where = "WHERE mes IS NOT NULL AND mes != ''" + cond_mes + cond_anio
        group_mes = " GROUP BY mes ORDER BY FIELD(mes, " + placeholders + ")"
        params_por_mes = (params_filtro or ()) + _MESES_VALIDOS

        por_mes = query(
            "SELECT mes, COUNT(*) AS total FROM incidencia_at " + base_where + group_mes,
            params_por_mes,
        )
        severidad_mes = query(
            "SELECT mes, COALESCE(SUM(COALESCE(dias_incapacidad, 0)), 0) AS total FROM incidencia_at " + base_where + group_mes,
            params_por_mes,
        )

        base_where_resto = "WHERE 1=1" + cond_mes + cond_anio
        pr = params_filtro or ()

        por_tipo = query(
            "SELECT tipo_evento AS tipo, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND tipo_evento IS NOT NULL AND tipo_evento != '' GROUP BY tipo_evento",
            pr,
        )
        por_dia_semana = query(
            "SELECT dia_semana AS dia, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND dia_semana IS NOT NULL AND dia_semana != '' GROUP BY dia_semana ORDER BY FIELD(dia_semana, 'LUNES','MARTES','MIÉRCOLES','MIERCOLES','JUEVES','VIERNES','SÁBADO','SABADO','DOMINGO'), dia_semana",
            pr,
        )
        por_lugar = query(
            "SELECT area_seccion_ocurrencia AS lugar, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND area_seccion_ocurrencia IS NOT NULL AND area_seccion_ocurrencia != '' GROUP BY area_seccion_ocurrencia ORDER BY total DESC",
            pr,
        )
        por_forma = query(
            "SELECT forma_accidente AS forma, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND forma_accidente IS NOT NULL AND forma_accidente != '' GROUP BY forma_accidente ORDER BY total DESC",
            pr,
        )
        por_parte_cuerpo = query(
            "SELECT parte_cuerpo_afectada AS parte, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND parte_cuerpo_afectada IS NOT NULL AND parte_cuerpo_afectada != '' GROUP BY parte_cuerpo_afectada ORDER BY total DESC",
            pr,
        )
        por_tipo_lesion = query(
            "SELECT tipo_lesion AS lesion, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND tipo_lesion IS NOT NULL AND tipo_lesion != '' GROUP BY tipo_lesion ORDER BY total DESC",
            pr,
        )
        por_agente = query(
            "SELECT agente_lesion AS agente, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND agente_lesion IS NOT NULL AND agente_lesion != '' GROUP BY agente_lesion ORDER BY total DESC",
            pr,
        )
        por_genero = query(
            "SELECT genero AS genero, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND genero IS NOT NULL AND genero != '' GROUP BY genero ORDER BY total DESC",
            pr,
        )
        por_cargo = query(
            "SELECT cargo AS cargo, COUNT(*) AS total FROM incidencia_at " + base_where_resto + " AND cargo IS NOT NULL AND cargo != '' GROUP BY cargo ORDER BY total DESC",
            pr,
        )
    except Exception:
        total = 0
        años_disponibles = []
        por_mes = por_tipo = por_dia_semana = por_lugar = por_forma = []
        por_parte_cuerpo = por_tipo_lesion = por_agente = por_genero = por_cargo = []
        severidad_mes = []

    return render_template(
        "incidencias_dashboard.html",
        total=total,
        por_tipo=por_tipo,
        por_mes=por_mes,
        severidad_mes=severidad_mes,
        por_dia_semana=por_dia_semana,
        por_lugar=por_lugar,
        por_forma=por_forma,
        por_parte_cuerpo=por_parte_cuerpo,
        por_tipo_lesion=por_tipo_lesion,
        por_agente=por_agente,
        por_genero=por_genero,
        por_cargo=por_cargo,
        filtro_mes=filtro_mes,
        filtro_anio=filtro_anio,
        meses_validos=_MESES_VALIDOS,
        años_disponibles=años_disponibles,
        active_page="Dashboard Incidencias",
    )


def _incidencia_from_form():
    """Extrae dict de incidencia desde request.form."""
    def _d(k, default=None):
        v = (request.form.get(k) or "").strip()
        return v if v else default
    def _date(k):
        v = _d(k)
        if not v:
            return None
        try:
            from datetime import datetime
            return datetime.strptime(v[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    return {
        "numero_registro": int(request.form.get("numero_registro") or 0) or None,
        "mes": _d("mes"),
        "fecha_accidente": _date("fecha_accidente"),
        "dia_semana": _d("dia_semana"),
        "hora_ocurrencia": _d("hora_ocurrencia"),
        "tipo_evento": _d("tipo_evento"),
        "nombre_trabajador": _d("nombre_trabajador"),
        "cedula": _d("cedula"),
        "genero": _d("genero"),
        "cargo": _d("cargo"),
        "fecha_ingreso": _date("fecha_ingreso"),
        "antiguedad_meses": int(request.form.get("antiguedad_meses") or 0) or None,
        "area_seccion_ocurrencia": _d("area_seccion_ocurrencia"),
        "tipo_vinculacion": _d("tipo_vinculacion"),
        "dias_incapacidad": int(request.form.get("dias_incapacidad") or 0) or None,
        "prorroga": _d("prorroga"),
        "parte_cuerpo_afectada": _d("parte_cuerpo_afectada"),
        "tipo_lesion": _d("tipo_lesion"),
        "forma_accidente": _d("forma_accidente"),
        "clasificacion_origen": _d("clasificacion_origen"),
        "agente_lesion": _d("agente_lesion"),
        "reincidente": _d("reincidente"),
        "descripcion_accidente": _d("descripcion_accidente"),
        "investigado": _d("investigado"),
        "causas": _d("causas"),
        "seguimiento_clinico": _d("seguimiento_clinico"),
    }


@app.route("/incidencias/nueva", methods=["GET", "POST"])
@login_required
@module_required("incidencias")
def incidencias_nueva():
    """Alta de incidencia. Solo SISO."""
    if request.method == "POST":
        data = _incidencia_from_form()
        user = get_current_user()
        execute(
            """INSERT INTO incidencia_at (
                numero_registro, mes, fecha_accidente, dia_semana, hora_ocurrencia, tipo_evento,
                nombre_trabajador, cedula, genero, cargo, fecha_ingreso, antiguedad_meses,
                area_seccion_ocurrencia, tipo_vinculacion, dias_incapacidad, prorroga,
                parte_cuerpo_afectada, tipo_lesion, forma_accidente, clasificacion_origen,
                agente_lesion, reincidente, descripcion_accidente, investigado, causas,
                seguimiento_clinico, creado_por
            ) VALUES (
                %(numero_registro)s, %(mes)s, %(fecha_accidente)s, %(dia_semana)s, %(hora_ocurrencia)s, %(tipo_evento)s,
                %(nombre_trabajador)s, %(cedula)s, %(genero)s, %(cargo)s, %(fecha_ingreso)s, %(antiguedad_meses)s,
                %(area_seccion_ocurrencia)s, %(tipo_vinculacion)s, %(dias_incapacidad)s, %(prorroga)s,
                %(parte_cuerpo_afectada)s, %(tipo_lesion)s, %(forma_accidente)s, %(clasificacion_origen)s,
                %(agente_lesion)s, %(reincidente)s, %(descripcion_accidente)s, %(investigado)s, %(causas)s,
                %(seguimiento_clinico)s, %(creado_por)s
            )""",
            {**data, "creado_por": user.get("email") or user.get("id_user")},
        )
        flash("Incidencia registrada correctamente.", "success")
        return redirect(url_for("incidencias_index"))
    return render_template("incidencias_form.html", incidencia=None, active_page="Nueva incidencia")


@app.route("/incidencias/<int:id>/editar", methods=["GET", "POST"])
@login_required
@module_required("incidencias")
def incidencias_editar(id):
    """Editar incidencia. Solo SISO."""
    incidencia = query("SELECT * FROM incidencia_at WHERE id = %s", (id,), one=True)
    if not incidencia:
        flash("Incidencia no encontrada.", "error")
        return redirect(url_for("incidencias_index"))
    if request.method == "POST":
        data = _incidencia_from_form()
        execute(
            """UPDATE incidencia_at SET
                numero_registro=%(numero_registro)s, mes=%(mes)s, fecha_accidente=%(fecha_accidente)s,
                dia_semana=%(dia_semana)s, hora_ocurrencia=%(hora_ocurrencia)s, tipo_evento=%(tipo_evento)s,
                nombre_trabajador=%(nombre_trabajador)s, cedula=%(cedula)s, genero=%(genero)s, cargo=%(cargo)s,
                fecha_ingreso=%(fecha_ingreso)s, antiguedad_meses=%(antiguedad_meses)s,
                area_seccion_ocurrencia=%(area_seccion_ocurrencia)s, tipo_vinculacion=%(tipo_vinculacion)s,
                dias_incapacidad=%(dias_incapacidad)s, prorroga=%(prorroga)s,
                parte_cuerpo_afectada=%(parte_cuerpo_afectada)s, tipo_lesion=%(tipo_lesion)s,
                forma_accidente=%(forma_accidente)s, clasificacion_origen=%(clasificacion_origen)s,
                agente_lesion=%(agente_lesion)s, reincidente=%(reincidente)s,
                descripcion_accidente=%(descripcion_accidente)s, investigado=%(investigado)s,
                causas=%(causas)s, seguimiento_clinico=%(seguimiento_clinico)s
            WHERE id = %(id)s""",
            {**data, "id": id},
        )
        flash("Incidencia actualizada.", "success")
        return redirect(url_for("incidencias_index"))
    return render_template("incidencias_form.html", incidencia=incidencia, active_page="Editar incidencia")


@app.route("/incidencias/<int:id>/eliminar", methods=["POST"])
@login_required
@module_required("incidencias")
def incidencias_eliminar(id):
    """Eliminar incidencia. Solo SISO."""
    incidencia = query("SELECT id FROM incidencia_at WHERE id = %s", (id,), one=True)
    if not incidencia:
        flash("Incidencia no encontrada.", "error")
        return redirect(url_for("incidencias_index"))
    execute("DELETE FROM incidencia_at WHERE id = %s", (id,))
    flash("Incidencia eliminada.", "success")
    return redirect(url_for("incidencias_index"))


PERMISOS_EXPORT_COLUMNS = [
    ("id_cedula", "Cédula"),
    ("apellidos_nombre", "Nombre"),
    ("area", "Área"),
    ("tipo", "Tipo"),
    ("fecha_desde", "Fecha desde"),
    ("fecha_hasta", "Fecha hasta"),
    ("motivo", "Motivo"),
    ("estado", "Estado"),
    ("fecha_solicitud_str", "Fecha solicitud"),
    ("observaciones", "Observaciones"),
]


@app.route("/permisos/export")
@login_required
@module_required("permisos")
def permisos_export():
    """Exporta listado de solicitudes de permiso con los mismos filtros que la vista.
    Respeta el filtro por encargado (jefe ve solo los suyos).
    """
    if not _puede_ver_listado_solicitudes():
        flash("No tiene permiso para exportar.", "error")
        return redirect(url_for("permisos_index"))
    filtro_estado = request.args.get("estado", "").strip().upper()
    if filtro_estado not in ("PENDIENTE", "APROBADO", "RECHAZADO"):
        filtro_estado = None
    buscar = request.args.get("buscar", "").strip()
    area = request.args.get("area", "").strip() or None
    tipo = request.args.get("tipo", "").strip() or None
    orden = request.args.get("orden", "").strip() or None
    sql, params = _permisos_query(filtro_estado=filtro_estado, buscar=buscar, area=area, tipo=tipo, orden=orden)
    rows = query(sql, params)
    for r in rows:
        if r.get("fecha_solicitud"):
            d = r["fecha_solicitud"]
            r["fecha_solicitud_str"] = d.strftime("%d/%m/%Y %H:%M") if hasattr(d, "strftime") else str(d)
        else:
            r["fecha_solicitud_str"] = ""
    return export_excel_response_generic(rows, PERMISOS_EXPORT_COLUMNS, "Solicitudes_permiso")


@app.route("/api/aniversario")
@login_required
def api_aniversario():
    year = request.args.get("year", date.today().year, type=int)
    month = request.args.get("month", date.today().month, type=int)

    # Una sola consulta de perfiles (evita N+1)
    perfil_rows = query("SELECT id_perfil, perfil_ocupacional FROM perfil_ocupacional")
    perfil_map = {str(r["id_perfil"]).strip(): r["perfil_ocupacional"] or "" for r in perfil_rows}

    tipo_map, nivel_map, prof_map = _calendar_label_maps()

    rows = query(
        "SELECT e.id_cedula, e.apellidos_nombre, e.fecha_ingreso, e.fecha_nacimiento, "
        "e.departamento, e.area, e.id_perfil_ocupacional, e.sexo, e.celular, "
        "e.direccion_email, e.estado, e.tipo_documento, e.nivel_educativo, e.profesion "
        "FROM empleado e WHERE e.estado = 'ACTIVO'"
    )

    results = []
    for r in rows:
        fi = parse_fecha(r["fecha_ingreso"])
        if fi and fi.month == month:
            try:
                aniv_date = date(year, fi.month, fi.day)
            except ValueError:
                continue
            antiguedad = year - fi.year

            perfil = perfil_map.get(str(r["id_perfil_ocupacional"] or "").strip(), "")

            bd = parse_fecha(r["fecha_nacimiento"])
            mes_cumple = bd.month if bd else ""
            cumple_str = ""
            if bd:
                try:
                    cumple_str = date(year, bd.month, bd.day).strftime("%d/%m/%Y")
                except ValueError:
                    pass

            item = {
                "id_cedula": r["id_cedula"],
                "apellidos_nombre": r["apellidos_nombre"],
                "fecha_ingreso": r["fecha_ingreso"],
                "departamento": r["departamento"] or "",
                "area": r["area"] or "",
                "perfil_ocupacional": perfil,
                "sexo": r["sexo"] or "",
                "celular": r["celular"] or "",
                "correo": r["direccion_email"] or "",
                "estado": r["estado"] or "",
                "tipo_documento": r["tipo_documento"] or "",
                "nivel_educativo": r["nivel_educativo"] or "",
                "profesion": r["profesion"] or "",
                "mes_cumple": mes_cumple,
                "cumpleanos": cumple_str,
                "aniversario_laboral": aniv_date.strftime("%d/%m/%Y"),
                "antiguedad": antiguedad,
                "dia": fi.day,
            }
            enrich_calendar_row(item, tipo_map, nivel_map, prof_map)
            results.append(item)

    _resolve_calendar_ids_in_results(results)
    return jsonify(results)


@app.route("/api/empleado/<id_cedula>", methods=["GET"])
@login_required
@module_required("personal")
def api_empleado_get(id_cedula):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (id_cedula,), one=True)
    if not emp:
        return jsonify({"error": "No encontrado"}), 404
    return jsonify(emp)


@app.route("/api/empleado/<id_cedula>", methods=["PUT"])
@login_required
@module_required("personal")
@role_required("WRITE")
def api_empleado_update(id_cedula):
    data = request.get_json()
    allowed = [
        "apellidos_nombre", "departamento", "area", "sexo", "celular",
        "direccion_email", "fecha_nacimiento", "estado", "tipo_documento",
        "nivel_educativo", "profesion", "contacto_emergencia",
        "telefono_contacto", "parentezco",
    ]
    sets = []
    vals = []
    for key in allowed:
        if key in data:
            sets.append(f"{key} = %s")
            vals.append(data[key])
    if not sets:
        return jsonify({"error": "Sin cambios"}), 400
    vals.append(id_cedula)
    execute(f"UPDATE empleado SET {', '.join(sets)} WHERE id_cedula = %s", tuple(vals))
    return jsonify({"ok": True})


# ── PERSONAL ACTIVO ──────────────────────────────────────────

@app.route("/personal-activo")
@login_required
@module_required("personal")
def personal_activo():
    # Solo un modo: inactivos=1 en la URL → inactivos; si no → activos
    show_inactivos = request.args.get("inactivos", "").strip().lower() in ("1", "true", "yes")
    show_activos = not show_inactivos
    if show_inactivos:
        rows = query(
            "SELECT id_cedula, apellidos_nombre, tipo_documento, departamento, "
            "area, sexo, fecha_ingreso, celular, eps, estado "
            "FROM empleado WHERE estado = 'INACTIVO' ORDER BY apellidos_nombre"
        )
        columns = [
            {"key": "id_cedula",        "label": "Cédula"},
            {"key": "apellidos_nombre", "label": "Nombre"},
            {"key": "tipo_documento",   "label": "Tipo Doc"},
            {"key": "departamento",     "label": "Departamento", "type": "dept"},
            {"key": "area",             "label": "Área"},
            {"key": "sexo",             "label": "Sexo", "type": "sex"},
            {"key": "fecha_ingreso",    "label": "Fecha Ingreso"},
            {"key": "celular",          "label": "Celular"},
            {"key": "eps",              "label": "EPS"},
        ]
        filter_columns = [
            {"index": 3, "label": "Departamento"},
            {"index": 5, "label": "Sexo"},
            {"index": 8, "label": "EPS"},
        ]
    else:
        rows = query(
            "SELECT id_cedula, apellidos_nombre, tipo_documento, departamento, "
            "area, sexo, fecha_ingreso, celular, eps, estado "
            "FROM empleado WHERE estado = 'ACTIVO' ORDER BY apellidos_nombre"
        )
        columns = [
            {"key": "id_cedula",        "label": "Cédula"},
            {"key": "apellidos_nombre", "label": "Nombre"},
            {"key": "tipo_documento",   "label": "Tipo Doc"},
            {"key": "departamento",     "label": "Departamento", "type": "dept"},
            {"key": "area",             "label": "Área"},
            {"key": "sexo",             "label": "Sexo", "type": "sex"},
            {"key": "fecha_ingreso",    "label": "Fecha Ingreso"},
            {"key": "celular",          "label": "Celular"},
            {"key": "eps",              "label": "EPS"},
        ]
        filter_columns = [
            {"index": 3, "label": "Departamento"},
            {"index": 5, "label": "Sexo"},
            {"index": 8, "label": "EPS"},
        ]
    sex_m = sum(1 for r in rows if r.get("sexo") == "M")
    sex_f = sum(1 for r in rows if r.get("sexo") == "F")
    deptos = len(set(r["departamento"] for r in rows if r.get("departamento")))
    n_activos = sum(1 for r in rows if r.get("estado") == "ACTIVO")
    n_inactivos = sum(1 for r in rows if r.get("estado") == "INACTIVO")
    stats = [
        {"value": len(rows), "label": "Total",           "icon": "group",      "color": "green"},
        {"value": n_activos, "label": "Activos",         "icon": "person_check", "color": "green"},
        {"value": n_inactivos, "label": "Inactivos",     "icon": "person_off", "color": "orange"},
        {"value": deptos,    "label": "Departamentos",   "icon": "business",   "color": "blue"},
    ]
    user = get_current_user()
    perm = get_role_permission(user["rol"] or "") if user else "READ"
    show_add_btn = perm in ("WRITE", "ALL")  # Solo Contratación, Coord. GH y Admin pueden agregar
    return render_template(
        "data_table.html", active_page="Personal Activo",
        rows=rows, columns=columns, stats=stats,
        detail_route="detalle_empleado", pk="id_cedula",
        export_key="personal_activo", filter_columns=filter_columns,
        personal_tipo_selector=True,
        personal_show_activos=show_activos,
        personal_show_inactivos=show_inactivos,
        personal_activo_url=url_for("personal_activo"),
        show_add_btn=show_add_btn, add_url=url_for("crear_empleado"),
    )


EMPLEADO_FIELDS = [
    {"key": "id_cedula",            "label": "Cédula"},
    {"key": "apellidos_nombre",     "label": "Nombre Completo"},
    {"key": "tipo_documento",       "label": "Tipo Documento"},
    {"key": "lugar_expedicion",     "label": "Lugar Expedición"},
    {"key": "fecha_expedicion",     "label": "Fecha Expedición"},
    {"key": "fecha_nacimiento",     "label": "Fecha Nacimiento"},
    {"key": "sexo",                 "label": "Sexo"},
    {"key": "rh",                   "label": "RH"},
    {"key": "departamento",         "label": "Departamento"},
    {"key": "area",                 "label": "Área"},
    {"key": "fecha_ingreso",        "label": "Fecha Ingreso"},
    {"key": "direccion_residencia", "label": "Dirección"},
    {"key": "barrio_residencia",    "label": "Barrio"},
    {"key": "ciudad_residencia",    "label": "Ciudad"},
    {"key": "telefono",             "label": "Teléfono"},
    {"key": "celular",              "label": "Celular"},
    {"key": "direccion_email",      "label": "Email"},
    {"key": "eps",                  "label": "EPS"},
    {"key": "fondo_pensiones",      "label": "Fondo Pensiones"},
    {"key": "nivel_educativo",      "label": "Nivel Educativo"},
    {"key": "profesion",            "label": "Profesión"},
    {"key": "hijos",                "label": "¿Tiene Hijos?"},
    {"key": "contacto_emergencia",  "label": "Contacto Emergencia"},
    {"key": "telefono_contacto",    "label": "Tel. Contacto"},
    {"key": "parentezco",           "label": "Parentesco"},
    {"key": "estado",               "label": "Estado", "type": "badge"},
]


def _procesar_hijos_form(id_cedula, request_form, reemplazar=False):
    """Lee las listas hijo_id[], hijo_apellidos_nombre[], hijo_identificacion[],
    hijo_fecha_nacimiento[], hijo_sexo[], hijo_estado[] del formulario y sincroniza
    la tabla `hijo` para el empleado dado.
    - Filas sin nombre se ignoran (se consideran vacías).
    - Si reemplazar=True (caso edición), los hijos existentes que no figuren en el form
      se eliminan.
    - Si una fila trae hijo_id lo actualiza; si no, se inserta con un id generado.
    Devuelve la cantidad de hijos resultantes para el empleado.
    """
    import uuid
    if not id_cedula:
        return 0
    nombres = request_form.getlist("hijo_apellidos_nombre[]")
    identificaciones = request_form.getlist("hijo_identificacion[]")
    fechas = request_form.getlist("hijo_fecha_nacimiento[]")
    sexos = request_form.getlist("hijo_sexo[]")
    estados = request_form.getlist("hijo_estado[]")
    ids = request_form.getlist("hijo_id[]")

    def _g(lst, i):
        return (lst[i].strip() if i < len(lst) and lst[i] is not None else "")

    total_rows = max(len(nombres), len(ids))
    ids_en_form = set()
    for i in range(total_rows):
        nombre = _g(nombres, i).upper()
        if not nombre:
            continue
        hijo_id = _g(ids, i)
        identificacion = _g(identificaciones, i) or None
        fecha_nac = _g(fechas, i) or None
        sexo = _g(sexos, i) or None
        estado = _g(estados, i) or "ACTIVO"
        if hijo_id:
            ids_en_form.add(hijo_id)
            execute(
                "UPDATE hijo SET apellidos_nombre=%s, identificacion_hijo=%s, fecha_nacimiento=%s, "
                "sexo=%s, estado=%s WHERE id_hijo=%s AND id_cedula=%s",
                (nombre, identificacion, fecha_nac, sexo, estado, hijo_id, id_cedula),
            )
        else:
            new_id = uuid.uuid4().hex[:8]
            execute(
                "INSERT INTO hijo (id_hijo, id_cedula, apellidos_nombre, identificacion_hijo, "
                "fecha_nacimiento, sexo, estado) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (new_id, id_cedula, nombre, identificacion, fecha_nac, sexo, estado),
            )
            ids_en_form.add(new_id)

    if reemplazar:
        existentes = query("SELECT id_hijo FROM hijo WHERE id_cedula = %s", (id_cedula,))
        for row in existentes:
            if row["id_hijo"] not in ids_en_form:
                execute("DELETE FROM hijo WHERE id_hijo = %s AND id_cedula = %s", (row["id_hijo"], id_cedula))

    total = query("SELECT COUNT(*) AS c FROM hijo WHERE id_cedula = %s", (id_cedula,), one=True)
    return int(total["c"]) if total else 0


def _form_context():
    """Lookup data for the empleado form dropdowns."""
    try:
        encargados_list = query(
            "SELECT id_user, nombre, rol FROM usuario "
            "WHERE COALESCE(estado, 1) = 1 AND UPPER(COALESCE(rol, '')) <> 'EMPLEADO' "
            "ORDER BY nombre"
        )
    except Exception:
        encargados_list = []
    return dict(
        tipos_doc=query("SELECT DISTINCT tipo_documento FROM tipo_documento ORDER BY tipo_documento"),
        departamentos=query("SELECT nombre FROM departamento ORDER BY nombre"),
        areas=query(
            "SELECT a.id, a.nombre, d.nombre AS departamento "
            "FROM area a JOIN departamento d ON a.departamento_id = d.id ORDER BY a.nombre"
        ),
        perfiles=query(
            "SELECT p.id_perfil, p.perfil_ocupacional, a.nombre AS area "
            "FROM perfil_ocupacional p JOIN area a ON p.area_id = a.id "
            "ORDER BY p.perfil_ocupacional"
        ),
        eps_list=query("SELECT nombre FROM eps ORDER BY nombre"),
        fondos_list=query("SELECT nombre FROM fondo_pensiones ORDER BY nombre"),
        niveles=query("SELECT DISTINCT nivel FROM nivel_educativo ORDER BY nivel"),
        profesiones=query("SELECT DISTINCT profesion FROM profesion ORDER BY profesion"),
        encargados_list=encargados_list,
    )


@app.route("/personal-activo/<id>")
@login_required
@module_required("personal")
def detalle_empleado(id):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (id,), one=True)
    if not emp:
        flash("Empleado no encontrado", "error")
        return redirect(url_for("personal_activo"))
    hijos = query("SELECT * FROM hijo WHERE id_cedula = %s ORDER BY fecha_nacimiento", (id,))
    retirados = query(
        "SELECT id_retiro, fecha_retiro, tipo_retiro, dias_laborados, motivo "
        "FROM retirado WHERE id_cedula = %s ORDER BY fecha_retiro DESC", (id,)
    )
    encargado_info = None
    if emp.get("id_user_encargado"):
        try:
            encargado_info = query(
                "SELECT id_user, nombre, email, rol FROM usuario WHERE id_user = %s",
                (emp["id_user_encargado"],), one=True,
            )
        except Exception:
            encargado_info = None
    return render_template(
        "detail.html", active_page="Personal Activo",
        data=emp, fields=EMPLEADO_FIELDS, children=hijos,
        retirados=retirados, encargado_info=encargado_info,
        back_url=url_for("personal_activo"),
    )


@app.route("/personal-activo/nuevo", methods=["GET", "POST"])
@login_required
@module_required("personal")
@role_required("WRITE")
def crear_empleado():
    if request.method == "POST":
        cedula = request.form.get("id_cedula", "").strip()
        nombre = request.form.get("apellidos_nombre", "").strip().upper()
        if not cedula or not nombre:
            flash("Cédula y Nombre son obligatorios", "error")
            return redirect(url_for("crear_empleado"))
        existing = query("SELECT id_cedula FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
        if existing:
            flash(f"Ya existe un empleado con cédula {cedula}", "error")
            return redirect(url_for("crear_empleado"))
        fields_list = [
            "id_cedula", "apellidos_nombre", "tipo_documento", "lugar_expedicion",
            "fecha_expedicion", "departamento", "area", "id_perfil_ocupacional",
            "fecha_ingreso", "sexo", "rh", "direccion_residencia", "barrio_residencia",
            "ciudad_residencia", "telefono", "celular", "direccion_email", "eps",
            "fondo_pensiones", "fecha_nacimiento", "hijos", "estado", "nivel_educativo",
            "profesion", "contacto_emergencia", "telefono_contacto", "parentezco",
            "id_user_encargado",
        ]
        vals = []
        for f in fields_list:
            v = request.form.get(f, "").strip()
            vals.append(v if v else None)
        try:
            execute(f"INSERT INTO empleado ({', '.join(fields_list)}) VALUES ({', '.join(['%s'] * len(fields_list))})", tuple(vals))
        except Exception as e:
            # Compat: si la BD aún no tiene la columna id_user_encargado, reintentar sin ella.
            if "id_user_encargado" in str(e).lower():
                fields_fallback = [f for f in fields_list if f != "id_user_encargado"]
                vals_fallback = [request.form.get(f, "").strip() or None for f in fields_fallback]
                execute(
                    f"INSERT INTO empleado ({', '.join(fields_fallback)}) VALUES ({', '.join(['%s'] * len(fields_fallback))})",
                    tuple(vals_fallback),
                )
            else:
                raise
        try:
            total_hijos = _procesar_hijos_form(cedula, request.form, reemplazar=False)
            if total_hijos > 0:
                execute("UPDATE empleado SET hijos = 'SI' WHERE id_cedula = %s", (cedula,))
        except Exception:
            pass
        try:
            ruta_foto = _guardar_foto_empleado(cedula, request.files.get("foto"))
            if ruta_foto:
                _actualizar_foto_empleado_db(cedula, ruta_foto)
        except ValueError as e:
            flash(str(e), "warning")
        # Crear usuario para portal del empleado (contraseña inicial Colbeef2026*)
        id_user_emp = "EMP-" + cedula
        email_emp = (request.form.get("direccion_email") or "").strip().lower()
        if not email_emp:
            email_emp = cedula + "@empleado.colbeef.local"
        if not query("SELECT id_user FROM usuario WHERE id_user = %s", (id_user_emp,), one=True):
            try:
                execute(
                    "INSERT INTO usuario (id_user, email, password_hash, nombre, rol, estado, id_cedula, debe_cambiar_clave) VALUES (%s, %s, %s, %s, 'EMPLEADO', 1, %s, 1)",
                    (id_user_emp, email_emp, generate_password_hash(EMPLEADO_PASSWORD_DEFAULT), nombre, cedula),
                )
            except Exception as e:
                if "debe_cambiar_clave" in str(e):
                    execute(
                        "INSERT INTO usuario (id_user, email, password_hash, nombre, rol, estado, id_cedula) VALUES (%s, %s, %s, %s, 'EMPLEADO', 1, %s)",
                        (id_user_emp, email_emp, generate_password_hash(EMPLEADO_PASSWORD_DEFAULT), nombre, cedula),
                    )
                else:
                    raise
            flash(
                f"Empleado {nombre} creado. Para el portal: correo {email_emp}, contraseña inicial Colbeef2026*. Podrá cambiarla al ingresar.",
                "success",
            )
        else:
            flash(f"Empleado {nombre} creado exitosamente", "success")
        return redirect(url_for("detalle_empleado", id=cedula))

    ctx = _form_context()
    return render_template(
        "empleado_form.html", active_page="Personal Activo",
        emp=None, hijos_emp=[], back_url=url_for("personal_activo"), **ctx,
    )


@app.route("/personal-activo/<id>/editar", methods=["GET", "POST"])
@login_required
@module_required("personal")
@role_required("WRITE")
def editar_empleado(id):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (id,), one=True)
    if not emp:
        flash("Empleado no encontrado", "error")
        return redirect(url_for("personal_activo"))

    if request.method == "POST":
        update_fields = [
            "apellidos_nombre", "tipo_documento", "lugar_expedicion", "fecha_expedicion",
            "departamento", "area", "id_perfil_ocupacional", "fecha_ingreso", "sexo", "rh",
            "direccion_residencia", "barrio_residencia", "ciudad_residencia", "telefono",
            "celular", "direccion_email", "eps", "fondo_pensiones", "fecha_nacimiento",
            "hijos", "estado", "nivel_educativo", "profesion", "contacto_emergencia",
            "telefono_contacto", "parentezco", "id_user_encargado",
        ]
        sets = []
        vals = []
        for f in update_fields:
            v = request.form.get(f, "").strip()
            sets.append(f"{f} = %s")
            vals.append(v if v else None)
        vals.append(id)
        try:
            execute(f"UPDATE empleado SET {', '.join(sets)} WHERE id_cedula = %s", tuple(vals))
        except Exception as e:
            if "id_user_encargado" in str(e).lower():
                fb_fields = [f for f in update_fields if f != "id_user_encargado"]
                fb_sets = [f"{f} = %s" for f in fb_fields]
                fb_vals = [request.form.get(f, "").strip() or None for f in fb_fields] + [id]
                execute(f"UPDATE empleado SET {', '.join(fb_sets)} WHERE id_cedula = %s", tuple(fb_vals))
            else:
                raise
        try:
            total_hijos = _procesar_hijos_form(id, request.form, reemplazar=True)
            estado_hijos = "SI" if total_hijos > 0 else (request.form.get("hijos", "").strip() or None)
            execute("UPDATE empleado SET hijos = %s WHERE id_cedula = %s", (estado_hijos, id))
        except Exception:
            pass
        try:
            ruta_foto = _guardar_foto_empleado(id, request.files.get("foto"))
            if ruta_foto:
                _actualizar_foto_empleado_db(id, ruta_foto)
        except ValueError as e:
            flash(str(e), "warning")
        flash("Empleado actualizado exitosamente", "success")
        return redirect(url_for("detalle_empleado", id=id))

    ctx = _form_context()
    hijos_emp = query(
        "SELECT id_hijo, identificacion_hijo, apellidos_nombre, fecha_nacimiento, sexo, estado "
        "FROM hijo WHERE id_cedula = %s ORDER BY apellidos_nombre", (id,)
    )
    return render_template(
        "empleado_form.html", active_page="Personal Activo",
        emp=emp, hijos_emp=hijos_emp, back_url=url_for("detalle_empleado", id=id), **ctx,
    )


@app.route("/personal-activo/<id>/eliminar", methods=["POST"])
@login_required
@module_required("personal")
@role_required("ALL")
def eliminar_empleado(id):
    execute("DELETE FROM hijo WHERE id_cedula = %s", (id,))
    execute("DELETE FROM retirado WHERE id_cedula = %s", (id,))
    execute("DELETE FROM empleado WHERE id_cedula = %s", (id,))
    flash("Empleado eliminado", "success")
    return redirect(url_for("personal_activo"))


@app.route("/personal-activo/<id>/retirar", methods=["GET", "POST"])
@login_required
@module_required("personal")
@role_required("WRITE")
def retirar_empleado(id):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (id,), one=True)
    if not emp:
        flash("Empleado no encontrado", "error")
        return redirect(url_for("personal_activo"))

    if request.method == "POST":
        fecha_retiro = request.form.get("fecha_retiro", "").strip()
        tipo_retiro = request.form.get("tipo_retiro", "").strip()
        dias_laborados = request.form.get("dias_laborados", "").strip()
        motivo = request.form.get("motivo", "").strip()
        if not fecha_retiro or not tipo_retiro:
            flash("Fecha y Tipo de retiro son obligatorios", "error")
            return redirect(url_for("retirar_empleado", id=id))

        last = query("SELECT id_retiro FROM retirado ORDER BY id_retiro DESC LIMIT 1", one=True)
        if last:
            try:
                num = int(str(last["id_retiro"]).replace("RET-", "")) + 1
            except (ValueError, TypeError):
                num = query("SELECT COUNT(*) AS c FROM retirado", one=True)["c"] + 1
        else:
            num = 1
        new_id = f"RET-{num:04d}"

        fr_formatted = fecha_retiro
        if "-" in fecha_retiro:
            parts = fecha_retiro.split("-")
            fr_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"

        execute(
            "INSERT INTO retirado (id_retiro, id_cedula, apellidos_nombre, departamento, "
            "area, id_perfil_ocupacional, fecha_ingreso, fecha_retiro, dias_laborados, "
            "tipo_retiro, motivo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (new_id, emp["id_cedula"], emp["apellidos_nombre"], emp["departamento"],
             emp["area"], emp.get("id_perfil_ocupacional"), emp["fecha_ingreso"],
             fr_formatted, int(dias_laborados) if dias_laborados else None,
             tipo_retiro, motivo or None),
        )
        execute("UPDATE empleado SET estado = 'INACTIVO' WHERE id_cedula = %s", (id,))
        flash(f"Retiro registrado. Empleado {emp['apellidos_nombre']} marcado como INACTIVO", "success")
        return redirect(url_for("detalle_empleado", id=id))

    motivos = query("SELECT tipo_retiro FROM motivo_retiro ORDER BY tipo_retiro")
    return render_template(
        "retirar_form.html", active_page="Personal Activo",
        emp=emp, motivos=motivos,
        back_url=url_for("detalle_empleado", id=id),
    )


# ── CRUD HIJOS ───────────────────────────────────────────────

def _safe_redirect_to(target, fallback_endpoint, **fallback_kwargs):
    """Redirige a target si es una ruta interna segura, si no al fallback."""
    if target and target.startswith("/") and not target.startswith("//"):
        return redirect(target)
    return redirect(url_for(fallback_endpoint, **fallback_kwargs))


@app.route("/hijo/nuevo", methods=["POST"])
@login_required
@module_required("familia")
@role_required("WRITE")
def crear_hijo():
    cedula = request.form.get("id_cedula", "").strip()
    nombre = request.form.get("apellidos_nombre", "").strip().upper()
    redirect_to = request.form.get("redirect_to", "").strip()
    if not cedula or not nombre:
        flash("Cédula del padre y nombre del hijo son obligatorios", "error")
        return _safe_redirect_to(redirect_to, "detalle_empleado", id=cedula)

    padre = query("SELECT id_cedula FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
    if not padre:
        flash(f"No existe un empleado con cédula {cedula}. Verifica el número.", "error")
        return _safe_redirect_to(redirect_to, "hijos_gestion")

    last = query("SELECT id_hijo FROM hijo ORDER BY id_hijo DESC LIMIT 1", one=True)
    if last:
        try:
            num = int(str(last["id_hijo"]).replace("H-", "")) + 1
        except (ValueError, TypeError):
            num = query("SELECT COUNT(*) AS c FROM hijo", one=True)["c"] + 1
    else:
        num = 1
    new_id = f"H-{num:04d}"

    execute(
        "INSERT INTO hijo (id_hijo, identificacion_hijo, id_cedula, apellidos_nombre, "
        "fecha_nacimiento, sexo, estado) VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (new_id,
         request.form.get("identificacion_hijo", "").strip() or None,
         cedula, nombre,
         request.form.get("fecha_nacimiento", "").strip() or None,
         request.form.get("sexo", "").strip() or None,
         request.form.get("estado", "ACTIVO").strip()),
    )
    flash(f"Hijo {nombre} agregado exitosamente", "success")
    return _safe_redirect_to(redirect_to, "detalle_empleado", id=cedula)


@app.route("/hijo/<id>/editar", methods=["POST"])
@login_required
@module_required("familia")
@role_required("WRITE")
def editar_hijo(id):
    cedula = request.form.get("id_cedula", "").strip()
    nombre = request.form.get("apellidos_nombre", "").strip().upper()
    redirect_to = request.form.get("redirect_to", "").strip()
    execute(
        "UPDATE hijo SET identificacion_hijo=%s, apellidos_nombre=%s, "
        "fecha_nacimiento=%s, sexo=%s, estado=%s WHERE id_hijo=%s",
        (request.form.get("identificacion_hijo", "").strip() or None,
         nombre,
         request.form.get("fecha_nacimiento", "").strip() or None,
         request.form.get("sexo", "").strip() or None,
         request.form.get("estado", "ACTIVO").strip(),
         id),
    )
    flash(f"Hijo {nombre} actualizado", "success")
    return _safe_redirect_to(redirect_to, "detalle_empleado", id=cedula)


@app.route("/hijo/<id>/eliminar", methods=["POST"])
@login_required
@module_required("familia")
@role_required("ALL")
def eliminar_hijo(id):
    redirect_cedula = request.args.get("redirect_cedula", "") or request.form.get("redirect_cedula", "")
    redirect_to = request.args.get("redirect_to", "") or request.form.get("redirect_to", "")
    execute("DELETE FROM hijo WHERE id_hijo = %s", (id,))
    flash("Hijo eliminado", "success")
    if redirect_to:
        return _safe_redirect_to(redirect_to, "hijos_gestion")
    if redirect_cedula:
        return redirect(url_for("detalle_empleado", id=redirect_cedula))
    return redirect(url_for("hijos_gestion"))


# ── CRUD RETIRADOS ───────────────────────────────────────────

@app.route("/retirado/<id>/eliminar", methods=["POST"])
@login_required
@module_required("retiro")
@role_required("ALL")
def eliminar_retirado(id):
    redirect_cedula = request.args.get("redirect_cedula", "")
    execute("DELETE FROM retirado WHERE id_retiro = %s", (id,))
    flash("Registro de retiro eliminado", "success")
    if redirect_cedula:
        return redirect(url_for("detalle_empleado", id=redirect_cedula))
    return redirect(url_for("retiro_personal"))


# ── PERSONAL INACTIVO ────────────────────────────────────────

@app.route("/personal-inactivo")
@login_required
@module_required("personal_inactivo")
def personal_inactivo():
    rows = query(
        "SELECT id_cedula, apellidos_nombre, tipo_documento, departamento, "
        "area, sexo, fecha_ingreso, celular, eps, estado "
        "FROM empleado WHERE estado = 'INACTIVO' ORDER BY apellidos_nombre"
    )
    sex_m = sum(1 for r in rows if r.get("sexo") == "M")
    sex_f = sum(1 for r in rows if r.get("sexo") == "F")
    columns = [
        {"key": "id_cedula",        "label": "Cédula"},
        {"key": "apellidos_nombre", "label": "Nombre"},
        {"key": "tipo_documento",   "label": "Tipo Doc"},
        {"key": "departamento",     "label": "Departamento", "type": "dept"},
        {"key": "area",             "label": "Área"},
        {"key": "sexo",             "label": "Sexo", "type": "sex"},
        {"key": "fecha_ingreso",    "label": "Fecha Ingreso"},
        {"key": "celular",          "label": "Celular"},
        {"key": "eps",              "label": "EPS"},
    ]
    stats = [
        {"value": len(rows), "label": "Total Inactivos", "icon": "person_off",  "color": "red"},
        {"value": sex_m,     "label": "Masculino",       "icon": "male",        "color": "blue"},
        {"value": sex_f,     "label": "Femenino",        "icon": "female",      "color": "purple"},
    ]
    filter_columns = [
        {"index": 3, "label": "Departamento"},
        {"index": 5, "label": "Sexo"},
    ]
    return render_template(
        "data_table.html", active_page="Personal Inactivo",
        rows=rows, columns=columns, stats=stats,
        detail_route="detalle_empleado", pk="id_cedula",
        export_key="personal_inactivo", filter_columns=filter_columns,
    )


# ── HIJOS (ACTIVOS / INACTIVOS) ──────────────────────────────

@app.route("/hijos-activos")
@login_required
@module_required("familia")
def hijos_activos():
    estado = request.args.get("estado", "ACTIVO").strip().upper()
    if estado not in ("ACTIVO", "INACTIVO"):
        estado = "ACTIVO"
    rows = query(
        "SELECT h.id_hijo, h.identificacion_hijo, h.id_cedula, "
        "e.apellidos_nombre AS nombre_padre, h.apellidos_nombre AS nombre_hijo, "
        "h.fecha_nacimiento, h.sexo, h.estado "
        "FROM hijo h "
        "LEFT JOIN empleado e ON e.id_cedula = h.id_cedula "
        "WHERE h.estado = %s ORDER BY COALESCE(e.apellidos_nombre, h.id_cedula), h.apellidos_nombre",
        (estado,),
    )
    sex_m = sum(1 for r in rows if r.get("sexo") == "M")
    sex_f = sum(1 for r in rows if r.get("sexo") == "F")
    padres_set = set(r["id_cedula"] for r in rows if r.get("id_cedula"))
    padres = len(padres_set)
    columns = [
        {"key": "id_cedula",      "label": "Cédula Padre"},
        {"key": "nombre_padre",   "label": "Nombre Padre"},
        {"key": "nombre_hijo",    "label": "Nombre Hijo"},
        {"key": "identificacion_hijo", "label": "Identificación"},
        {"key": "fecha_nacimiento",   "label": "Fecha Nacimiento"},
        {"key": "sexo",           "label": "Sexo", "type": "sex"},
    ]
    filter_columns = [
        {"index": 0, "label": "Cédula Padre"},
        {"index": 5, "label": "Sexo"},
    ]
    base_label = "Hijos Activos" if estado == "ACTIVO" else "Hijos Inactivos"
    base_icon = "child_care" if estado == "ACTIVO" else "child_friendly"
    base_color = "green" if estado == "ACTIVO" else "orange"
    stats = [
        {"value": len(rows), "label": base_label,         "icon": base_icon,        "color": base_color, "filter": {"col": "sexo", "value": None}},
        {"value": sex_m,     "label": "Niños",            "icon": "male",           "color": "blue",     "filter": {"col": "sexo", "value": "M"}},
        {"value": sex_f,     "label": "Niñas",            "icon": "female",         "color": "purple",   "filter": {"col": "sexo", "value": "F"}},
        {"value": padres,    "label": "Empleados Padres", "icon": "family_restroom","color": "orange",   "filter": {"col": "id_cedula", "value": "__parent__"}},
    ]
    active_page = "Hijos Activos" if estado == "ACTIVO" else "Hijos Inactivos"
    export_key = "hijos_activos" if estado == "ACTIVO" else "hijos_inactivos"
    return render_template(
        "data_table.html", active_page=active_page,
        rows=rows, columns=columns, stats=stats,
        export_key=export_key, filter_columns=filter_columns,
        hijos_estado_selector=True,
        hijos_estado_actual=estado,
        hijos_base_url=url_for("hijos_activos"),
    )


@app.route("/hijos-inactivos")
@login_required
@module_required("familia")
def hijos_inactivos():
    # Redirige a la vista unificada con el estado INACTIVO seleccionado
    return redirect(url_for("hijos_activos", estado="INACTIVO"))


# ── API para autocompletar nombre del padre al escribir cedula ──
@app.route("/api/padre/<id_cedula>", methods=["GET"])
@login_required
@module_required("familia")
def api_padre_por_cedula(id_cedula):
    emp = query(
        "SELECT id_cedula, apellidos_nombre, departamento, area, estado "
        "FROM empleado WHERE id_cedula = %s",
        (id_cedula.strip(),), one=True,
    )
    if not emp:
        return jsonify({"ok": False, "error": "No existe un empleado con esa cédula"}), 404
    return jsonify({"ok": True, "empleado": emp})


# ── GESTION DE HIJOS (vista dedicada con busqueda + CRUD) ────
@app.route("/hijos-gestion")
@login_required
@module_required("familia")
def hijos_gestion():
    estado = request.args.get("estado", "TODOS").strip().upper()
    q_texto = request.args.get("q", "").strip()

    sql = (
        "SELECT h.id_hijo, h.identificacion_hijo, h.id_cedula, "
        "COALESCE(e.apellidos_nombre, '(padre no registrado)') AS nombre_padre, "
        "e.departamento, e.area, e.estado AS estado_padre, "
        "h.apellidos_nombre AS nombre_hijo, "
        "h.fecha_nacimiento, h.sexo, h.estado AS estado_hijo "
        "FROM hijo h "
        "LEFT JOIN empleado e ON e.id_cedula = h.id_cedula "
    )
    where = []
    params = []
    if estado in ("ACTIVO", "INACTIVO"):
        where.append("h.estado = %s")
        params.append(estado)
    if q_texto:
        where.append(
            "(h.id_cedula LIKE %s OR h.apellidos_nombre LIKE %s "
            "OR e.apellidos_nombre LIKE %s OR h.identificacion_hijo LIKE %s)"
        )
        like = f"%{q_texto}%"
        params.extend([like, like, like, like])
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY e.apellidos_nombre IS NULL, e.apellidos_nombre, h.apellidos_nombre"

    rows = query(sql, tuple(params))

    total = len(rows)
    activos = sum(1 for r in rows if (r.get("estado_hijo") or "").upper() == "ACTIVO")
    inactivos = total - activos
    padres_distintos = len({r["id_cedula"] for r in rows if r.get("id_cedula")})

    stats = [
        {"value": total,     "label": "Hijos encontrados",  "icon": "groups",           "color": "blue"},
        {"value": activos,   "label": "Activos",            "icon": "child_care",       "color": "green"},
        {"value": inactivos, "label": "Inactivos",          "icon": "child_friendly",   "color": "orange"},
        {"value": padres_distintos, "label": "Padres distintos", "icon": "family_restroom", "color": "purple"},
    ]

    return render_template(
        "hijos_gestion.html",
        active_page="Gestión de Hijos",
        rows=rows,
        stats=stats,
        estado_actual=estado,
        q_texto=q_texto,
        redirect_to=request.full_path.rstrip("?") or url_for("hijos_gestion"),
    )


# ── RETIRO DE PERSONAL ───────────────────────────────────────

@app.route("/retiro-personal")
@login_required
@module_required("retiro")
def retiro_personal():
    rows = query(
        "SELECT id_retiro, id_cedula, apellidos_nombre, departamento, area, "
        "fecha_ingreso, fecha_retiro, dias_laborados, tipo_retiro "
        "FROM retirado ORDER BY fecha_retiro DESC"
    )
    from collections import Counter
    tipo_counts = Counter(r.get("tipo_retiro", "") for r in rows)
    renuncia = tipo_counts.get("RENUNCIA VOLUNTARIA", 0) + tipo_counts.get("RENUNCIA", 0)
    terminacion = tipo_counts.get("TERMINACION DEL CONTRATO", 0) + tipo_counts.get("TERMINACION", 0)
    pendiente = tipo_counts.get("PENDIENTE", 0)
    columns = [
        {"key": "id_cedula",        "label": "Cédula"},
        {"key": "apellidos_nombre", "label": "Nombre"},
        {"key": "departamento",     "label": "Departamento", "type": "dept"},
        {"key": "area",             "label": "Área"},
        {"key": "fecha_ingreso",    "label": "Ingreso"},
        {"key": "fecha_retiro",     "label": "Retiro"},
        {"key": "dias_laborados",   "label": "Días", "type": "number"},
        {"key": "tipo_retiro",      "label": "Tipo Retiro", "type": "retiro"},
    ]
    stats = [
        {"value": len(rows),  "label": "Total Retirados", "icon": "person_remove", "color": "red"},
        {"value": renuncia,   "label": "Renuncias",       "icon": "directions_walk", "color": "orange"},
        {"value": terminacion,"label": "Terminaciones",   "icon": "gavel",         "color": "blue"},
        {"value": pendiente,  "label": "Pendientes",      "icon": "pending",       "color": "purple"},
    ]
    filter_columns = [
        {"index": 2, "label": "Departamento"},
        {"index": 7, "label": "Tipo Retiro"},
    ]
    return render_template(
        "data_table.html", active_page="Retiro de Personal",
        rows=rows, columns=columns, stats=stats,
        export_key="retiro_personal", filter_columns=filter_columns,
    )


# ── VIEW ÁREAS (módulo completo) ──────────────────────────────

@app.route("/areas")
@login_required
@module_required("organizacion")
def areas():
    area_rows = query(
        "SELECT a.id, a.nombre AS area, a.presupuestados, d.id AS depto_id, d.nombre AS departamento "
        "FROM area a JOIN departamento d ON a.departamento_id = d.id "
        "ORDER BY d.nombre, a.nombre"
    )
    emp_counts = query(
        "SELECT area, COUNT(*) AS cnt FROM empleado WHERE estado = 'ACTIVO' GROUP BY area"
    )
    count_map = {r["area"]: int(r["cnt"]) for r in emp_counts}

    from collections import OrderedDict
    grouped = OrderedDict()
    total_pres = total_ejec = 0
    for a in area_rows:
        depto = a["departamento"]
        if depto not in grouped:
            grouped[depto] = {"areas": [], "pres": 0, "ejec": 0, "depto_id": a.get("depto_id", 0)}
        pres = int(a["presupuestados"]) if a["presupuestados"] else 0
        ejec = count_map.get(a["area"], 0)
        pend = pres - ejec
        grouped[depto]["areas"].append({
            "id": a["id"],
            "area": a["area"],
            "presupuestados": pres if pres else "",
            "ejecutados": ejec,
            "pendientes": pend,
        })
        grouped[depto]["pres"] += pres
        grouped[depto]["ejec"] += ejec
        total_pres += pres
        total_ejec += ejec

    return render_template("areas_view.html", active_page="Area",
                           grouped=grouped, total=len(area_rows),
                           total_pres=total_pres, total_ejec=total_ejec)


@app.route("/areas/<int:area_id>")
@login_required
def area_detail(area_id):
    area = query(
        "SELECT a.id, a.nombre AS area, a.presupuestados, d.nombre AS departamento "
        "FROM area a JOIN departamento d ON a.departamento_id = d.id WHERE a.id = %s",
        (area_id,), one=True,
    )
    if not area:
        flash("Área no encontrada", "error")
        return redirect(url_for("areas"))

    pres = int(area["presupuestados"]) if area["presupuestados"] else 0

    perfiles = query(
        "SELECT p.id_perfil, p.perfil_ocupacional, d.nombre AS departamento, a.nombre AS area "
        "FROM perfil_ocupacional p "
        "JOIN area a ON p.area_id = a.id "
        "JOIN departamento d ON a.departamento_id = d.id "
        "WHERE a.id = %s ORDER BY p.perfil_ocupacional",
        (area_id,),
    )
    empleados = query(
        "SELECT id_cedula, apellidos_nombre, lugar_expedicion "
        "FROM empleado WHERE area = %s AND estado = 'ACTIVO' ORDER BY apellidos_nombre",
        (area["area"],),
    )
    retirados = query(
        "SELECT id_cedula, id_retiro, apellidos_nombre "
        "FROM retirado WHERE area = %s ORDER BY apellidos_nombre",
        (area["area"],),
    )

    ejec = len(empleados)
    pend = pres - ejec

    return render_template(
        "area_detail.html", active_page="Area",
        area=area, presupuestados=pres, ejecutados=ejec, pendientes=pend,
        perfiles=perfiles, empleados=empleados, retirados=retirados,
    )


@app.route("/areas/<int:area_id>/perfil/<perfil_id>")
@login_required
def perfil_detail(area_id, perfil_id):
    perfil = query(
        "SELECT p.id_perfil, p.perfil_ocupacional, p.presupuestados, "
        "d.nombre AS departamento, a.nombre AS area, a.id AS area_id "
        "FROM perfil_ocupacional p "
        "JOIN area a ON p.area_id = a.id "
        "JOIN departamento d ON a.departamento_id = d.id "
        "WHERE p.id_perfil = %s",
        (str(perfil_id),), one=True,
    )
    if not perfil:
        flash("Perfil no encontrado", "error")
        return redirect(url_for("area_detail", area_id=area_id))

    empleados = query(
        "SELECT id_cedula, apellidos_nombre, lugar_expedicion "
        "FROM empleado WHERE area = %s AND id_perfil_ocupacional = %s AND estado = 'ACTIVO' "
        "ORDER BY apellidos_nombre",
        (perfil["area"], str(perfil_id)),
    )
    retirados = query(
        "SELECT id_cedula, id_retiro, apellidos_nombre "
        "FROM retirado WHERE area = %s AND id_perfil_ocupacional = %s "
        "ORDER BY apellidos_nombre",
        (perfil["area"], str(perfil_id)),
    )

    return render_template(
        "perfil_detail.html", active_page="Area",
        perfil=perfil, area_id=area_id,
        empleados=empleados, retirados=retirados,
    )


@app.route("/areas/<int:area_id>/empleado/<cedula>")
@login_required
def area_empleado_detail(area_id, cedula):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
    if not emp:
        flash("Empleado no encontrado", "error")
        return redirect(url_for("area_detail", area_id=area_id))

    area = query("SELECT a.nombre AS area FROM area a WHERE a.id = %s", (area_id,), one=True)
    area_nombre = area["area"] if area else ""

    hijos = query("SELECT * FROM hijo WHERE id_cedula = %s ORDER BY fecha_nacimiento", (cedula,))
    retirados = query(
        "SELECT id_cedula, id_retiro, apellidos_nombre, departamento, area, "
        "id_perfil_ocupacional, fecha_ingreso, fecha_retiro, dias_laborados, tipo_retiro "
        "FROM retirado WHERE id_cedula = %s", (cedula,),
    )

    perfil_nombre = ""
    if emp.get("id_perfil_ocupacional"):
        p = query(
            "SELECT perfil_ocupacional FROM perfil_ocupacional WHERE id_perfil = %s",
            (str(emp["id_perfil_ocupacional"]),), one=True,
        )
        if p:
            perfil_nombre = p["perfil_ocupacional"]

    bd = parse_fecha(emp.get("fecha_nacimiento", ""))
    fi = parse_fecha(emp.get("fecha_ingreso", ""))
    today = date.today()
    mes_cumple = bd.month if bd else ""
    cumple_str = ""
    if bd:
        try:
            cumple_str = date(today.year, bd.month, bd.day).strftime("%d/%m/%Y")
        except ValueError:
            pass
    aniv_str = ""
    antiguedad = ""
    if fi:
        try:
            aniv_str = date(today.year, fi.month, fi.day).strftime("%d/%m/%Y")
        except ValueError:
            pass
        antiguedad = today.year - fi.year

    return render_template(
        "area_empleado_detail.html", active_page="Area",
        emp=emp, area_id=area_id, area_nombre=area_nombre,
        perfil_nombre=perfil_nombre,
        hijos=hijos, retirados=retirados,
        mes_cumple=mes_cumple, cumple_str=cumple_str,
        aniv_str=aniv_str, antiguedad=antiguedad,
    )


# ── DEPARTAMENTOS ─────────────────────────────────────────────

@app.route("/departamentos")
@login_required
@module_required("organizacion")
def departamentos():
    rows = query("SELECT * FROM departamento ORDER BY nombre")
    total_pres = sum(int(r["presupuestados"] or 0) for r in rows)
    columns = [
        {"key": "nombre",         "label": "Departamento"},
        {"key": "presupuestados", "label": "Presupuestados", "type": "number"},
    ]
    stats = [
        {"value": len(rows),   "label": "Departamentos",    "icon": "business",      "color": "green"},
        {"value": total_pres,  "label": "Presupuestados",   "icon": "group_add",     "color": "blue"},
    ]
    return render_template(
        "departamentos.html", active_page="Departamentos",
        rows=rows, columns=columns, stats=stats,
        export_key="departamentos",
    )


# ── PERFIL OCUPACIONAL ────────────────────────────────────────

@app.route("/perfil-ocupacional")
@login_required
@module_required("organizacion")
def perfil_ocupacional():
    rows = query(
        "SELECT p.id_perfil, d.id AS departamento_id, d.nombre AS departamento, "
        "a.id AS area_id, a.nombre AS area, p.perfil_ocupacional, p.presupuestados "
        "FROM perfil_ocupacional p "
        "JOIN area a ON p.area_id = a.id "
        "JOIN departamento d ON a.departamento_id = d.id "
        "ORDER BY d.nombre, a.nombre, p.perfil_ocupacional"
    )
    total_pres = sum(int(r["presupuestados"] or 0) for r in rows)
    deptos = len(set(r["departamento"] for r in rows if r.get("departamento")))
    columns = [
        {"key": "departamento",      "label": "Departamento", "type": "dept"},
        {"key": "area",              "label": "Área"},
        {"key": "perfil_ocupacional","label": "Perfil Ocupacional"},
        {"key": "presupuestados",    "label": "Presupuestados", "type": "number"},
    ]
    stats = [
        {"value": len(rows),  "label": "Perfiles",         "icon": "badge",     "color": "green"},
        {"value": deptos,     "label": "Departamentos",    "icon": "business",  "color": "blue"},
        {"value": total_pres, "label": "Presupuestados",   "icon": "group_add", "color": "orange"},
    ]
    filter_columns = [
        {"index": 0, "label": "Departamento"},
        {"index": 1, "label": "Área"},
    ]
    departamentos_list = query("SELECT id, nombre FROM departamento ORDER BY nombre")
    areas_list = query(
        "SELECT a.id, a.nombre, a.departamento_id FROM area a ORDER BY a.nombre"
    )
    return render_template(
        "perfil_ocupacional.html", active_page="Perfil Ocupacional",
        rows=rows, columns=columns, stats=stats,
        export_key="perfil_ocupacional", filter_columns=filter_columns,
        departamentos_list=departamentos_list, areas_list=areas_list,
    )


# ── VIEW EPS (módulo completo) ────────────────────────────────

@app.route("/view-eps")
@login_required
@module_required("eps")
def view_eps():
    rows = query(
        "SELECT eps, COUNT(*) AS cnt "
        "FROM empleado WHERE eps IS NOT NULL AND eps != '' "
        "GROUP BY eps ORDER BY eps"
    )
    eps_list = []
    for r in rows:
        eps_list.append({
            "eps": r["eps"],
            "related": f"Related DBases ({r['cnt']})",
            "cnt": int(r["cnt"]),
        })
    return render_template("eps_view.html", active_page="EPS", eps_list=eps_list)


@app.route("/view-eps/add", methods=["POST"])
@login_required
@module_required("eps")
@role_required("WRITE")
def eps_add():
    nombre = (request.form.get("nombre") or "").strip().upper()
    if not nombre:
        flash("El nombre de la EPS es obligatorio", "error")
        return redirect(url_for("view_eps"))
    existing = query("SELECT id FROM eps WHERE nombre = %s", (nombre,), one=True)
    if existing:
        flash(f"La EPS '{nombre}' ya existe", "error")
        return redirect(url_for("view_eps"))
    execute("INSERT INTO eps (nombre) VALUES (%s)", (nombre,))
    flash(f"EPS '{nombre}' creada exitosamente", "success")
    return redirect(url_for("view_eps"))


@app.route("/view-eps/<path:eps_name>")
@login_required
@module_required("eps")
def eps_detail(eps_name):
    empleados = query(
        "SELECT id_cedula, apellidos_nombre, lugar_expedicion, fecha_expedicion "
        "FROM empleado WHERE eps = %s ORDER BY apellidos_nombre",
        (eps_name,),
    )
    return render_template(
        "eps_detail.html", active_page="EPS",
        eps_name=eps_name, empleados=empleados,
    )


@app.route("/view-eps/<path:eps_name>/inline")
@login_required
@module_required("eps")
def eps_inline(eps_name):
    empleados = query(
        "SELECT * FROM empleado WHERE eps = %s ORDER BY apellidos_nombre",
        (eps_name,),
    )
    empleados = enrich_empleados(empleados)
    return render_template(
        "inline_view.html", active_page="EPS",
        context_label="EPS", context_name=eps_name,
        back_url=url_for("eps_detail", eps_name=eps_name),
        list_url=url_for("view_eps"),
        empleados=empleados, columns=INLINE_COLUMNS,
        export_url=url_for("eps_export", eps_name=eps_name),
        detail_route="eps_empleado_detail",
        detail_kwargs={"eps_name": eps_name},
    )


@app.route("/view-eps/<path:eps_name>/export")
@login_required
def eps_export(eps_name):
    empleados = query(
        "SELECT * FROM empleado WHERE eps = %s ORDER BY apellidos_nombre",
        (eps_name,),
    )
    empleados = enrich_empleados(empleados)
    return export_excel_response(empleados, f"EPS_{eps_name}")


@app.route("/view-eps/<path:eps_name>/empleado/<cedula>")
@login_required
def eps_empleado_detail(eps_name, cedula):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
    if not emp:
        flash("Empleado no encontrado", "error")
        return redirect(url_for("eps_detail", eps_name=eps_name))

    perfil_nombre = ""
    if emp.get("id_perfil_ocupacional"):
        p = query(
            "SELECT perfil_ocupacional FROM perfil_ocupacional WHERE id_perfil = %s",
            (str(emp["id_perfil_ocupacional"]),), one=True,
        )
        if p:
            perfil_nombre = p["perfil_ocupacional"]

    hijos = query("SELECT * FROM hijo WHERE id_cedula = %s ORDER BY fecha_nacimiento", (cedula,))
    retirados = query(
        "SELECT id_cedula, id_retiro, apellidos_nombre, departamento, area, "
        "id_perfil_ocupacional, fecha_ingreso, fecha_retiro, dias_laborados, tipo_retiro "
        "FROM retirado WHERE id_cedula = %s", (cedula,),
    )

    bd = parse_fecha(emp.get("fecha_nacimiento", ""))
    fi = parse_fecha(emp.get("fecha_ingreso", ""))
    today = date.today()
    mes_cumple = bd.month if bd else ""
    cumple_str = ""
    if bd:
        try:
            cumple_str = date(today.year, bd.month, bd.day).strftime("%d/%m/%Y")
        except ValueError:
            pass
    aniv_str = ""
    antiguedad = ""
    if fi:
        try:
            aniv_str = date(today.year, fi.month, fi.day).strftime("%d/%m/%Y")
        except ValueError:
            pass
        antiguedad = today.year - fi.year

    return render_template(
        "eps_empleado_detail.html", active_page="EPS",
        emp=emp, eps_name=eps_name, perfil_nombre=perfil_nombre,
        hijos=hijos, retirados=retirados,
        mes_cumple=mes_cumple, cumple_str=cumple_str,
        aniv_str=aniv_str, antiguedad=antiguedad,
    )


@app.route("/view-eps/<path:eps_name>/empleado/<cedula>/retirado/<retiro_id>")
@login_required
def eps_retirado_detail(eps_name, cedula, retiro_id):
    ret = query(
        "SELECT * FROM retirado WHERE id_retiro = %s", (retiro_id,), one=True,
    )
    if not ret:
        flash("Registro de retiro no encontrado", "error")
        return redirect(url_for("eps_empleado_detail", eps_name=eps_name, cedula=cedula))

    emp_nombre = ""
    emp = query("SELECT apellidos_nombre FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
    if emp:
        emp_nombre = emp["apellidos_nombre"]

    return render_template(
        "eps_retirado_detail.html", active_page="EPS",
        ret=ret, eps_name=eps_name, cedula=cedula, emp_nombre=emp_nombre,
    )


# ── VIEW FONDO DE PENSIONES (módulo completo) ────────────────

@app.route("/view-fondos")
@login_required
@module_required("fondos")
def view_fondos():
    rows = query(
        "SELECT fondo_pensiones, COUNT(*) AS cnt "
        "FROM empleado WHERE fondo_pensiones IS NOT NULL AND fondo_pensiones != '' "
        "GROUP BY fondo_pensiones ORDER BY fondo_pensiones"
    )
    fondos_list = []
    for r in rows:
        fondos_list.append({
            "fondo": r["fondo_pensiones"],
            "cnt": int(r["cnt"]),
        })
    return render_template("fondos_view.html", active_page="Fondo de Pensiones",
                           fondos_list=fondos_list)


@app.route("/view-fondos/add", methods=["POST"])
@login_required
@module_required("fondos")
@role_required("WRITE")
def fondos_add():
    nombre = (request.form.get("nombre") or "").strip().upper()
    if not nombre:
        flash("El nombre del Fondo es obligatorio", "error")
        return redirect(url_for("view_fondos"))
    existing = query("SELECT id FROM fondo_pensiones WHERE nombre = %s", (nombre,), one=True)
    if existing:
        flash(f"El Fondo '{nombre}' ya existe", "error")
        return redirect(url_for("view_fondos"))
    execute("INSERT INTO fondo_pensiones (nombre) VALUES (%s)", (nombre,))
    flash(f"Fondo '{nombre}' creado exitosamente", "success")
    return redirect(url_for("view_fondos"))


@app.route("/view-fondos/<path:fondo_name>")
@login_required
@module_required("fondos")
def fondo_detail(fondo_name):
    empleados = query(
        "SELECT id_cedula, apellidos_nombre, lugar_expedicion, fecha_expedicion "
        "FROM empleado WHERE fondo_pensiones = %s ORDER BY apellidos_nombre",
        (fondo_name,),
    )
    return render_template(
        "fondo_detail.html", active_page="Fondo de Pensiones",
        fondo_name=fondo_name, empleados=empleados,
    )


@app.route("/view-fondos/<path:fondo_name>/inline")
@login_required
@module_required("fondos")
def fondo_inline(fondo_name):
    empleados = query(
        "SELECT * FROM empleado WHERE fondo_pensiones = %s ORDER BY apellidos_nombre",
        (fondo_name,),
    )
    empleados = enrich_empleados(empleados)
    return render_template(
        "inline_view.html", active_page="Fondo de Pensiones",
        context_label="Fondo de Pensiones", context_name=fondo_name,
        back_url=url_for("fondo_detail", fondo_name=fondo_name),
        list_url=url_for("view_fondos"),
        empleados=empleados, columns=INLINE_COLUMNS,
        export_url=url_for("fondo_export", fondo_name=fondo_name),
        detail_route="fondo_empleado_detail",
        detail_kwargs={"fondo_name": fondo_name},
    )


@app.route("/view-fondos/<path:fondo_name>/export")
@login_required
def fondo_export(fondo_name):
    empleados = query(
        "SELECT * FROM empleado WHERE fondo_pensiones = %s ORDER BY apellidos_nombre",
        (fondo_name,),
    )
    empleados = enrich_empleados(empleados)
    return export_excel_response(empleados, f"Fondo_{fondo_name}")


@app.route("/view-fondos/<path:fondo_name>/empleado/<cedula>")
@login_required
def fondo_empleado_detail(fondo_name, cedula):
    emp = query("SELECT * FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
    if not emp:
        flash("Empleado no encontrado", "error")
        return redirect(url_for("fondo_detail", fondo_name=fondo_name))

    perfil_nombre = ""
    if emp.get("id_perfil_ocupacional"):
        p = query(
            "SELECT perfil_ocupacional FROM perfil_ocupacional WHERE id_perfil = %s",
            (str(emp["id_perfil_ocupacional"]),), one=True,
        )
        if p:
            perfil_nombre = p["perfil_ocupacional"]

    hijos = query("SELECT * FROM hijo WHERE id_cedula = %s ORDER BY fecha_nacimiento", (cedula,))
    retirados = query(
        "SELECT id_cedula, id_retiro, apellidos_nombre, departamento, area, "
        "id_perfil_ocupacional, fecha_ingreso, fecha_retiro, dias_laborados, tipo_retiro "
        "FROM retirado WHERE id_cedula = %s", (cedula,),
    )

    bd = parse_fecha(emp.get("fecha_nacimiento", ""))
    fi = parse_fecha(emp.get("fecha_ingreso", ""))
    today = date.today()
    mes_cumple = bd.month if bd else ""
    cumple_str = ""
    if bd:
        try:
            cumple_str = date(today.year, bd.month, bd.day).strftime("%d/%m/%Y")
        except ValueError:
            pass
    aniv_str = ""
    antiguedad = ""
    if fi:
        try:
            aniv_str = date(today.year, fi.month, fi.day).strftime("%d/%m/%Y")
        except ValueError:
            pass
        antiguedad = today.year - fi.year

    return render_template(
        "fondo_empleado_detail.html", active_page="Fondo de Pensiones",
        emp=emp, fondo_name=fondo_name, perfil_nombre=perfil_nombre,
        hijos=hijos, retirados=retirados,
        mes_cumple=mes_cumple, cumple_str=cumple_str,
        aniv_str=aniv_str, antiguedad=antiguedad,
    )


@app.route("/view-fondos/<path:fondo_name>/empleado/<cedula>/retirado/<retiro_id>")
@login_required
def fondo_retirado_detail(fondo_name, cedula, retiro_id):
    ret = query("SELECT * FROM retirado WHERE id_retiro = %s", (retiro_id,), one=True)
    if not ret:
        flash("Registro de retiro no encontrado", "error")
        return redirect(url_for("fondo_empleado_detail", fondo_name=fondo_name, cedula=cedula))

    emp_nombre = ""
    emp = query("SELECT apellidos_nombre FROM empleado WHERE id_cedula = %s", (cedula,), one=True)
    if emp:
        emp_nombre = emp["apellidos_nombre"]

    return render_template(
        "fondo_retirado_detail.html", active_page="Fondo de Pensiones",
        ret=ret, fondo_name=fondo_name, cedula=cedula, emp_nombre=emp_nombre,
    )


# ── REACTIVAR EMPLEADO ───────────────────────────────────────

@app.route("/personal-activo/<id>/reactivar", methods=["POST"])
@login_required
@module_required("personal")
@role_required("WRITE")
def reactivar_empleado(id):
    execute("UPDATE empleado SET estado = 'ACTIVO' WHERE id_cedula = %s", (id,))
    flash("Empleado reactivado exitosamente", "success")
    return redirect(url_for("detalle_empleado", id=id))


# ── CRUD RETIRO (editar) ──────────────────────────────────────

@app.route("/retirado/<id>/editar", methods=["GET", "POST"])
@login_required
@module_required("retiro")
@role_required("WRITE")
def editar_retirado(id):
    ret = query("SELECT * FROM retirado WHERE id_retiro = %s", (id,), one=True)
    if not ret:
        flash("Registro no encontrado", "error")
        return redirect(url_for("retiro_personal"))
    if request.method == "POST":
        fecha_retiro = request.form.get("fecha_retiro", "").strip()
        tipo_retiro = request.form.get("tipo_retiro", "").strip()
        dias_laborados = request.form.get("dias_laborados", "").strip()
        motivo = request.form.get("motivo", "").strip()
        if "-" in fecha_retiro:
            parts = fecha_retiro.split("-")
            fecha_retiro = f"{parts[2]}/{parts[1]}/{parts[0]}"
        execute(
            "UPDATE retirado SET fecha_retiro=%s, tipo_retiro=%s, dias_laborados=%s, motivo=%s "
            "WHERE id_retiro=%s",
            (fecha_retiro or None,
             tipo_retiro or None,
             int(dias_laborados) if dias_laborados else None,
             motivo or None,
             id),
        )
        flash("Registro de retiro actualizado", "success")
        redirect_cedula = request.form.get("redirect_cedula", "")
        if redirect_cedula:
            return redirect(url_for("detalle_empleado", id=redirect_cedula))
        return redirect(url_for("retiro_personal"))
    motivos = query("SELECT tipo_retiro FROM motivo_retiro ORDER BY tipo_retiro")
    redirect_cedula = request.args.get("redirect_cedula", ret.get("id_cedula", ""))
    return render_template(
        "retirado_edit_form.html", active_page="Retiro de Personal",
        ret=ret, motivos=motivos,
        back_url=url_for("detalle_empleado", id=redirect_cedula) if redirect_cedula else url_for("retiro_personal"),
        redirect_cedula=redirect_cedula,
    )


# ── API helpers para cascadas ────────────────────────────────

@app.route("/api/areas-por-depto/<int:depto_id>")
@login_required
def api_areas_por_depto(depto_id):
    rows = query(
        "SELECT id, nombre FROM area WHERE departamento_id=%s ORDER BY nombre", (depto_id,)
    )
    return jsonify([{"id": r["id"], "nombre": r["nombre"]} for r in rows])


# ── CRUD DEPARTAMENTOS ────────────────────────────────────────

@app.route("/departamentos/nuevo", methods=["POST"])
@login_required
@role_required("ALL")
def departamento_nuevo():
    nombre = (request.form.get("nombre") or "").strip().upper()
    presupuestados = request.form.get("presupuestados", "").strip()
    if not nombre:
        flash("El nombre del departamento es obligatorio", "error")
        return redirect(url_for("departamentos"))
    existing = query("SELECT id FROM departamento WHERE nombre = %s", (nombre,), one=True)
    if existing:
        flash(f"Ya existe el departamento '{nombre}'", "error")
        return redirect(url_for("departamentos"))
    execute(
        "INSERT INTO departamento (nombre, presupuestados) VALUES (%s, %s)",
        (nombre, int(presupuestados) if presupuestados else None),
    )
    flash(f"Departamento '{nombre}' creado", "success")
    return redirect(url_for("departamentos"))


@app.route("/departamentos/<int:id>/editar", methods=["POST"])
@login_required
@role_required("ALL")
def departamento_editar(id):
    nombre = (request.form.get("nombre") or "").strip().upper()
    presupuestados = request.form.get("presupuestados", "").strip()
    execute(
        "UPDATE departamento SET nombre=%s, presupuestados=%s WHERE id=%s",
        (nombre, int(presupuestados) if presupuestados else None, id),
    )
    flash("Departamento actualizado", "success")
    return redirect(url_for("departamentos"))


@app.route("/departamentos/<int:id>/eliminar", methods=["POST"])
@login_required
@role_required("ALL")
def departamento_eliminar(id):
    try:
        execute("DELETE FROM departamento WHERE id = %s", (id,))
        flash("Departamento eliminado", "success")
    except Exception:
        flash("No se puede eliminar: tiene áreas asociadas", "error")
    return redirect(url_for("departamentos"))


# ── CRUD ÁREAS ────────────────────────────────────────────────

@app.route("/areas/nueva", methods=["POST"])
@login_required
@role_required("ALL")
def area_nueva():
    nombre = (request.form.get("nombre") or "").strip().upper()
    departamento_id = request.form.get("departamento_id", "").strip()
    presupuestados = request.form.get("presupuestados", "").strip()
    if not nombre or not departamento_id:
        flash("Nombre y Departamento son obligatorios", "error")
        return redirect(url_for("areas"))
    execute(
        "INSERT INTO area (nombre, departamento_id, presupuestados) VALUES (%s, %s, %s)",
        (nombre, int(departamento_id), int(presupuestados) if presupuestados else None),
    )
    flash(f"Área '{nombre}' creada", "success")
    return redirect(url_for("areas"))


@app.route("/areas/<int:id>/editar", methods=["POST"])
@login_required
@role_required("ALL")
def area_editar(id):
    nombre = (request.form.get("nombre") or "").strip().upper()
    departamento_id = request.form.get("departamento_id", "").strip()
    presupuestados = request.form.get("presupuestados", "").strip()
    execute(
        "UPDATE area SET nombre=%s, departamento_id=%s, presupuestados=%s WHERE id=%s",
        (nombre, int(departamento_id), int(presupuestados) if presupuestados else None, id),
    )
    flash("Área actualizada", "success")
    return redirect(url_for("areas"))


@app.route("/areas/<int:id>/eliminar", methods=["POST"])
@login_required
@role_required("ALL")
def area_eliminar(id):
    try:
        execute("DELETE FROM area WHERE id = %s", (id,))
        flash("Área eliminada", "success")
    except Exception:
        flash("No se puede eliminar: tiene perfiles o empleados asociados", "error")
    return redirect(url_for("areas"))


# ── CRUD PERFIL OCUPACIONAL ───────────────────────────────────

@app.route("/perfil/nuevo", methods=["POST"])
@login_required
@role_required("ALL")
def perfil_nuevo():
    import uuid
    perfil = (request.form.get("perfil_ocupacional") or "").strip().upper()
    area_id = request.form.get("area_id", "").strip()
    presupuestados = request.form.get("presupuestados", "").strip()
    if not perfil or not area_id:
        flash("Perfil y Área son obligatorios", "error")
        return redirect(url_for("perfil_ocupacional"))
    new_id = uuid.uuid4().hex[:8]
    execute(
        "INSERT INTO perfil_ocupacional (id_perfil, area_id, perfil_ocupacional, presupuestados) "
        "VALUES (%s, %s, %s, %s)",
        (new_id, int(area_id), perfil, int(presupuestados) if presupuestados else None),
    )
    flash(f"Perfil '{perfil}' creado", "success")
    return redirect(url_for("perfil_ocupacional"))


@app.route("/perfil/<id>/editar", methods=["POST"])
@login_required
@role_required("ALL")
def perfil_editar(id):
    perfil = (request.form.get("perfil_ocupacional") or "").strip().upper()
    area_id = request.form.get("area_id", "").strip()
    presupuestados = request.form.get("presupuestados", "").strip()
    execute(
        "UPDATE perfil_ocupacional SET perfil_ocupacional=%s, area_id=%s, presupuestados=%s "
        "WHERE id_perfil=%s",
        (perfil, int(area_id), int(presupuestados) if presupuestados else None, id),
    )
    flash("Perfil actualizado", "success")
    return redirect(url_for("perfil_ocupacional"))


@app.route("/perfil/<id>/eliminar", methods=["POST"])
@login_required
@role_required("ALL")
def perfil_eliminar(id):
    execute("DELETE FROM perfil_ocupacional WHERE id_perfil = %s", (id,))
    flash("Perfil eliminado", "success")
    return redirect(url_for("perfil_ocupacional"))


# ── CRUD EPS ──────────────────────────────────────────────────

@app.route("/view-eps/<path:eps_name>/editar", methods=["POST"])
@login_required
@role_required("ALL")
def eps_editar(eps_name):
    nuevo_nombre = (request.form.get("nombre") or "").strip().upper()
    if not nuevo_nombre:
        flash("El nombre no puede estar vacío", "error")
        return redirect(url_for("view_eps"))
    execute("UPDATE eps SET nombre=%s WHERE nombre=%s", (nuevo_nombre, eps_name))
    execute("UPDATE empleado SET eps=%s WHERE eps=%s", (nuevo_nombre, eps_name))
    flash(f"EPS renombrada a '{nuevo_nombre}'", "success")
    return redirect(url_for("view_eps"))


@app.route("/view-eps/<path:eps_name>/eliminar", methods=["POST"])
@login_required
@role_required("ALL")
def eps_eliminar(eps_name):
    execute("DELETE FROM eps WHERE nombre = %s", (eps_name,))
    flash(f"EPS '{eps_name}' eliminada", "success")
    return redirect(url_for("view_eps"))


# ── CRUD FONDO DE PENSIONES ───────────────────────────────────

@app.route("/view-fondos/<path:fondo_name>/editar", methods=["POST"])
@login_required
@role_required("ALL")
def fondo_editar(fondo_name):
    nuevo_nombre = (request.form.get("nombre") or "").strip().upper()
    if not nuevo_nombre:
        flash("El nombre no puede estar vacío", "error")
        return redirect(url_for("view_fondos"))
    execute("UPDATE fondo_pensiones SET nombre=%s WHERE nombre=%s", (nuevo_nombre, fondo_name))
    execute("UPDATE empleado SET fondo_pensiones=%s WHERE fondo_pensiones=%s", (nuevo_nombre, fondo_name))
    flash(f"Fondo renombrado a '{nuevo_nombre}'", "success")
    return redirect(url_for("view_fondos"))


@app.route("/view-fondos/<path:fondo_name>/eliminar", methods=["POST"])
@login_required
@role_required("ALL")
def fondo_eliminar(fondo_name):
    execute("DELETE FROM fondo_pensiones WHERE nombre = %s", (fondo_name,))
    flash(f"Fondo '{fondo_name}' eliminado", "success")
    return redirect(url_for("view_fondos"))


# ── USUARIOS (ADMIN full; COORD. GH ver+inactivar; GESTOR CONTRATACIÓN ver+crear+editar roles+inactivar) ───

_ROLES_CAN_CREATE_EDIT_RESET = ("ADMIN", "GESTOR DE CONTRATACION")
_ROLES_CAN_TOGGLE = ("ADMIN", "COORD. GH", "GESTOR DE CONTRATACION")


def _user_management_allowed(action):
    """action in ('create', 'edit', 'toggle', 'reset')."""
    user = get_current_user()
    if not user:
        return False
    rol = (user.get("rol") or "").strip()
    if action == "toggle":
        return rol in _ROLES_CAN_TOGGLE
    if action in ("create", "edit", "reset"):
        return rol in _ROLES_CAN_CREATE_EDIT_RESET
    return False


def _validar_password(password):
    """Exige mínimo 8 caracteres, al menos una letra y un número. Devuelve (True, None) o (False, mensaje)."""
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    if not any(c.isalpha() for c in password):
        return False, "La contraseña debe contener al menos una letra"
    if not any(c.isdigit() for c in password):
        return False, "La contraseña debe contener al menos un número"
    return True, None


@app.route("/users/nuevo", methods=["POST"])
@login_required
@module_required("admin_usuarios")
def usuario_nuevo():
    if not _user_management_allowed("create"):
        flash("No tienes permiso para crear usuarios.", "error")
        return redirect(url_for("usuarios"))
    nombre   = (request.form.get("nombre")   or "").strip().upper()
    email    = (request.form.get("email")    or "").strip().lower()
    password = (request.form.get("password") or "").strip()
    confirm  = (request.form.get("confirmar_password") or "").strip()
    rol      = (request.form.get("rol")      or "").strip()

    if not all([nombre, email, password, rol]):
        flash("Todos los campos son obligatorios", "error")
        return redirect(url_for("usuarios"))

    if rol == "ADMIN":
        flash("No se puede crear un usuario con rol ADMIN. Ese rol está reservado.", "error")
        return redirect(url_for("usuarios"))

    if password != confirm:
        flash("La contraseña y su confirmación no coinciden", "error")
        return redirect(url_for("usuarios"))

    ok, msg = _validar_password(password)
    if not ok:
        flash(msg, "error")
        return redirect(url_for("usuarios"))

    existing = query("SELECT id_user FROM usuario WHERE LOWER(email) = %s", (email,), one=True)
    if existing:
        flash(f"Ya existe un usuario con el correo {email}", "error")
        return redirect(url_for("usuarios"))

    last = query("SELECT id_user FROM usuario ORDER BY id_user DESC LIMIT 1", one=True)
    num = (int(last["id_user"].replace("US-", "")) + 1) if last else 1
    new_id = f"US-{num:04d}"

    try:
        execute(
            "INSERT INTO usuario (id_user, email, password_hash, nombre, rol, estado, acciones, debe_cambiar_clave) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, 1)",
            (new_id, email, generate_password_hash(password),
             nombre, rol, True, get_acciones_for_rol(rol)),
        )
    except Exception as e:
        if "debe_cambiar_clave" in str(e):
            execute(
                "INSERT INTO usuario (id_user, email, password_hash, nombre, rol, estado, acciones) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (new_id, email, generate_password_hash(password),
                 nombre, rol, True, get_acciones_for_rol(rol)),
            )
        else:
            raise
    registrar_audit("Usuario creado", "admin", f"id={new_id} rol={rol}")
    flash(f"Usuario {nombre} creado. Deberá cambiar la contraseña al ingresar por primera vez.", "success")
    return redirect(url_for("usuarios"))


_ROLE_DESCRIPTIONS = {
    "ADMIN": "Administrador del sistema",
    "COORD. GH": "Coordinación Gestión Humana",
    "EMPLEADO": "Empleado (portal y solicitudes)",
    "GESTOR DE CONTRATACION": "Gestor de Contratación",
    "GESTOR DE NOMINA": "Gestor de Nómina",
    "GESTOR SST": "Gestor SST",
    "BIENESTAR SOCIAL": "Bienestar Social",
}


@app.route("/users")
@login_required
@module_required("admin_usuarios")
def usuarios():
    roles = query("SELECT nombre FROM rol ORDER BY nombre")
    rows = query(
        "SELECT id_user, email, nombre, rol, estado, acciones FROM usuario ORDER BY id_user"
    )
    activos = sum(1 for r in rows if r.get("estado"))
    roles_con_desc = [{"nombre": r["nombre"], "desc": _ROLE_DESCRIPTIONS.get(r["nombre"], "")} for r in roles]
    roles_crear = [r for r in roles_con_desc if r["nombre"] != "ADMIN"]
    admin_email = current_app.config.get("ADMIN_EMAIL", "tecnologia@colbeef.com").strip().lower()
    user = get_current_user()
    rol = (user.get("rol") or "").strip()
    can_create = rol in _ROLES_CAN_CREATE_EDIT_RESET
    can_edit = rol in _ROLES_CAN_CREATE_EDIT_RESET
    can_toggle = rol in _ROLES_CAN_TOGGLE
    can_reset = rol in _ROLES_CAN_CREATE_EDIT_RESET
    columns = [
        {"key": "id_user",  "label": "ID"},
        {"key": "email",    "label": "Email"},
        {"key": "nombre",   "label": "Nombre"},
        {"key": "rol",      "label": "Rol"},
        {"key": "estado",   "label": "Estado", "type": "badge"},
        {"key": "acciones", "label": "Acciones"},
    ]
    stats = [
        {"value": len(rows), "label": "Total Usuarios",  "icon": "manage_accounts", "color": "green"},
        {"value": activos,   "label": "Activos",         "icon": "verified_user",   "color": "blue"},
    ]
    return render_template(
        "users.html", active_page="User",
        rows=rows, columns=columns, stats=stats, roles=roles, roles_con_desc=roles_con_desc,
        roles_crear=roles_crear, admin_email=admin_email,
        can_create=can_create, can_edit=can_edit, can_toggle=can_toggle, can_reset=can_reset,
    )


@app.route("/users/<id>/editar", methods=["POST"])
@login_required
@module_required("admin_usuarios")
def usuario_editar(id):
    if not _user_management_allowed("edit"):
        flash("No tienes permiso para editar usuarios.", "error")
        return redirect(url_for("usuarios"))
    nombre = (request.form.get("nombre") or "").strip().upper()
    email = (request.form.get("email") or "").strip().lower()
    rol = (request.form.get("rol") or "").strip()
    admin_email = current_app.config.get("ADMIN_EMAIL", "tecnologia@colbeef.com").strip().lower()

    # Solo el usuario con ADMIN_EMAIL puede tener rol ADMIN
    if rol == "ADMIN" and email != admin_email:
        flash("El rol ADMIN solo puede asignarse al usuario autorizado (tecnología).", "error")
        return redirect(url_for("usuarios"))

    # No se puede quitar ADMIN al usuario autorizado
    actual = query("SELECT email, rol FROM usuario WHERE id_user = %s", (id,), one=True)
    if actual and (actual.get("email") or "").strip().lower() == admin_email and rol != "ADMIN":
        flash("No se puede quitar el rol ADMIN a la cuenta de tecnología.", "error")
        return redirect(url_for("usuarios"))

    execute(
        "UPDATE usuario SET nombre=%s, email=%s, rol=%s, acciones=%s WHERE id_user=%s",
        (nombre, email, rol, get_acciones_for_rol(rol), id),
    )
    flash("Usuario actualizado", "success")
    return redirect(url_for("usuarios"))


@app.route("/users/<id>/toggle-estado", methods=["POST"])
@login_required
@module_required("admin_usuarios")
def usuario_toggle_estado(id):
    if not _user_management_allowed("toggle"):
        flash("No tienes permiso para activar/desactivar usuarios.", "error")
        return redirect(url_for("usuarios"))
    user = query("SELECT estado, nombre FROM usuario WHERE id_user=%s", (id,), one=True)
    if not user:
        flash("Usuario no encontrado", "error")
        return redirect(url_for("usuarios"))
    nuevo = not user["estado"]
    execute("UPDATE usuario SET estado=%s WHERE id_user=%s", (nuevo, id))
    accion = "activado" if nuevo else "desactivado"
    flash(f"Usuario {user['nombre']} {accion}", "success")
    return redirect(url_for("usuarios"))


@app.route("/users/<id>/reset-password", methods=["POST"])
@login_required
@module_required("admin_usuarios")
def usuario_reset_password(id):
    if not _user_management_allowed("reset"):
        flash("No tienes permiso para restablecer contraseñas.", "error")
        return redirect(url_for("usuarios"))
    nueva = (request.form.get("nueva_password") or "").strip()
    confirm = (request.form.get("confirmar_password") or "").strip()
    if nueva != confirm:
        flash("La contraseña y su confirmación no coinciden", "error")
        return redirect(url_for("usuarios"))
    ok, msg = _validar_password(nueva)
    if not ok:
        flash(msg, "error")
        return redirect(url_for("usuarios"))
    try:
        execute(
            "UPDATE usuario SET password_hash=%s, debe_cambiar_clave=0 WHERE id_user=%s",
            (generate_password_hash(nueva), id),
        )
    except Exception as e:
        if "debe_cambiar_clave" in str(e):
            execute(
                "UPDATE usuario SET password_hash=%s WHERE id_user=%s",
                (generate_password_hash(nueva), id),
            )
        else:
            raise
    flash("Contraseña restablecida. El usuario puede ingresar con la nueva clave.", "success")
    return redirect(url_for("usuarios"))


@app.route("/users/<id>/reset-password-estandar", methods=["POST"])
@login_required
@module_required("admin_usuarios")
def usuario_reset_password_estandar(id):
    """Restablece la contraseña del usuario a la estándar; deberá cambiarla al ingresar."""
    if not _user_management_allowed("reset"):
        flash("No tienes permiso para restablecer contraseñas.", "error")
        return redirect(url_for("usuarios"))
    target = query("SELECT id_user, nombre FROM usuario WHERE id_user = %s", (id,), one=True)
    if not target:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("usuarios"))
    try:
        execute(
            "UPDATE usuario SET password_hash=%s, debe_cambiar_clave=1 WHERE id_user=%s",
            (generate_password_hash(PASSWORD_ESTANDAR), id),
        )
    except Exception as e:
        if "debe_cambiar_clave" in str(e):
            execute(
                "UPDATE usuario SET password_hash=%s WHERE id_user=%s",
                (generate_password_hash(PASSWORD_ESTANDAR), id),
            )
        else:
            raise
    registrar_audit("Contraseña restablecida a estándar", "admin", f"id={id}")
    flash(f"Contraseña de {target['nombre']} restablecida a la estándar. Deberá cambiarla al iniciar sesión.", "success")
    return redirect(url_for("usuarios"))


# ── CATÁLOGOS (tipo documento, nivel educativo, profesión, motivo retiro) ──

@app.route("/admin/catalogos")
@login_required
@module_required("admin")
@role_required("ALL")
@admin_only
def admin_catalogos():
    """Lista y permite agregar registros a catálogos usados en empleados y retiros."""
    tipo_doc = query("SELECT id_tipo_documento, tipo_documento FROM tipo_documento ORDER BY tipo_documento")
    niveles = query("SELECT id_nivel, nivel FROM nivel_educativo ORDER BY nivel")
    profesiones = query("SELECT id_profesion, profesion FROM profesion ORDER BY profesion")
    motivos = query("SELECT id, tipo_retiro FROM motivo_retiro ORDER BY tipo_retiro")
    return render_template(
        "catalogos.html", active_page="Catalogos",
        tipo_doc=tipo_doc, niveles=niveles, profesiones=profesiones, motivos=motivos,
    )


@app.route("/admin/catalogos/tipo-documento", methods=["POST"])
@login_required
@module_required("admin")
@role_required("ALL")
@admin_only
def catalogos_tipo_documento_add():
    sigla = (request.form.get("id_tipo_documento") or "").strip().upper()[:50]
    nombre = (request.form.get("tipo_documento") or "").strip()
    if not sigla or not nombre:
        flash("Sigla y nombre son obligatorios", "error")
        return redirect(url_for("admin_catalogos"))
    existing = query("SELECT id_tipo_documento FROM tipo_documento WHERE id_tipo_documento = %s", (sigla,), one=True)
    if existing:
        flash(f"Ya existe el tipo de documento con sigla {sigla}", "error")
        return redirect(url_for("admin_catalogos"))
    execute("INSERT INTO tipo_documento (id_tipo_documento, tipo_documento) VALUES (%s, %s)", (sigla, nombre))
    flash(f"Tipo de documento '{nombre}' agregado", "success")
    return redirect(url_for("admin_catalogos"))


@app.route("/admin/catalogos/nivel-educativo", methods=["POST"])
@login_required
@module_required("admin")
@role_required("ALL")
@admin_only
def catalogos_nivel_add():
    sigla = (request.form.get("id_nivel") or "").strip().upper()[:50]
    nombre = (request.form.get("nivel") or "").strip()
    if not sigla or not nombre:
        flash("Sigla y nombre son obligatorios", "error")
        return redirect(url_for("admin_catalogos"))
    existing = query("SELECT id_nivel FROM nivel_educativo WHERE id_nivel = %s", (sigla,), one=True)
    if existing:
        flash(f"Ya existe el nivel con sigla {sigla}", "error")
        return redirect(url_for("admin_catalogos"))
    execute("INSERT INTO nivel_educativo (id_nivel, nivel) VALUES (%s, %s)", (sigla, nombre))
    flash(f"Nivel educativo '{nombre}' agregado", "success")
    return redirect(url_for("admin_catalogos"))


@app.route("/admin/catalogos/profesion", methods=["POST"])
@login_required
@module_required("admin")
@role_required("ALL")
@admin_only
def catalogos_profesion_add():
    id_prof = (request.form.get("id_profesion") or "").strip()[:100]
    nombre = (request.form.get("profesion") or "").strip()
    if not nombre:
        flash("El nombre de la profesión es obligatorio", "error")
        return redirect(url_for("admin_catalogos"))
    if not id_prof:
        import uuid
        id_prof = str(uuid.uuid4())[:8]
    existing = query("SELECT id_profesion FROM profesion WHERE id_profesion = %s", (id_prof,), one=True)
    if existing:
        flash("Ese ID ya existe; use otro o deje vacío para auto-generar", "error")
        return redirect(url_for("admin_catalogos"))
    execute("INSERT INTO profesion (id_profesion, profesion) VALUES (%s, %s)", (id_prof, nombre))
    flash(f"Profesión '{nombre}' agregada", "success")
    return redirect(url_for("admin_catalogos"))


@app.route("/admin/catalogos/motivo-retiro", methods=["POST"])
@login_required
@module_required("admin")
@role_required("ALL")
@admin_only
def catalogos_motivo_retiro_add():
    tipo_retiro = (request.form.get("tipo_retiro") or "").strip()
    if not tipo_retiro:
        flash("El motivo de retiro es obligatorio", "error")
        return redirect(url_for("admin_catalogos"))
    existing = query("SELECT id FROM motivo_retiro WHERE tipo_retiro = %s", (tipo_retiro,), one=True)
    if existing:
        flash(f"Ya existe el motivo '{tipo_retiro}'", "error")
        return redirect(url_for("admin_catalogos"))
    execute("INSERT INTO motivo_retiro (tipo_retiro) VALUES (%s)", (tipo_retiro,))
    flash(f"Motivo de retiro '{tipo_retiro}' agregado", "success")
    return redirect(url_for("admin_catalogos"))


# ── VIEW TOTAL HIJOS ──────────────────────────────────────────

@app.route("/view-total-hijos")
@login_required
@module_required("total_hijos")
def view_total_hijos():
    rows = query(
        "SELECT e.id_cedula, e.apellidos_nombre, "
        "COUNT(h.id_hijo) AS total_hijos, "
        "SUM(CASE WHEN h.estado = 'ACTIVO' THEN 1 ELSE 0 END) AS activos, "
        "SUM(CASE WHEN h.estado = 'INACTIVO' THEN 1 ELSE 0 END) AS inactivos "
        "FROM empleado e "
        "JOIN hijo h ON e.id_cedula = h.id_cedula "
        "GROUP BY e.id_cedula, e.apellidos_nombre "
        "ORDER BY total_hijos DESC"
    )
    for r in rows:
        r["activos"] = int(r["activos"] or 0)
        r["inactivos"] = int(r["inactivos"] or 0)
        r["total_hijos"] = int(r["total_hijos"] or 0)

    total_activos = sum(r["activos"] for r in rows)
    total_inactivos = sum(r["inactivos"] for r in rows)

    columns = [
        {"key": "id_cedula",        "label": "Cédula"},
        {"key": "apellidos_nombre", "label": "Empleado"},
        {"key": "total_hijos",      "label": "Total Hijos"},
        {"key": "activos",          "label": "Activos"},
        {"key": "inactivos",        "label": "Inactivos"},
    ]
    stats = [
        {"value": len(rows), "label": "Empleados con hijos", "icon": "family_restroom", "color": "blue"},
        {"value": total_activos, "label": "Hijos Activos", "icon": "child_care", "color": "green"},
        {"value": total_inactivos, "label": "Hijos Inactivos", "icon": "child_friendly", "color": "orange"},
    ]
    filter_columns = [
        {"index": 0, "label": "Cédula"},
        {"index": 2, "label": "Total Hijos"},
    ]
    return render_template(
        "data_table.html", active_page="View Total Hijos",
        rows=rows, columns=columns, stats=stats,
        export_key="view_total_hijos", filter_columns=filter_columns,
    )


# ── GENERIC EXPORT ────────────────────────────────────────────

def _parse_export_date(val):
    """Convierte valor de celda a date para filtrar exportación. Acepta date, str YYYY-MM-DD o DD/MM/YYYY."""
    if val is None:
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


EXPORT_CONFIGS = {
    "personal_activo": {
        "sql": "SELECT id_cedula, apellidos_nombre, tipo_documento, departamento, area, sexo, fecha_ingreso, celular, eps, estado FROM empleado WHERE estado = 'ACTIVO' ORDER BY apellidos_nombre",
        "columns": [
            ("id_cedula", "Cédula"), ("apellidos_nombre", "Nombre"), ("tipo_documento", "Tipo Doc"),
            ("departamento", "Departamento"), ("area", "Área"), ("sexo", "Sexo"),
            ("fecha_ingreso", "Fecha Ingreso"), ("celular", "Celular"), ("eps", "EPS"), ("estado", "Estado"),
        ],
        "prefix": "Personal_Activo",
        "date_column": "fecha_ingreso",
    },
    "personal_inactivo": {
        "sql": "SELECT id_cedula, apellidos_nombre, tipo_documento, departamento, area, sexo, fecha_ingreso, celular, eps, estado FROM empleado WHERE estado = 'INACTIVO' ORDER BY apellidos_nombre",
        "columns": [
            ("id_cedula", "Cédula"), ("apellidos_nombre", "Nombre"), ("tipo_documento", "Tipo Doc"),
            ("departamento", "Departamento"), ("area", "Área"), ("sexo", "Sexo"),
            ("fecha_ingreso", "Fecha Ingreso"), ("celular", "Celular"), ("eps", "EPS"), ("estado", "Estado"),
        ],
        "prefix": "Personal_Inactivo",
        "date_column": "fecha_ingreso",
    },
    "hijos_activos": {
        "sql": (
            "SELECT h.id_cedula, e.apellidos_nombre AS nombre_padre, h.apellidos_nombre AS nombre_hijo, "
            "h.identificacion_hijo, h.fecha_nacimiento, h.sexo, h.estado "
            "FROM hijo h LEFT JOIN empleado e ON e.id_cedula = h.id_cedula "
            "WHERE h.estado = 'ACTIVO' ORDER BY e.apellidos_nombre, h.apellidos_nombre"
        ),
        "columns": [
            ("id_cedula", "Cédula Padre"), ("nombre_padre", "Nombre Padre"), ("nombre_hijo", "Nombre Hijo"),
            ("identificacion_hijo", "Identificación"), ("fecha_nacimiento", "Fecha Nacimiento"),
            ("sexo", "Sexo"), ("estado", "Estado"),
        ],
        "prefix": "Hijos_Activos",
        "date_column": "fecha_nacimiento",
    },
    "hijos_inactivos": {
        "sql": (
            "SELECT h.id_cedula, e.apellidos_nombre AS nombre_padre, h.apellidos_nombre AS nombre_hijo, "
            "h.identificacion_hijo, h.fecha_nacimiento, h.sexo, h.estado "
            "FROM hijo h LEFT JOIN empleado e ON e.id_cedula = h.id_cedula "
            "WHERE h.estado = 'INACTIVO' ORDER BY e.apellidos_nombre, h.apellidos_nombre"
        ),
        "columns": [
            ("id_cedula", "Cédula Padre"), ("nombre_padre", "Nombre Padre"), ("nombre_hijo", "Nombre Hijo"),
            ("identificacion_hijo", "Identificación"), ("fecha_nacimiento", "Fecha Nacimiento"),
            ("sexo", "Sexo"), ("estado", "Estado"),
        ],
        "prefix": "Hijos_Inactivos",
        "date_column": "fecha_nacimiento",
    },
    "retiro_personal": {
        "sql": "SELECT id_retiro, id_cedula, apellidos_nombre, departamento, area, fecha_ingreso, fecha_retiro, dias_laborados, tipo_retiro FROM retirado ORDER BY fecha_retiro DESC",
        "columns": [
            ("id_cedula", "Cédula"), ("apellidos_nombre", "Nombre"), ("departamento", "Departamento"),
            ("area", "Área"), ("fecha_ingreso", "Ingreso"), ("fecha_retiro", "Retiro"),
            ("dias_laborados", "Días Laborados"), ("tipo_retiro", "Tipo Retiro"),
        ],
        "prefix": "Retiro_Personal",
        "date_column": "fecha_retiro",
    },
    "departamentos": {
        "sql": "SELECT nombre, presupuestados FROM departamento ORDER BY nombre",
        "columns": [("nombre", "Departamento"), ("presupuestados", "Presupuestados")],
        "prefix": "Departamentos",
    },
    "perfil_ocupacional": {
        "sql": (
            "SELECT d.nombre AS departamento, a.nombre AS area, p.perfil_ocupacional, p.presupuestados "
            "FROM perfil_ocupacional p JOIN area a ON p.area_id = a.id "
            "JOIN departamento d ON a.departamento_id = d.id ORDER BY d.nombre, a.nombre"
        ),
        "columns": [
            ("departamento", "Departamento"), ("area", "Área"),
            ("perfil_ocupacional", "Perfil Ocupacional"), ("presupuestados", "Presupuestados"),
        ],
        "prefix": "Perfil_Ocupacional",
    },
    "view_total_hijos": {
        "sql": (
            "SELECT e.id_cedula, e.apellidos_nombre, COUNT(h.id_hijo) AS total_hijos, "
            "SUM(CASE WHEN h.estado='ACTIVO' THEN 1 ELSE 0 END) AS activos, "
            "SUM(CASE WHEN h.estado='INACTIVO' THEN 1 ELSE 0 END) AS inactivos "
            "FROM empleado e JOIN hijo h ON e.id_cedula=h.id_cedula "
            "GROUP BY e.id_cedula, e.apellidos_nombre ORDER BY total_hijos DESC"
        ),
        "columns": [
            ("id_cedula", "Cédula"), ("apellidos_nombre", "Empleado"),
            ("total_hijos", "Total Hijos"), ("activos", "Activos"), ("inactivos", "Inactivos"),
        ],
        "prefix": "Total_Hijos",
    },
}


@app.route("/export/<page_key>")
@login_required
def generic_export(page_key):
    cfg = EXPORT_CONFIGS.get(page_key)
    if not cfg:
        flash("Exportación no disponible", "error")
        return redirect(url_for("home"))
    rows = query(cfg["sql"])
    date_col = cfg.get("date_column")
    desde_str = request.args.get("desde", "").strip()
    hasta_str = request.args.get("hasta", "").strip()
    if date_col and (desde_str or hasta_str):
        desde = _parse_export_date(desde_str) if desde_str else None
        hasta = _parse_export_date(hasta_str) if hasta_str else None
        filtered = []
        for r in rows:
            cell_date = _parse_export_date(r.get(date_col))
            if cell_date is None:
                continue
            if desde and cell_date < desde:
                continue
            if hasta and cell_date > hasta:
                continue
            filtered.append(r)
        rows = filtered
    return export_excel_response_generic(rows, cfg["columns"], cfg["prefix"])


# ── TELEMETRÍA / AUDITORÍA (reportes empresariales) ─────────────

def registrar_audit(accion, modulo=None, detalle=None):
    """Registra una acción en audit_log para telemetría y reportes."""
    try:
        user = get_current_user()
        id_user = (user.get("id_user") if user else None) or ""
        execute(
            "INSERT INTO audit_log (id_user, accion, modulo, detalle) VALUES (%s, %s, %s, %s)",
            (id_user[:50] if id_user else None, accion[:100], (modulo or "")[:80], (detalle or "")[:500]),
        )
    except Exception:
        pass  # Si no existe la tabla o falla, no romper el flujo


# ── DASHBOARD ─────────────────────────────────────────────────

DASHBOARD_CHARTS = {
    "tipos_retiro": {
        "title": "Tipos de Retiro",
        "query": "SELECT COALESCE(tipo_retiro,'SIN ASIGNAR') AS label, COUNT(*) AS cnt FROM retirado GROUP BY label ORDER BY cnt DESC",
        "table": "retirado",
    },
    "eps": {
        "title": "EPS",
        "query": "SELECT COALESCE(eps,'N/A') AS label, COUNT(*) AS cnt FROM empleado GROUP BY label ORDER BY cnt DESC",
        "table": "empleado",
    },
    "fondo_pensiones": {
        "title": "Fondo de Pensiones",
        "query": "SELECT COALESCE(fondo_pensiones,'N/A') AS label, COUNT(*) AS cnt FROM empleado GROUP BY label ORDER BY cnt DESC",
        "table": "empleado",
    },
    "hijos": {
        "title": "Hijos",
        "subtitle": "¿Tiene Hijos?",
        "query": "SELECT COALESCE(hijos,'N/A') AS label, COUNT(*) AS cnt FROM empleado GROUP BY label ORDER BY cnt DESC",
        "table": "empleado",
    },
    "sexo": {
        "title": "Sexo",
        "query": "SELECT COALESCE(sexo,'N/A') AS label, COUNT(*) AS cnt FROM empleado GROUP BY label ORDER BY cnt DESC",
        "table": "empleado",
    },
    "rh": {
        "title": "RH",
        "query": "SELECT COALESCE(rh,'N/A') AS label, COUNT(*) AS cnt FROM empleado GROUP BY label ORDER BY cnt DESC",
        "table": "empleado",
    },
}

RETIRADO_COLUMNS = [
    ("id_cedula", "ID_Cedula"),
    ("id_retiro", "ID_Retiro"),
    ("apellidos_nombre", "Apellidos y Nombres"),
    ("departamento", "Departamento"),
    ("area", "Area"),
    ("id_perfil_ocupacional", "ID_Perfil_Ocupacional"),
    ("fecha_ingreso", "Fecha de Ingreso"),
    ("fecha_retiro", "Fecha de Retiro"),
    ("dias_laborados", "Dias laborados"),
    ("tipo_retiro", "Tipo de Retiro"),
    ("motivo", "Motivo"),
    ("numero_cedula", "Numero de Cedula"),
    ("mes_retiro", "Mes de Retiro"),
    ("ano_retiro", "Año de Retiro"),
    ("contador", "Contador"),
]


def enrich_retirados(rows):
    """Add calculated fields to retirado rows."""
    for r in rows:
        r["numero_cedula"] = r.get("id_cedula", "")
        fr = parse_fecha(r.get("fecha_retiro", ""))
        r["mes_retiro"] = fr.month if fr else ""
        r["ano_retiro"] = fr.year if fr else ""
        r["contador"] = 1
    return rows


def _get_retiros_por_mes():
    """Últimos 12 meses con cantidad de retiros (para gráfica de barras)."""
    from datetime import datetime, date
    try:
        end = date.today()
        rows = query(
            "SELECT YEAR(fecha_retiro) AS y, MONTH(fecha_retiro) AS m, COUNT(*) AS cnt "
            "FROM retirado WHERE fecha_retiro >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH) "
            "GROUP BY y, m ORDER BY y, m",
        )
        by_month = {(r["y"], r["m"]): int(r["cnt"]) for r in rows}
        labels = []
        values = []
        meses_nombres = ("Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic")
        for i in range(11, -1, -1):
            d = date(end.year, end.month, 1)
            for _ in range(i):
                if d.month == 1:
                    d = date(d.year - 1, 12, 1)
                else:
                    d = date(d.year, d.month - 1, 1)
            labels.append(meses_nombres[d.month - 1] + " " + str(d.year))
            values.append(by_month.get((d.year, d.month), 0))
        labels.reverse()
        values.reverse()
        return labels, values
    except Exception:
        return [], []


def _get_actividad_reciente(limite=10):
    """Últimas acciones registradas en audit_log (telemetría)."""
    try:
        rows = query(
            "SELECT id_user, accion, modulo, detalle, fecha_hora FROM audit_log "
            "ORDER BY fecha_hora DESC LIMIT %s",
            (limite,),
        )
        for r in rows or []:
            fh = r.get("fecha_hora")
            if hasattr(fh, "strftime"):
                r["fecha_str"] = fh.strftime("%d/%m %H:%M")
            else:
                r["fecha_str"] = str(fh)[:16] if fh else "—"
        return rows or []
    except Exception:
        return []


@app.route("/view-total-personal")
@login_required
@module_required("dashboard")
def view_total_personal():
    charts_data = {}
    for key, cfg in DASHBOARD_CHARTS.items():
        rows = query(cfg["query"])
        labels = [r["label"] for r in rows]
        values = [int(r["cnt"]) for r in rows]
        charts_data[key] = {
            "title": cfg["title"],
            "subtitle": cfg.get("subtitle", cfg["title"]),
            "labels": labels,
            "values": values,
        }

    # KPIs nivel empresarial
    try:
        kpi_activos = query("SELECT COUNT(*) AS c FROM empleado WHERE estado = 'ACTIVO'", one=True)["c"]
    except Exception:
        kpi_activos = 0
    try:
        from datetime import date
        hoy = date.today()
        kpi_retiros_mes = query(
            "SELECT COUNT(*) AS c FROM retirado WHERE YEAR(fecha_retiro)=%s AND MONTH(fecha_retiro)=%s",
            (hoy.year, hoy.month), one=True,
        )["c"]
    except Exception:
        kpi_retiros_mes = 0
    try:
        kpi_solicitudes_pend = query(
            "SELECT COUNT(*) AS c FROM solicitud_permiso WHERE estado = 'PENDIENTE'",
            one=True,
        )["c"]
    except Exception:
        kpi_solicitudes_pend = 0
    try:
        from datetime import date
        hoy = date.today()
        kpi_resueltas_hoy = query(
            "SELECT COUNT(*) AS c FROM solicitud_permiso WHERE DATE(fecha_resolucion) = %s AND estado IN ('APROBADO','RECHAZADO')",
            (hoy,), one=True,
        )["c"]
    except Exception:
        kpi_resueltas_hoy = 0

    retiros_mes_labels, retiros_mes_values = _get_retiros_por_mes()
    actividad_reciente = _get_actividad_reciente(10)
    from datetime import datetime
    ultima_actualizacion = datetime.now().strftime("%d/%m/%Y %H:%M")

    return render_template(
        "dashboard.html", active_page="Dashboard",
        charts_data=charts_data,
        kpi_activos=kpi_activos,
        kpi_retiros_mes=kpi_retiros_mes,
        kpi_solicitudes_pend=kpi_solicitudes_pend,
        kpi_resueltas_hoy=kpi_resueltas_hoy,
        retiros_mes_labels=retiros_mes_labels,
        retiros_mes_values=retiros_mes_values,
        actividad_reciente=actividad_reciente,
        ultima_actualizacion=ultima_actualizacion,
    )


@app.route("/dashboard/<chart_key>")
@login_required
@module_required("dashboard")
def dashboard_chart(chart_key):
    cfg = DASHBOARD_CHARTS.get(chart_key)
    if not cfg:
        flash("Gráfico no encontrado", "error")
        return redirect(url_for("view_total_personal"))
    rows = query(cfg["query"])
    labels = [r["label"] for r in rows]
    values = [int(r["cnt"]) for r in rows]
    return render_template(
        "dashboard_chart.html", active_page="Dashboard",
        chart_key=chart_key, title=cfg["title"],
        subtitle=cfg.get("subtitle", cfg["title"]),
        labels=labels, values=values,
        table_name=cfg["table"],
    )


@app.route("/dashboard/<chart_key>/data")
@login_required
@module_required("dashboard")
def dashboard_data(chart_key):
    cfg = DASHBOARD_CHARTS.get(chart_key)
    if not cfg:
        flash("Gráfico no encontrado", "error")
        return redirect(url_for("view_total_personal"))
    filter_val = request.args.get("filter", "")
    if cfg["table"] == "retirado":
        sql = "SELECT * FROM retirado"
        params = ()
        if filter_val:
            sql += " WHERE tipo_retiro = %s"
            params = (filter_val,)
        sql += " ORDER BY apellidos_nombre"
        rows = query(sql, params)
        rows = enrich_retirados(rows)
        cols = RETIRADO_COLUMNS
        table_title = "Retirados_Sli"
    else:
        field_map = {
            "eps": "eps", "fondo_pensiones": "fondo_pensiones",
            "hijos": "hijos", "sexo": "sexo", "rh": "rh",
        }
        field = field_map.get(chart_key, chart_key)
        sql = "SELECT * FROM empleado"
        params = ()
        if filter_val:
            sql += f" WHERE {field} = %s"
            params = (filter_val,)
        sql += " ORDER BY apellidos_nombre"
        rows = query(sql, params)
        rows = enrich_empleados(rows)
        cols = INLINE_COLUMNS
        table_title = "DBase_Sli"
    return render_template(
        "dashboard_data.html", active_page="Dashboard",
        chart_key=chart_key, chart_title=cfg["title"],
        table_title=table_title,
        rows=rows, columns=cols,
        filter_val=filter_val,
    )


@app.route("/dashboard/<chart_key>/export")
@login_required
@module_required("dashboard")
def dashboard_export(chart_key):
    cfg = DASHBOARD_CHARTS.get(chart_key)
    if not cfg:
        flash("Gráfico no encontrado", "error")
        return redirect(url_for("view_total_personal"))
    filter_val = request.args.get("filter", "")
    if cfg["table"] == "retirado":
        sql = "SELECT * FROM retirado"
        params = ()
        if filter_val:
            sql += " WHERE tipo_retiro = %s"
            params = (filter_val,)
        sql += " ORDER BY apellidos_nombre"
        rows = query(sql, params)
        rows = enrich_retirados(rows)
        cols = RETIRADO_COLUMNS
    else:
        field_map = {
            "eps": "eps", "fondo_pensiones": "fondo_pensiones",
            "hijos": "hijos", "sexo": "sexo", "rh": "rh",
        }
        field = field_map.get(chart_key, chart_key)
        sql = "SELECT * FROM empleado"
        params = ()
        if filter_val:
            sql += f" WHERE {field} = %s"
            params = (filter_val,)
        sql += " ORDER BY apellidos_nombre"
        rows = query(sql, params)
        rows = enrich_empleados(rows)
        cols = INLINE_COLUMNS
    return export_excel_response_generic(rows, cols, f"Dashboard_{cfg['title']}")


# ── HOME SETTING (solo ADMIN) ────────────────────────────────

@app.route("/home-setting")
@login_required
@module_required("admin")
@role_required("ALL")
@admin_only
def home_setting():
    rows = query("SELECT * FROM menu ORDER BY nombre")
    columns = [
        {"key": "id_menu", "label": "ID"},
        {"key": "nombre",  "label": "Nombre"},
        {"key": "vista",   "label": "Vista"},
        {"key": "imagen",  "label": "Imagen"},
    ]
    stats = [
        {"value": len(rows), "label": "Total Menús", "icon": "menu", "color": "green"},
    ]
    return render_template(
        "data_table.html", active_page="Home Setting",
        rows=rows, columns=columns, stats=stats,
    )


# ── ABOUT ─────────────────────────────────────────────────────

@app.route("/about")
@login_required
def about():
    return render_template("about.html", active_page="About")


# ── TEST DB ───────────────────────────────────────────────────

@app.route("/test-db")
def test_db():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchone()
        cursor.close()
        conn.close()
        return {"status": "ok", "message": "Conexión a MySQL exitosa"}
    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


# ── RUN ───────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
