# -*- coding: utf-8 -*-
"""
Carga incidencias (INATEL) desde el Excel CARACTERIZACION AT -2026.xlsx
en la tabla incidencia_at. Para el gestor SISO.

Uso (desde la raíz del proyecto):
  python scripts/cargar_incidencias_at_desde_excel.py "ruta/a/CARACTERIZACION AT -2026.xlsx"
"""
import os
import sys
from datetime import datetime, date, time

# Cargar config del proyecto
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

import mysql.connector
from config import Config

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl")
    sys.exit(1)


def _val(v):
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    if isinstance(v, time):
        return v.strftime("%H:%M") if v else None
    s = str(v).strip()
    return s if s else None


def _int_val(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


# Meses en español (abreviados o completos) para parsear "1 de ago. de 2025"
_MESES = {"ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3,
          "abr": 4, "abril": 4, "may": 5, "mayo": 5, "jun": 6, "junio": 6,
          "jul": 7, "julio": 7, "ago": 8, "agosto": 8, "sep": 9, "sept": 9, "septiembre": 9,
          "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11, "dic": 12, "diciembre": 12}


def _parse_fecha_texto(v):
    """Convierte '1 de ago. de 2025' o '22 de dic. de 2025' a date."""
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    if not s:
        return None
    import re
    m = re.match(r"(\d{1,2})\s+de\s+(\w+)\.?\s+de\s+(\d{4})", s, re.I)
    if not m:
        return None
    dia, mes_str, año = int(m.group(1)), m.group(2).lower()[:4], int(m.group(3))
    mes = None
    for k, val in _MESES.items():
        if mes_str.startswith(k) or k.startswith(mes_str):
            mes = val
            break
    if mes is None:
        return None
    try:
        from datetime import date as date_type
        return date_type(año, mes, min(dia, 28))
    except (ValueError, TypeError):
        return None


def main():
    if len(sys.argv) < 2:
        print("Uso: python scripts/cargar_incidencias_at_desde_excel.py <ruta Excel>")
        print('Ejemplo: python scripts/cargar_incidencias_at_desde_excel.py "C:\\Users\\...\\CARACTERIZACION AT -2026.xlsx"')
        sys.exit(1)
    excel_path = sys.argv[1]
    if not os.path.isfile(excel_path):
        print("No se encontró el archivo:", excel_path)
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "ACCIDENTES 2026" not in wb.sheetnames:
        print('La hoja "ACCIDENTES 2026" no existe en el Excel.')
        wb.close()
        sys.exit(1)
    ws = wb["ACCIDENTES 2026"]
    # Fila 4 = encabezados (1-based), datos desde fila 5
    rows_data = list(ws.iter_rows(min_row=5, values_only=True))
    wb.close()

    conn = mysql.connector.connect(
        host=Config.MYSQL_HOST,
        port=Config.MYSQL_PORT,
        user=Config.MYSQL_USER,
        password=Config.MYSQL_PASSWORD,
        database=Config.MYSQL_DATABASE,
    )
    cursor = conn.cursor()

    insert_sql = """
    INSERT INTO incidencia_at (
        numero_registro, mes, fecha_accidente, dia_semana, hora_ocurrencia, tipo_evento,
        nombre_trabajador, cedula, genero, cargo, fecha_ingreso, antiguedad_meses,
        area_seccion_ocurrencia, tipo_vinculacion, dias_incapacidad, prorroga,
        parte_cuerpo_afectada, tipo_lesion, forma_accidente, clasificacion_origen,
        agente_lesion, reincidente, descripcion_accidente, investigado, causas,
        seguimiento_clinico
    ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
    )
    """
    n = 0
    for row in rows_data:
        if not row or row[0] is None:
            continue
        # Columnas según el Excel: 0 No, 1 MES, 2 FECHA, 3 DÍA SEMANA, 4 HORA, 5 TIPO EVENTO, 6 NOMBRE, 7 CEDULA, 8 GENERO, 9 CARGO, 10 F. INGRESO, 11 ANTIG MESES, 12 ÁREA, 13 VINCULACIÓN, 14 DIAS INC, 15 PRÓRROGA, 16 PARTE CUERPO, 17 TIPO LESION, 18 FORMA, 19 CLASIF ORIGEN, 20 AGENTE, 21 REINCIDENTE, 22 DESCRIPCIÓN, 23 INVESTIGADO, 24 CAUSAS, 25 SEGUIMIENTO
        def c(i, default=None):
            return row[i] if i < len(row) else default
        fecha_acc = c(2)
        if hasattr(fecha_acc, "date"):
            fecha_acc = fecha_acc.date()
        fecha_ing = c(10)
        if hasattr(fecha_ing, "date"):
            fecha_ing = fecha_ing.date()
        elif fecha_ing is not None and not isinstance(fecha_ing, (date, datetime)):
            fecha_ing = _parse_fecha_texto(fecha_ing)
        hora = c(4)
        if hasattr(hora, "strftime"):
            hora = hora.strftime("%H:%M")
        vals = (
            _int_val(c(0)),
            _val(c(1)),
            fecha_acc,
            _val(c(3)),
            _val(hora) or _val(c(4)),
            _val(c(5)),
            _val(c(6)),
            str(c(7)) if c(7) is not None else None,
            _val(c(8)),
            _val(c(9)),
            fecha_ing,
            _int_val(c(11)),
            _val(c(12)),
            _val(c(13)),
            _int_val(c(14)),
            _val(c(15)),
            _val(c(16)),
            _val(c(17)),
            _val(c(18)),
            _val(c(19)),
            _val(c(20)),
            _val(c(21)),
            _val(c(22)),
            _val(c(23)),
            _val(c(24)),
            _val(c(25)),
        )
        try:
            cursor.execute(insert_sql, vals)
            n += 1
        except Exception as e:
            print("Error en fila:", row[:8], ":", e)
    conn.commit()
    cursor.close()
    conn.close()
    print("Listo. Se cargaron", n, "incidencias en incidencia_at.")


if __name__ == "__main__":
    main()
