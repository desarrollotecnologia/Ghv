"""
Genera un SQL global para corregir fechas de texto desde la fuente real.

No intenta adivinar fechas ambiguas como 10/11/1976. En su lugar copia las
fechas desde el Excel maestro:

    BDATOS_XLSX_PATH en .env, o --xlsx "ruta\\BDatos_APPGH.xlsx"

Corrige:
- empleado.fecha_expedicion
- empleado.fecha_ingreso
- empleado.fecha_nacimiento
- hijo.fecha_nacimiento
- retirado.fecha_ingreso
- retirado.fecha_retiro

Opcionalmente, si se entrega --cumpleanos-xlsx, esa fuente tiene prioridad para
empleado.fecha_nacimiento porque es el archivo revisado de cumpleaños.
"""
import argparse
import os
import re
import unicodedata
from datetime import date, datetime, timedelta

import openpyxl
from dotenv import load_dotenv


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(ROOT, ".env"))

DEFAULT_XLSX = (os.getenv("BDATOS_XLSX_PATH") or "").strip()
DEFAULT_OUT = os.path.join(ROOT, "database", "actualizar_fechas_global.sql")
DEFAULT_CUMPLEANOS = r"C:\Users\USUARIO\Downloads\CUMPLEAÑOS 2026.xlsx"


def esc(value):
    return str(value).replace("\\", "\\\\").replace("'", "''")


def strip_accents(value):
    text = str(value or "")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm(value):
    text = strip_accents(value).upper()
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return " ".join(text.split())


def clean_id(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        try:
            return str(int(value))
        except Exception:
            return str(value).strip()
    text = str(value).strip()
    try:
        return str(int(float(text)))
    except (ValueError, TypeError):
        return text


def clean_date_str(value):
    """Convierte Excel/date/string a DD/MM/YYYY sin aceptar MM/DD/YYYY."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float)):
        try:
            if float(value) > 20000:
                return (date(1899, 12, 30) + timedelta(days=int(value))).strftime("%d/%m/%Y")
        except Exception:
            return None
    text = str(value).strip()
    if not text or text.lower() in ("none", "nan"):
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None


def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm(h) if h else f"COL_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    return data


def get(row, *names):
    for name in names:
        key = norm(name)
        if key in row:
            return row.get(key)
    return None


def leer_cumpleanos_override(path):
    """Retorna {nombre_norm: fecha_ddmmyyyy} desde CUMPLEAÑOS 2026.xlsx."""
    if not path or not os.path.isfile(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result = {}
    for ws in wb.worksheets:
        header = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            normalized = [norm(v) for v in row]
            if "APELLIDOS Y NOMBRES" in normalized and "FECHA DE NACIMIENTO" in normalized:
                header = {name: pos for pos, name in enumerate(normalized)}
                start = idx + 1
                break
        if not header:
            continue
        for row in ws.iter_rows(min_row=start, values_only=True):
            nombre = row[header["APELLIDOS Y NOMBRES"]]
            fecha = row[header["FECHA DE NACIMIENTO"]]
            nombre_key = norm(nombre)
            fecha_fmt = clean_date_str(fecha)
            if nombre_key and nombre_key != "APELLIDOS Y NOMBRES" and fecha_fmt:
                result.setdefault(nombre_key, fecha_fmt)
    wb.close()
    return result


def update_sql(table, assignments, where):
    set_sql = ", ".join(f"{col} = {val}" for col, val in assignments)
    return f"UPDATE {table} SET {set_sql} WHERE {where};"


def sql_value(value):
    return "NULL" if value is None else f"'{esc(value)}'"


def generar(args):
    if not args.xlsx or not os.path.isfile(args.xlsx):
        raise SystemExit(
            "No existe el Excel maestro. Configure BDATOS_XLSX_PATH en .env "
            'o use --xlsx "ruta\\BDatos_APPGH.xlsx".'
        )

    cumpleanos = leer_cumpleanos_override(args.cumpleanos_xlsx)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)

    statements = []
    stats = {"empleado": 0, "hijo": 0, "retirado": 0, "sin_fecha": 0}

    for row in read_sheet(wb, "DBase"):
        cedula = clean_id(get(row, "ID_Cedula"))
        if not cedula:
            continue
        nombre = get(row, "Apellidos_Nombre")
        fecha_nacimiento = clean_date_str(get(row, "Fecha_Nacimiento"))
        if norm(nombre) in cumpleanos:
            fecha_nacimiento = cumpleanos[norm(nombre)]

        values = {
            "fecha_expedicion": clean_date_str(get(row, "Fecha_Expedicion")),
            "fecha_ingreso": clean_date_str(get(row, "Fecha_Ingreso")),
            "fecha_nacimiento": fecha_nacimiento,
        }
        assignments = [(col, sql_value(val)) for col, val in values.items() if val]
        if not assignments:
            stats["sin_fecha"] += 1
            continue
        statements.append((
            f"-- empleado {cedula} | {nombre or ''}",
            update_sql("empleado", assignments, f"id_cedula = '{esc(cedula)}'"),
        ))
        stats["empleado"] += 1

    for row in read_sheet(wb, "Hijos"):
        hijo_id = clean_id(get(row, "ID_Hijo"))
        cedula = clean_id(get(row, "ID_Cedula"))
        fecha = clean_date_str(get(row, "Fecha_Nacimiento"))
        if not hijo_id or not fecha:
            stats["sin_fecha"] += 1
            continue
        where = f"id_hijo = '{esc(hijo_id)}'"
        if cedula:
            where += f" AND id_cedula = '{esc(cedula)}'"
        statements.append((
            f"-- hijo {hijo_id} | cedula padre {cedula}",
            update_sql("hijo", [("fecha_nacimiento", sql_value(fecha))], where),
        ))
        stats["hijo"] += 1

    for row in read_sheet(wb, "Retirados"):
        retiro_id = clean_id(get(row, "ID_Retiro"))
        cedula = clean_id(get(row, "ID_Cedula"))
        values = {
            "fecha_ingreso": clean_date_str(get(row, "Fecha_Ingreso")),
            "fecha_retiro": clean_date_str(get(row, "Fecha_Retiro")),
        }
        assignments = [(col, sql_value(val)) for col, val in values.items() if val]
        if not retiro_id or not assignments:
            stats["sin_fecha"] += 1
            continue
        where = f"id_retiro = '{esc(retiro_id)}'"
        if cedula:
            where += f" AND id_cedula = '{esc(cedula)}'"
        statements.append((
            f"-- retirado {retiro_id} | cedula {cedula}",
            update_sql("retirado", assignments, where),
        ))
        stats["retirado"] += 1

    wb.close()

    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write("-- Correccion global de fechas desde fuente Excel\n")
        fh.write("-- Copia fechas reales; no intenta adivinar dia/mes en valores ambiguos.\n")
        fh.write("USE gestio_humana;\nSTART TRANSACTION;\n\n")
        for comment, sql in statements:
            fh.write(f"{comment}\n{sql}\n")
        fh.write("\nCOMMIT;\n")

    return stats, len(statements)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--cumpleanos-xlsx", default=DEFAULT_CUMPLEANOS)
    parser.add_argument("--out", default=DEFAULT_OUT)
    args = parser.parse_args()

    stats, total = generar(args)
    print(f"SQL generado: {args.out}")
    print(f"Updates totales: {total}")
    print(f"Empleado: {stats['empleado']}")
    print(f"Hijo: {stats['hijo']}")
    print(f"Retirado: {stats['retirado']}")
    print(f"Filas sin fecha util: {stats['sin_fecha']}")
    if not os.path.isfile(args.cumpleanos_xlsx):
        print("Aviso: no se encontro Excel de cumpleaños; se uso fecha_nacimiento del Excel maestro.")


if __name__ == "__main__":
    main()
