"""Genera SQL por nombre desde CUMPLEAÑOS 2026.xlsx, sin conectar a MySQL."""
import argparse
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from openpyxl import load_workbook


XLSX = r"C:\Users\USUARIO\Downloads\CUMPLEAÑOS 2026.xlsx"
OUT = r"c:\Users\USUARIO\Documents\app_ghv\database\actualizar_cumpleanos_2026_por_nombre.sql"


def esc(value):
    return str(value).replace("\\", "\\\\").replace("'", "''")


def strip_accents(value):
    text = str(value or "")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm(value):
    text = strip_accents(value).upper()
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return " ".join(text.split())


def fdate(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float)):
        try:
            return (date(1899, 12, 30) + timedelta(days=int(value))).strftime("%d/%m/%Y")
        except Exception:
            return None
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=XLSX)
    parser.add_argument("--out", default=OUT)
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        raise SystemExit(f"No existe el Excel: {args.xlsx}")

    wb = load_workbook(args.xlsx, data_only=True, read_only=True)
    rows = {}
    skipped = 0
    for ws in wb.worksheets:
        header = None
        start = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [norm(x) for x in row]
            if "APELLIDOS Y NOMBRES" in vals and "FECHA DE NACIMIENTO" in vals:
                header = {name: idx for idx, name in enumerate(vals)}
                start = i + 1
                break
        if not header:
            continue
        for row in ws.iter_rows(min_row=start, values_only=True):
            nombre = row[header["APELLIDOS Y NOMBRES"]]
            fecha = row[header["FECHA DE NACIMIENTO"]]
            area = row[header.get("AREA", -1)] if "AREA" in header else None
            nombre_norm = norm(nombre)
            fecha_fmt = fdate(fecha)
            if not nombre_norm or nombre_norm == "APELLIDOS Y NOMBRES" or not fecha_fmt:
                skipped += 1
                continue
            key = (nombre_norm, norm(area))
            rows.setdefault(key, (str(nombre).strip(), str(area or "").strip(), fecha_fmt, ws.title))
    wb.close()

    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write("-- Actualiza fecha_nacimiento usando CUMPLEAÑOS 2026.xlsx como fuente correcta\n")
        fh.write("-- Ejecutar en MySQL Workbench sobre gestio_humana. No adivina dia/mes: copia la fecha del Excel.\n")
        fh.write("USE gestio_humana;\nSTART TRANSACTION;\n\n")
        for nombre, area, fecha, hoja in rows.values():
            fh.write(f"-- {hoja}: {esc(nombre)} / {esc(area)} -> {fecha}\n")
            area_clause = ""
            if area:
                area_clause = f" AND UPPER(TRIM(area)) = UPPER('{esc(area)}')"
            fh.write(
                "UPDATE empleado SET fecha_nacimiento = "
                f"'{esc(fecha)}' WHERE estado = 'ACTIVO' "
                f"AND UPPER(TRIM(apellidos_nombre)) = UPPER('{esc(nombre)}')"
                f"{area_clause};\n"
            )
        fh.write("\nCOMMIT;\n")
    print(f"Generado {args.out} con {len(rows)} UPDATEs. Filas omitidas: {skipped}.")


if __name__ == "__main__":
    main()
