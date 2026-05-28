"""
Sincroniza empleado.fecha_nacimiento desde el Excel revisado de cumpleaños.

El Excel CUMPLEAÑOS 2026.xlsx es la fuente correcta para cumpleaños.
Este script actualiza solo fecha_nacimiento en empleado, sin tocar otros datos.

Uso:
    python database/sincronizar_cumpleanos_2026.py --xlsx "C:\\Users\\USUARIO\\Downloads\\CUMPLEAÑOS 2026.xlsx"
    python database/sincronizar_cumpleanos_2026.py --apply
    python database/sincronizar_cumpleanos_2026.py --write-sql database/actualizar_cumpleanos_2026.sql
"""
import argparse
import os
import re
import sys
import unicodedata
from datetime import date, datetime

import mysql.connector
import openpyxl
from dotenv import load_dotenv


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(ROOT, ".env"))

DEFAULT_XLSX = r"C:\Users\USUARIO\Downloads\CUMPLEAÑOS 2026.xlsx"

DB_CONFIG = {
    "host": os.getenv("MYSQL_HOST", "localhost"),
    "port": int(os.getenv("MYSQL_PORT", 3306)),
    "user": os.getenv("MYSQL_USER", "gh_admin"),
    "password": os.getenv("MYSQL_PASSWORD", ""),
    "database": os.getenv("MYSQL_DATABASE", "gestio_humana"),
}


def strip_accents(value):
    text = str(value or "")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm(value):
    text = strip_accents(value).upper()
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return " ".join(text.split())


def fecha_ddmmyyyy(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return None


def leer_cumpleanos(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    registros = []
    for ws in wb.worksheets:
        header_row = None
        headers = {}
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            normalized = [norm(v) for v in row]
            if "APELLIDOS Y NOMBRES" in normalized and "FECHA DE NACIMIENTO" in normalized:
                header_row = idx
                headers = {name: pos for pos, name in enumerate(normalized)}
                break
        if not header_row:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            nombre = row[headers["APELLIDOS Y NOMBRES"]]
            fecha = row[headers["FECHA DE NACIMIENTO"]]
            if not nombre or not fecha:
                continue
            registros.append({
                "nombre": str(nombre).strip(),
                "nombre_norm": norm(nombre),
                "area": str(row[headers.get("AREA", -1)] or "").strip() if "AREA" in headers else "",
                "fecha": fecha_ddmmyyyy(fecha),
                "hoja": ws.title,
            })
    wb.close()
    # Si un nombre aparece dos veces, conservar el primero y reportar visualmente.
    dedup = {}
    duplicados = []
    for r in registros:
        if r["nombre_norm"] in dedup and dedup[r["nombre_norm"]]["fecha"] != r["fecha"]:
            duplicados.append((dedup[r["nombre_norm"]], r))
        dedup.setdefault(r["nombre_norm"], r)
    return list(dedup.values()), duplicados


def sql_escape(value):
    return str(value).replace("\\", "\\\\").replace("'", "''")


def main():
    parser = argparse.ArgumentParser(description="Corrige fecha_nacimiento desde CUMPLEAÑOS 2026.xlsx")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--apply", action="store_true", help="Aplicar cambios en BD. Sin esto solo muestra vista previa.")
    parser.add_argument("--write-sql", default=None, help="Genera archivo SQL con los UPDATE.")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"ERROR: no existe Excel: {args.xlsx}")
        sys.exit(1)

    excel_rows, duplicados = leer_cumpleanos(args.xlsx)
    print(f"Excel: {len(excel_rows)} registros de cumpleaños leidos.")
    if duplicados:
        print(f"AVISO: {len(duplicados)} nombres duplicados con fechas distintas; revisar manualmente.")

    conn = mysql.connector.connect(**DB_CONFIG)
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id_cedula, apellidos_nombre, area, fecha_nacimiento FROM empleado WHERE estado = 'ACTIVO'")
    empleados = cur.fetchall()

    by_name = {}
    for emp in empleados:
        by_name.setdefault(norm(emp["apellidos_nombre"]), []).append(emp)

    cambios = []
    no_encontrados = []
    ambiguos = []
    for row in excel_rows:
        matches = by_name.get(row["nombre_norm"], [])
        if not matches:
            no_encontrados.append(row)
            continue
        if len(matches) > 1:
            area_norm = norm(row["area"])
            area_matches = [m for m in matches if norm(m.get("area")) == area_norm]
            if len(area_matches) == 1:
                matches = area_matches
            else:
                ambiguos.append((row, matches))
                continue
        emp = matches[0]
        actual = fecha_ddmmyyyy(emp.get("fecha_nacimiento")) or (emp.get("fecha_nacimiento") or "")
        correcta = row["fecha"]
        if correcta and actual != correcta:
            cambios.append((emp, row, actual, correcta))

    print(f"Cambios detectados: {len(cambios)}")
    print(f"No encontrados en BD: {len(no_encontrados)}")
    print(f"Ambiguos por nombre: {len(ambiguos)}")

    for emp, row, actual, correcta in cambios[:40]:
        print(f"  {emp['id_cedula']} | {emp['apellidos_nombre']} | {actual} -> {correcta} ({row['hoja']})")
    if len(cambios) > 40:
        print(f"  ... {len(cambios) - 40} cambios mas")

    if args.write_sql:
        with open(args.write_sql, "w", encoding="utf-8") as f:
            f.write("-- Correccion de fecha_nacimiento desde CUMPLEAÑOS 2026.xlsx\n")
            f.write("USE gestio_humana;\nSTART TRANSACTION;\n\n")
            for emp, _row, _actual, correcta in cambios:
                f.write(
                    "UPDATE empleado SET fecha_nacimiento = "
                    f"'{sql_escape(correcta)}' WHERE id_cedula = '{sql_escape(emp['id_cedula'])}';\n"
                )
            f.write("\nCOMMIT;\n")
        print(f"SQL generado: {args.write_sql}")

    if args.apply and cambios:
        for emp, _row, _actual, correcta in cambios:
            cur.execute(
                "UPDATE empleado SET fecha_nacimiento = %s WHERE id_cedula = %s",
                (correcta, emp["id_cedula"]),
            )
        conn.commit()
        print(f"Aplicados {len(cambios)} cambios.")
    elif not args.apply:
        print("Vista previa solamente. Use --apply para actualizar la BD.")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
